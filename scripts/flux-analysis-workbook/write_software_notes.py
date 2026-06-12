"""
Write Apr 2026 variance notes into the Software tab static table (column L)
of the Flux Pivot Template, for vendors with abs variance > $200.

Per the accountant's instruction: investigate only above $200, do not touch existing notes.
"""
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, Alignment

WB_PATH = Path(r"c:/Users/Accountant/Documents/Finance's Requests/Antigravity/Monthly-Accounting/Monthly Flux Analysis/2026/2026-04/Flux Pivot Template 2026-04.xlsx")

# (row, vendor_label, note) - rows pre-sorted by abs var on the Software tab
NOTES = [
    (6,  "Datadog, Inc.",              "Mar carried $6K Feb overage accrual; Apr base only ($17K JE) - overage settled via Apr bill + accrual reversal"),
    (7,  "Figma",                      "Mar had -$9.6K Brex credit/refund; Apr back to normal $2.5K monthly accrual"),
    (8,  "Workhorse HR",                   "Apr contract-minimum + overage fees higher than Mar (Mar had -$501 credit memo)"),
    (9,  "Pagerduty, Inc.",            "Apr added $5K Brex charge (no equivalent Mar charge); Apr-26 SW reclass dept-only net zero"),
    (10, "SEMRUSH",                    "Apr annual seat renewal $4.6K on Brex; Mar only $289 base"),
    (11, "WalkMe",                     "Mar over-accrued $4.2K (carrying Feb catch-up); Apr reversed Mar and booked correct $2.1K monthly"),
    (12, "Google",                     "Higher Apr Google Workspace + ads Brex charges; Apr-26 dept reclass net zero"),
    (13, "Temporal Technologies Inc.", "New vendor in Apr - $1.9K subscription start (JE#####)"),
    (14, "SSO Provider",                 "Fully reclassed from OpEx to COGS via Apr-26 SW Reclass JE##### ($17K moved out)"),
    (15, "OpenAI, LLC",                "Higher Apr OpenAI seat charges on Brex (~$950 per seat across multiple users)"),
    (16, "Cobalt Labs, Inc.",          "Lower Apr Cobalt subscription -$1.8K (contract step-down from Mar)"),
    (17, "MICROSOFT",                  "Higher Apr Microsoft Brex spend (one-time $5.7K license purchase)"),
    (18, "6Sense",                     "Apr campaign fees bill $1K lower than Mar ($1.8K vs $2.8K); base JE unchanged"),
    (19, "CoderPad",                   "Mar Canada Brex reimbursements ($995) not repeating; no Apr activity"),
    (20, "Matik, Inc",                 "Higher Apr Matik subscription $1.6K (Mar was partial $725)"),
    (21, "Dun & the AP Reviewerstreet",           "Apr base subscription $506 vs Mar $1.2K (contract step-down)"),
    (22, "JetBrains",                  "Higher Apr JetBrains license purchase ($431 + $448) vs Mar $177"),
    (23, "Parallels",                  "Mar accrual reversed; no Apr Parallels purchase"),
    (24, "NAME-CHEAP.COM",             "Mar $485 domain registration one-off; Apr only minor $79 renewal"),
    (25, "Chatgpt",                    "Higher Apr ChatGPT Brex usage $1.3K vs Mar $0.9K; Apr-26 dept reclass net zero"),
    (26, "Lodging",                    "Miscoded $430 lodging hitting Software (671100) - should be 661200; consider Apr reclass"),
    (27, "Buoyant, Inc.",              "Mar one-off $334 charge not repeating; no Apr activity"),
    (28, "Mineral, Inc.",              "Mar one-off $284 charge not repeating; no Apr activity"),
    (29, "WEBSHARE",                   "Mar $230 Brex + Q1 reclass cancellations; no Apr activity"),
    (30, "Software Subscription",      "Lower misc Brex software charges -$224 vs Mar (no specific vendor named)"),
    (31, "ZoomInfo Technologies LLC",  "Lower Apr ZoomInfo contract step-down -$202"),
]


def main():
    wb = openpyxl.load_workbook(WB_PATH)
    sw = wb["Software"]
    arial8 = Font(name="Arial", size=8)
    wrap = Alignment(wrap_text=True, vertical="top")

    written = 0
    skipped = 0
    for row, vendor_check, note in NOTES:
        # Verify vendor matches expected (with fragile non-breaking spaces tolerated)
        actual = (sw.cell(row=row, column=6).value or "").strip().replace("\xa0", "")
        if vendor_check.replace("\xa0", "") not in actual and actual not in vendor_check:
            print(f"  WARN R{row}: expected {vendor_check!r}, found {actual!r} - skipping")
            skipped += 1
            continue
        existing = sw.cell(row=row, column=12).value
        if existing and str(existing).strip():
            print(f"  SKIP R{row} ({actual}): already has L={existing!r}")
            skipped += 1
            continue
        for ch in note:
            if ord(ch) > 127:
                raise ValueError(f"Non-ASCII at R{row}: {ch!r}")
        c = sw.cell(row=row, column=12)
        c.value = note
        c.font = arial8
        c.alignment = wrap
        print(f"  R{row:3} {actual[:28]:28} -> {note[:75]}")
        written += 1

    wb.save(WB_PATH)
    print(f"\nWritten: {written}, skipped: {skipped}")
    print(f"Saved: {WB_PATH}")


if __name__ == "__main__":
    main()
