"""
Populate the Flux Pivot Template for a target month from the 4 cached NS reports.

Per SKILL.md flux-analysis-workbook:
- COGS (537): filter to 511400, 511425, 511450, 511510, 511520, 511550, 511600; negate amounts
- Contractors (540): filter to 511370, 611700
- Professional Fees (542): filter to 651100, 651101
- Software (721): filter to 671000, 671100

Master template lives at Monthly Flux Analysis/Flux Pivot Template.xlsx
Output: Monthly Flux Analysis/{YYYY}/{YYYY-MM}/Flux Pivot Template {YYYY-MM}.xlsx
"""
from pathlib import Path
from datetime import datetime
from email.utils import parsedate_to_datetime
import json, shutil, sys
import openpyxl

REPO = Path(r"c:/Users/Accountant/Documents/Finance's Requests/Antigravity/Monthly-Accounting")
MASTER = REPO / "Monthly Flux Analysis/Flux Pivot Template.xlsx"
MONTH = "2026-04"  # default; override via CLI arg `python populate_pivot_template.py 2026-05`
if len(sys.argv) > 1 and sys.argv[1].count("-") == 1:
    MONTH = sys.argv[1]
YEAR, MM = MONTH.split("-")
MONTH_DIR = REPO / f"Monthly Flux Analysis/{YEAR}/{MONTH}"
CACHE = MONTH_DIR / "_cache"
# v2 layout: output goes inside Flux Workbook/ subfolder
OUT = MONTH_DIR / "Flux Workbook" / f"{MONTH} Flux Pivot Template.xlsx"

COGS_FILTERS = {"511400", "511425", "511450", "511510", "511520", "511550", "511600"}
CONTRACTOR_FILTERS = {"511370", "611700"}
PROFFEES_FILTERS = {"651100", "651101"}
SOFTWARE_FILTERS = {"671000", "671100"}

SHEET_MAP = {
    "COGS":          ("IncomeStatementDetailCOGS", "537.json", "cogs"),
    "Contractors":   ("GeneralLedgerdetailContra", "540.json", "gl"),
    "ProfFees":      ("GeneralLedgerdetailProfes", "542.json", "gl"),
    "Software":      ("GeneralLedgerdetailSoftwa", "721.json", "gl"),
}

PERIOD_TEXT = "Feb 2026, Mar 2026, Apr 2026"


def parse_date(s):
    if not s:
        return None
    try:
        return parsedate_to_datetime(s).replace(tzinfo=None)
    except Exception:
        return None


def clean_acct(s):
    return (s or "").replace("\x01", ":")


def derive_final_name(entity, name):
    if entity and entity != "- No Entity -":
        return entity
    return name or ""


def load_report(path):
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def iter_detail_rows(rd):
    """Yield (row_dict, financial_row_context) where row_dict is detailLineValues kvp."""
    keys = sorted(rd.keys(), key=lambda k: int(k))
    fin_row_ctx = ""
    for k in keys:
        row = rd[k]
        if not row.get("isDetailLine"):
            v = row.get("value") or ""
            if v and ("-" in v) and any(c.isdigit() for c in v[:8]):
                # Looks like an account header e.g. "511400 - COGS - Hosting"
                fin_row_ctx = v
            continue
        dlv = row.get("detailLineValues") or []
        if not dlv:
            continue
        kvp = {}
        for d in dlv:
            if isinstance(d, dict):
                # may have one or more {columnName: value} pairs
                for kk, vv in d.items():
                    kvp[kk] = vv
        yield kvp, fin_row_ctx


def parse_cogs(rd):
    """Parse Report {NS_REPORT_ID} (COGS IS Detail). Returns list of 15-col rows for IncomeStatementDetailCOGS."""
    out = []
    for kvp, fin_ctx in iter_detail_rows(rd):
        typ = kvp.get("Type")
        if not typ:
            continue
        # Financial Row: from Account (Line): Name (GL-style), take last segment
        acct_path_raw = kvp.get("Account (Line): Name (GL-style)") or kvp.get("Account (Line): Name", "")
        acct_path = clean_acct(acct_path_raw)
        fin_row = acct_path.split(":")[-1].strip() if acct_path else fin_ctx
        # COGS filter: financial row must contain one of the target sub-accounts
        if not any(a in fin_row for a in COGS_FILTERS):
            continue
        amt_raw = kvp.get("Amount", 0) or 0
        try:
            amt = -float(amt_raw)  # negate
        except (TypeError, ValueError):
            amt = 0
        date = parse_date(kvp.get("Date"))
        name = kvp.get("Name") or ""
        entity = kvp.get("Entity (Line)") or ""
        final_name = derive_final_name(entity, name)
        row = [
            fin_row,
            typ,
            date,
            kvp.get("Document Number") or "",
            name,
            entity,
            final_name,
            kvp.get("Clr") or "",
            kvp.get("Split") or "",
            amt,
            kvp.get("Memo") or "",
            kvp.get("Message") or "",
            acct_path,
            kvp.get("Department: Name") or "",
            kvp.get("Accounting Period: Name") or "",
        ]
        out.append(row)
    return out


# Entity names that, if present on a row tagged for a Software / Contractors /
# ProfFees account, indicate a NetSuite report quirk where a multi-line Brex JE
# attributes all rows to the same account on the report even though the actual
# posting account differs. These are categorical T&E "vendor" labels used by
# Brex (not real vendor names). When seen alongside one of our target accounts,
# the row is almost certainly miscategorized and should be dropped.
_BREX_CATEGORY_DENYLIST = {
    "lodging",
    "airfare",
    "meals",
    "transportation",
    "taxi",
    "hotel",
    "uber",
}


def parse_gl_detail(rd, filters, warnings: list | None = None):
    """
    Parse Reports {NS_REPORT_ID}/542/721 (GL Detail). Returns list of 15-col rows.

    Defends against NetSuite's multi-line Brex-export quirk where a Brex JE
    with lines hitting multiple accounts gets ALL its lines stamped with the
    same "Account" in the report output. Rows whose Entity name is a known
    categorical T&E label (Lodging, Airfare, etc.) are dropped from the data
    sheet AND logged to the `warnings` list for the caller to surface to the
    user / tie-back.
    """
    out = []
    for kvp, _ in iter_detail_rows(rd):
        typ = kvp.get("Type")
        if not typ:
            continue
        acct_path_raw = kvp.get("Account") or ""
        acct_path = clean_acct(acct_path_raw)
        if not any(a in acct_path for a in filters):
            continue

        # Defensive filter: reject rows whose entity label is a Brex T&E
        # category — these are NS report attribution errors, not real activity
        # on this account.
        entity_raw = kvp.get("Entity (Line): Name") or ""
        entity_lower = entity_raw.strip().lower()
        if entity_lower in _BREX_CATEGORY_DENYLIST:
            if warnings is not None:
                net_for_log = kvp.get("_total") if kvp.get("_total") is not None else kvp.get("") or 0
                warnings.append(
                    {
                        "kind": "brex_category_misattribution",
                        "account_path": acct_path,
                        "entity": entity_raw,
                        "doc": kvp.get("Document Number"),
                        "date": kvp.get("Date"),
                        "amount": net_for_log,
                        "memo_first_line": (kvp.get("Memo") or "").split("\n", 1)[0],
                    }
                )
            continue

        # The net amount sits in the "_total" column (unnamed in report header but keyed as _total)
        net_raw = kvp.get("_total")
        if net_raw is None:
            net_raw = kvp.get("") or 0
        try:
            net = float(net_raw)
        except (TypeError, ValueError):
            net = 0
        debit = net if net > 0 else 0
        credit = abs(net) if net < 0 else 0
        date = parse_date(kvp.get("Date"))
        name = kvp.get("Name") or ""
        entity = entity_raw
        final_name = derive_final_name(entity, name)
        row = [
            acct_path,
            typ,
            date,
            kvp.get("Document Number") or "",
            name,
            entity,
            final_name,
            debit,
            credit,
            net,
            kvp.get("Memo") or "",
            kvp.get("Message") or "",
            kvp.get("Department: Name") or "",
            kvp.get("Subsidiary: Name") or "",
            kvp.get("Accounting Period: Name") or "",
        ]
        out.append(row)
    return out


# ============================================================
# Tie-back verification
# ============================================================

# Maps detail-tab key -> (IS row labels to sum) for the IS Drop tie-back.
TIEBACK_IS_ROWS = {
    "COGS": ("Total - 511000 - Cost of Goods Sold",),
    "Contractors": (
        "511370 - COGS - Contractor Payroll",
        "611700 - Contractor Payroll",
    ),
    "ProfFees": ("651100 - Professional Fees",),
    "Software": ("671100 - Software Subscriptions",),
}


def read_is_line_totals(is_drop_path, periods):
    """
    Read the IS Drop xlsx, return {row_label: {period: value}} for the rows
    we care about. Returns None if the file doesn't exist or isn't readable.
    """
    p = Path(is_drop_path)
    if not p.exists():
        return None
    try:
        wb = openpyxl.load_workbook(p, data_only=True)
        # Assume the first sheet is the IS (or a sheet named "Income Statement")
        sheet_name = "Income Statement" if "Income Statement" in wb.sheetnames else wb.sheetnames[0]
        ws = wb[sheet_name]
    except Exception:
        return None

    # Identify the period column for each desired period by scanning the header row.
    # IS has period labels in row 7 (per template-spec.md).
    period_cols = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(7, c).value
        if v in periods:
            period_cols[v] = c

    out = {}
    # Find the row that contains each label in column A
    target_labels = {lbl for labels in TIEBACK_IS_ROWS.values() for lbl in labels}
    for r in range(1, ws.max_row + 1):
        label = ws.cell(r, 1).value
        if label in target_labels:
            row_vals = {}
            for period, col in period_cols.items():
                val = ws.cell(r, col).value
                try:
                    row_vals[period] = float(val) if val is not None else 0.0
                except (TypeError, ValueError):
                    row_vals[period] = 0.0
            out[label] = row_vals
    return out


def tieback(parsed_rows_by_tab, dropped_amounts_by_tab, periods, is_drop_path,
            tolerance: float = 0.50):
    """
    Compare parsed detail sums against authoritative source.

    parsed_rows_by_tab: {tab_key: list-of-15-col-rows}
    dropped_amounts_by_tab: {tab_key: {period: sum_dropped}} — Brex-category
        rows that were filtered out by parse_gl_detail. Added back so the
        comparison matches the IS xlsx (which includes the NS bug).
    periods: list of period names, e.g. ["Feb 2026", "Mar 2026", "Apr 2026"]
    is_drop_path: path to Income Statement Drop xlsx, or None to skip primary.
    tolerance: $ tolerance per period.

    Returns: {"pass": bool, "diffs": [...], "source": "is_drop"|"missing", ...}
    """
    is_totals = read_is_line_totals(is_drop_path, periods) if is_drop_path else None

    diffs = []
    overall_pass = True

    for tab_key, rows in parsed_rows_by_tab.items():
        # Sum parsed rows per period
        parsed_sum = {p: 0.0 for p in periods}
        for row in rows:
            period = row[14] or ""
            if period in parsed_sum:
                try:
                    parsed_sum[period] += float(row[9] or 0)
                except (TypeError, ValueError):
                    pass
        # Add back dropped (Brex-category) amounts so comparison matches IS bug
        dropped_per_period = dropped_amounts_by_tab.get(tab_key, {})
        comparable_sum = {p: parsed_sum[p] + dropped_per_period.get(p, 0.0) for p in periods}

        if is_totals is None:
            diffs.append({
                "tab": tab_key,
                "status": "skipped",
                "note": "No IS Drop xlsx available; SuiteQL fallback not yet wired in this call.",
                "parsed_sum": parsed_sum,
            })
            continue

        # Expected sum = sum of IS rows for this tab
        expected = {p: 0.0 for p in periods}
        for label in TIEBACK_IS_ROWS.get(tab_key, ()):
            row_vals = is_totals.get(label) or {}
            for p in periods:
                expected[p] += row_vals.get(p, 0.0)

        # Compare
        tab_pass = True
        per_period = []
        for p in periods:
            diff = comparable_sum[p] - expected[p]
            ok = abs(diff) <= tolerance
            tab_pass &= ok
            per_period.append({
                "period": p,
                "parsed": parsed_sum[p],
                "dropped_added_back": dropped_per_period.get(p, 0.0),
                "comparable": comparable_sum[p],
                "expected_from_is": expected[p],
                "diff": diff,
                "pass": ok,
            })

        overall_pass &= tab_pass
        diffs.append({
            "tab": tab_key,
            "status": "pass" if tab_pass else "fail",
            "per_period": per_period,
        })

    return {
        "pass": overall_pass,
        "source": "is_drop" if is_totals else "missing",
        "diffs": diffs,
    }


def write_data_to_sheet(ws, rows, period_text):
    # Update Row 4 with period text
    ws.cell(row=4, column=1).value = period_text
    # Clear from row 8 down
    last = max(ws.max_row, 8)
    for r in range(8, last + 1):
        for c in range(1, 16):
            ws.cell(row=r, column=c).value = None
    # Write data starting at row 8
    for i, data in enumerate(rows):
        for c, val in enumerate(data, start=1):
            ws.cell(row=8 + i, column=c).value = val
    # Update pivot source ranges if any pivots are attached
    last_row = 7 + len(rows)
    try:
        for pt in getattr(ws, "_pivots", []):
            try:
                pt.cache.cacheSource.worksheetSource.ref = f"A7:O{last_row}"
            except Exception:
                pass
    except Exception:
        pass


def main():
    # Copy master template to output path
    OUT.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy(MASTER, OUT)
    print(f"Copied template -> {OUT}")

    wb = openpyxl.load_workbook(OUT)  # preserve pivots/formulas

    summary = {}
    parsed_by_tab = {}
    dropped_by_tab = {}
    periods = [p.strip() for p in PERIOD_TEXT.split(",")]

    for tab_key, (sheet_name, cache_file, kind) in SHEET_MAP.items():
        cache_path = CACHE / cache_file
        if not cache_path.exists():
            print(f"  WARN: missing {cache_path}")
            continue
        rep = load_report(cache_path)
        rd = rep.get("reportData") or {}
        warnings = []
        if kind == "cogs":
            rows = parse_cogs(rd)
        else:
            filters = {"Contractors": CONTRACTOR_FILTERS, "ProfFees": PROFFEES_FILTERS, "Software": SOFTWARE_FILTERS}[tab_key]
            rows = parse_gl_detail(rd, filters, warnings=warnings)

        # Track dropped Brex-category amounts per period so tieback can add them back
        dropped_per_period = {p: 0.0 for p in periods}
        for w in warnings:
            date_str = w.get("date") or ""
            for p in periods:
                # Match period by month abbreviation in the date (best-effort)
                month_abbr = p.split()[0] if p else ""
                if month_abbr and month_abbr in date_str:
                    try:
                        dropped_per_period[p] += float(w.get("amount") or 0)
                    except (TypeError, ValueError):
                        pass
                    break
        if warnings:
            print(f"  {tab_key:<13}: dropped {len(warnings)} Brex-category misattribution row(s):")
            for w in warnings:
                print(f"    - {w['account_path']} | {w['entity']} ${w['amount']:.2f} | {w['doc']} | {w['memo_first_line'][:60]}")

        if sheet_name not in wb.sheetnames:
            print(f"  WARN: sheet {sheet_name!r} not in template")
            continue
        ws = wb[sheet_name]
        write_data_to_sheet(ws, rows, PERIOD_TEXT)

        # Per-period totals
        period_totals = {p: 0.0 for p in periods}
        for row in rows:
            period = row[14] or ""
            if period in period_totals:
                try:
                    period_totals[period] += float(row[9] or 0)
                except (TypeError, ValueError):
                    pass
        summary[tab_key] = (len(rows), period_totals)
        parsed_by_tab[tab_key] = rows
        dropped_by_tab[tab_key] = dropped_per_period
        totals_str = ", ".join(f"{p}=${v:,.0f}" for p, v in period_totals.items())
        print(f"  {tab_key:<13}: wrote {len(rows):>5} rows to {sheet_name!r}; {totals_str}")

    # === Tie-back to IS Drop (if available) ===
    is_drop_path = MONTH_DIR / "Flux Workbook" / "Income Statement Drop.xlsx"
    tieback_result = tieback(
        parsed_rows_by_tab=parsed_by_tab,
        dropped_amounts_by_tab=dropped_by_tab,
        periods=periods,
        is_drop_path=is_drop_path,
    )
    print(f"\n== Tie-back ({tieback_result['source']}) ==")
    for d in tieback_result["diffs"]:
        if d.get("status") == "skipped":
            print(f"  {d['tab']:<13}: SKIPPED ({d.get('note', '')[:60]})")
            continue
        marker = "PASS" if d["status"] == "pass" else "FAIL"
        print(f"  {d['tab']:<13}: {marker}")
        for pp in d.get("per_period", []):
            ok = "OK" if pp["pass"] else "X"
            print(f"     {pp['period']}: parsed=${pp['parsed']:>12,.2f} +dropped=${pp['dropped_added_back']:>8,.2f} = ${pp['comparable']:>12,.2f}  vs IS ${pp['expected_from_is']:>12,.2f}  diff=${pp['diff']:>10,.2f}  [{ok}]")

    if not tieback_result["pass"] and tieback_result["source"] != "missing":
        # If IS Drop was present and tie-back failed, we abort the save
        # (this is the design — force re-investigation before delivery).
        # We do not abort if IS Drop is missing (v1 pre-accrual case).
        print("\nFAIL: tie-back diff exceeds tolerance. NOT saving workbook. Investigate above.")
        return 1

    wb.save(OUT)
    print(f"\nSaved: {OUT}")

    # === Strip pivot caches so Excel rebuilds on open (avoids copy-paste crash) ===
    try:
        from strip_pivot_cache import strip_pivot_cache
        cache_result = strip_pivot_cache(OUT)
        print(f"Stripped pivot cache: {cache_result}")
    except Exception as e:
        print(f"WARN: strip_pivot_cache failed: {e}")

    # === Build accrual suggestions ===
    try:
        from build_accrual_suggestions import build_suggestions
        # Reload after save to apply suggestions then re-save
        wb2 = openpyxl.load_workbook(OUT)
        result = build_suggestions(
            wb=wb2,
            month_dir=MONTH_DIR,
            m_minus_2_label=periods[0],
            m_minus_1_label=periods[1],
            m_current_label=periods[2],
            prior_csv_path=None,  # caller can wire in once paths are known
        )
        wb2.save(OUT)
        # Strip cache again after the second save
        strip_pivot_cache(OUT)
        print(f"Accrual suggestions: {result['candidates_written']} candidates "
              f"({result['by_confidence']})")
        print(f"  JSON: {result['json_path']}")
    except Exception as e:
        print(f"WARN: build_accrual_suggestions failed: {e}")

    return 0


if __name__ == "__main__":
    sys.exit(main())
