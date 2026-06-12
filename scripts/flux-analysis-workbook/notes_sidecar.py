"""
notes_sidecar.py

Persist manual annotations (column-H comments on IS/BS, column-L notes on
Software pivot tab, accrual approvals) across multiple flux-workbook re-pulls
so row shifts and refreshes can't wipe them.

The sidecar lives at:
    Monthly Flux Analysis/{YYYY}/{YYYY-MM}/State/flux_notes.json

Schema (all keys are case-insensitive on read, exact on write):

{
  "version": 1,
  "updated": "2026-05-12T...",
  "sheets": {
    "Income Statement": {
      "411150 - Professional Services": {
        "H": "Higher US PS revenue +$70K...",
        "edited_by": "claude" | "ryan",
        "ts": "2026-05-11T..."
      },
      ...
    },
    "Software": {
      "Anthropic": {"L": "Increased usage/seats", "edited_by": "ryan", ...},
      ...
    }
  },
  "accruals": {
    "IT Hardware Partner Networked Solutions Group, LLC": {
      "approved": true,
      "amount": 377091.00,
      "edited_by": "ryan",
      "notes": "April estimate based on Mar Activity",
      "ts": "..."
    },
    ...
  }
}

Lookups are by VENDOR / ROW LABEL, not row index, so row shifts on re-pull do
not wipe annotations.

Merge rule:
- On re-pull, Claude-generated notes refresh against current data.
- User-edited notes (edited_by == "ryan") are NEVER overwritten unless the
  caller passes force_overwrite=True.
"""
from __future__ import annotations

import json
from dataclasses import dataclass, asdict
from datetime import datetime
from pathlib import Path
from typing import Iterable

SIDECAR_VERSION = 1
RYAN = "ryan"
CLAUDE = "claude"


def _now() -> str:
    return datetime.now().strftime("%Y-%m-%dT%H:%M:%S")


@dataclass
class NoteEntry:
    """One annotation: a (sheet, row_label, column) → text mapping with author."""

    sheet: str
    row_label: str   # column-A label for IS/BS; column-F label for detail tabs
    column: str      # "H" / "L" / etc.
    text: str
    edited_by: str   # "ryan" or "claude"
    ts: str = ""

    def __post_init__(self):
        if not self.ts:
            self.ts = _now()


@dataclass
class AccrualEntry:
    """One accrual suggestion / approval row."""

    vendor: str
    account: str            # e.g. "511400"
    amount: float
    department: str
    suggested_by: str       # "claude"
    approved: bool          # set by the accountant
    notes: str = ""
    edited_by: str = ""     # last user to touch this
    ts: str = ""

    def __post_init__(self):
        if not self.ts:
            self.ts = _now()


class FluxNotesSidecar:
    """Read/write helper for flux_notes.json under {YYYY-MM}/State/."""

    def __init__(self, month_dir: str | Path):
        self.month_dir = Path(month_dir)
        self.state_dir = self.month_dir / "State"
        self.path = self.state_dir / "flux_notes.json"
        self.data: dict = self._empty()
        if self.path.exists():
            self.load()

    @staticmethod
    def _empty() -> dict:
        return {
            "version": SIDECAR_VERSION,
            "updated": _now(),
            "sheets": {},
            "accruals": {},
        }

    # ----- Persistence -----

    def load(self) -> None:
        with open(self.path, "r", encoding="utf-8") as f:
            self.data = json.load(f)
        if self.data.get("version") != SIDECAR_VERSION:
            raise ValueError(
                f"flux_notes.json version {self.data.get('version')} "
                f"does not match expected {SIDECAR_VERSION} at {self.path}"
            )

    def save(self) -> None:
        self.state_dir.mkdir(parents=True, exist_ok=True)
        self.data["updated"] = _now()
        with open(self.path, "w", encoding="utf-8") as f:
            json.dump(self.data, f, indent=2, ensure_ascii=False)

    # ----- Note CRUD -----

    def get_note(self, sheet: str, row_label: str, column: str = "H") -> dict | None:
        """Return the persisted note dict for a (sheet, row_label, column), or None."""
        if not row_label:
            return None
        key = row_label.strip()
        sheet_map = self.data["sheets"].get(sheet, {})
        entry = sheet_map.get(key)
        if entry and entry.get(column):
            return entry
        return None

    def set_note(
        self,
        sheet: str,
        row_label: str,
        column: str,
        text: str,
        edited_by: str,
        force: bool = False,
    ) -> bool:
        """
        Write a note. If a the accountant-edited note already exists at this cell and
        the caller is Claude, do NOT overwrite (unless force=True).
        Returns True if the note was written, False if skipped.
        """
        if not row_label:
            return False
        key = row_label.strip()

        existing = self.get_note(sheet, key, column)
        if existing and existing.get("edited_by") == RYAN and edited_by != RYAN and not force:
            return False

        sheets = self.data["sheets"].setdefault(sheet, {})
        entry = sheets.setdefault(key, {})
        entry[column] = text
        entry["edited_by"] = edited_by
        entry["ts"] = _now()
        return True

    def record_user_edit(self, sheet: str, row_label: str, column: str, text: str) -> None:
        """Capture an edit made directly in Excel by the accountant."""
        self.set_note(sheet, row_label, column, text, edited_by=RYAN, force=True)

    # ----- Accrual CRUD -----

    def list_accruals(self) -> list[dict]:
        return [
            {"vendor": v, **payload}
            for v, payload in self.data.get("accruals", {}).items()
        ]

    def get_accrual(self, vendor: str) -> dict | None:
        return self.data.get("accruals", {}).get(vendor.strip())

    def set_accrual(self, entry: AccrualEntry) -> bool:
        """
        Write an accrual suggestion. If the accountant has already touched this vendor's
        accrual (approved or modified), do NOT overwrite the amount/notes
        unless the caller is also the accountant.
        Returns True if written, False if skipped.
        """
        key = entry.vendor.strip()
        existing = self.data.setdefault("accruals", {}).get(key)
        if (
            existing
            and existing.get("edited_by") == RYAN
            and entry.edited_by != RYAN
        ):
            # Preserve the accountant's amount/approved/notes, but refresh the suggested context fields
            existing.setdefault("history", []).append(
                {
                    "claude_suggested_amount": entry.amount,
                    "claude_ts": entry.ts,
                }
            )
            return False

        payload = asdict(entry)
        payload.pop("vendor")  # vendor is the key
        self.data["accruals"][key] = payload
        return True

    def approve_accrual(self, vendor: str, amount: float | None = None, notes: str = "") -> bool:
        key = vendor.strip()
        existing = self.data.setdefault("accruals", {}).get(key)
        if not existing:
            return False
        existing["approved"] = True
        if amount is not None:
            existing["amount"] = float(amount)
        if notes:
            existing["notes"] = notes
        existing["edited_by"] = RYAN
        existing["ts"] = _now()
        return True

    def reject_accrual(self, vendor: str) -> bool:
        key = vendor.strip()
        existing = self.data.setdefault("accruals", {}).get(key)
        if not existing:
            return False
        existing["approved"] = False
        existing["edited_by"] = RYAN
        existing["ts"] = _now()
        return True

    def approved_accruals(self) -> list[dict]:
        return [
            {"vendor": v, **payload}
            for v, payload in self.data.get("accruals", {}).items()
            if payload.get("approved")
        ]

    # ----- Bulk import from workbook -----

    def scan_workbook_user_edits(
        self,
        wb,
        cells: Iterable[tuple[str, int, str, int]],
    ) -> int:
        """
        Walk a list of (sheet_name, row_label_column, note_column, start_row)
        descriptors and capture any cell value that differs from the persisted
        Claude-generated note as a user edit.

        Use case: the accountant opens the workbook in Excel, types into column H or L,
        saves. On the next skill run, before re-populating, call this to
        capture his edits into the sidecar.

        Returns count of edits captured.
        """
        from openpyxl import Workbook  # local import to avoid hard dep at import time

        captured = 0
        for sheet_name, label_col, note_col, start_row in cells:
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            for r in range(start_row, ws.max_row + 1):
                label = ws.cell(r, label_col).value
                note = ws.cell(r, note_col).value
                if not label or not note:
                    continue
                label_str = str(label).strip()
                note_str = str(note).strip()
                column_letter = chr(ord("A") + note_col - 1)
                persisted = self.get_note(sheet_name, label_str, column_letter)
                # If the persisted note matches the cell value, nothing to do.
                if persisted and persisted.get(column_letter) == note_str:
                    continue
                # If persisted was claude and the cell now differs, that's a the accountant edit.
                # If no persisted entry exists, treat the existing cell as the accountant's.
                self.record_user_edit(sheet_name, label_str, column_letter, note_str)
                captured += 1
        return captured
