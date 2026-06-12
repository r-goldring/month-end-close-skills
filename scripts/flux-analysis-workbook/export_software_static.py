"""
Export the Software tab from Flux Pivot Template 2026-04 as a clean, static-values
standalone xlsx with NO pivot dependencies. Safe to copy/paste into the Flux workbook.

The original template Software tab carries a PivotTable whose cache definition
points to the GeneralLedgerdetailSoftwa hidden data sheet. When Excel tries to
copy that pivot tab to another workbook, it must also bring the cache + cache
records + data sheet, and openpyxl's preserved-cache lengths are out of sync
with the post-refresh sheet contents - hence the crash.

This script reads cached cell values (data_only=True) and rewrites them as plain
values into a fresh workbook with one Software sheet only.
"""
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border
from copy import copy

SRC = Path(r"c:/Users/Accountant/Documents/Finance's Requests/Antigravity/Monthly-Accounting/Monthly Flux Analysis/2026/2026-04/Flux Pivot Template 2026-04.xlsx")
DST = SRC.parent / "Software Tab 2026-04 (static).xlsx"


def main():
    src = openpyxl.load_workbook(SRC, data_only=True)
    src_sw = src["Software"]

    # Build a fresh workbook with just the Software tab as static values
    dst = openpyxl.Workbook()
    dst_sw = dst.active
    dst_sw.title = "Software"

    # Copy values + per-cell formatting
    max_row = src_sw.max_row
    max_col = src_sw.max_column
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            src_cell = src_sw.cell(row=r, column=c)
            dst_cell = dst_sw.cell(row=r, column=c)
            dst_cell.value = src_cell.value
            if src_cell.has_style:
                try:
                    dst_cell.font = copy(src_cell.font)
                    dst_cell.fill = copy(src_cell.fill)
                    dst_cell.border = copy(src_cell.border)
                    dst_cell.alignment = copy(src_cell.alignment)
                    dst_cell.number_format = src_cell.number_format
                except Exception:
                    pass

    # Copy column widths
    for col_letter, col_dim in src_sw.column_dimensions.items():
        if col_dim.width:
            dst_sw.column_dimensions[col_letter].width = col_dim.width

    # Copy row heights
    for row_num, row_dim in src_sw.row_dimensions.items():
        if row_dim.height:
            dst_sw.row_dimensions[row_num].height = row_dim.height

    # Freeze panes if source had them
    if src_sw.freeze_panes:
        dst_sw.freeze_panes = src_sw.freeze_panes

    dst.save(DST)
    print(f"Saved standalone Software tab: {DST}")
    print(f"  Rows copied: {max_row}, Cols copied: {max_col}")


if __name__ == "__main__":
    main()
