"""
Canada Payroll -> NetSuite JE generator.

Reads the bi-weekly ADP Canada export (.xls), parses Sheet1 (per-employee earnings,
deductions, taxes by cost center) and Sheet2 (Total Net, RRSP totals, EE/ER tax
totals), maps each code to a GL account, routes COGS vs OpEx by department (with
a 70/30 split for Infrastructure), and writes:

  - {MM.DD.YYYY} Canada Payroll Backup.xlsx — raw + JE tabs
  - {MM.DD.YYYY} Canada Payroll JE Import.csv — NetSuite Import Assistant format

Cadence: bi-weekly (15th + last day of month). Currency: CAD.
Subsidiary: Acme Holdings : Acme, Inc. : Acme Canada.

Usage:
    python process_canada_payroll.py <input.xls> [output_dir] [dept_map.csv] [gl_map.csv]
"""

import os
import sys
from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, os.path.join(SCRIPT_DIR, "..", "_shared"))
from je_csv_writer import write_je_csv, make_external_id  # noqa: E402

SUBSIDIARY = "Acme Holdings : Acme, Inc. : Acme Canada"


def parse_amt(val):
    if pd.isna(val): return 0.0
    val = str(val).strip().replace(',', '').replace('"', '').replace('$', '')
    if not val: return 0.0
    if val.startswith('(') and val.endswith(')'):
        return -float(val[1:-1])
    try:
        return float(val)
    except:
        return 0.0

def generate_je_csv(raw_xls, dept_map_csv, gl_map_csv, output_csv):
    dept_df = pd.read_csv(dept_map_csv)
    dept_map = {}
    for _, row in dept_df.iterrows():
        cc_full = str(row['Payroll Report Cost Center']).strip()
        cc = cc_full.split("Cost Center: ")[1].strip() if "Cost Center: " in cc_full else cc_full
        dept_map[cc] = {
            'department': str(row['Department']).strip(),
            'subsidiary': str(row['Subsidiary']).strip()
        }

    gl_df = pd.read_csv(gl_map_csv)
    gl_map = {}
    for _, row in gl_df.iterrows():
        code = str(row['Payroll Code']).strip()
        if code not in gl_map:
            gl_map[code] = []
        gl_map[code].append({
            'name': str(row['GL Account Name']).strip(),
            'opex': str(row['GL Account - Opex']).strip() if not pd.isna(row['GL Account - Opex']) else '',
            'cogs': str(row['GL Account - COGS']).strip() if not pd.isna(row['GL Account - COGS']) else '',
            'type': str(row['Type']).strip(),
            'memo': str(row['Line Memo']).strip()
        })

    # Retrieve tax accounts for the department summary
    tax_m = [m for code, maps in gl_map.items() for m in maps if m['type'] == 'Expense' and 'Payroll Tax' in m['name']]
    if tax_m:
        tax_opex = tax_m[0]['opex']
        tax_cogs = tax_m[0]['cogs']
        tax_memo = "CAN PAYROLL - Payroll Taxes"
    else:
        tax_opex = "611450 Salary and Compensation : Payroll Taxes"
        tax_cogs = "511350 COGS - Salary and Compensation : COGS - Payroll Taxes"
        tax_memo = "CAN PAYROLL - Payroll Taxes"

    df1 = pd.read_excel(raw_xls, sheet_name='Sheet1', header=None)
    
    je_date = "2/28/2026" 
    je_memo = "2.28.26 Canada Payroll"
    for r in range(15):
        for c in range(10):
            val = str(df1.iat[r, c])
            if "Summary Payroll Register" in val:
                parts = val.split()
                for p in parts:
                    if p.count('.') == 2:
                        m, d, y = p.split('.')
                        je_date = f"{int(m)}/{int(d)}/20{y}" if len(y) == 2 else f"{int(m)}/{int(d)}/{y}"
                        je_memo = f"{p} Canada Payroll"
                        break

    row_8_idx = -1
    for r in range(5, 12):
        row_str = " ".join([str(x) for x in df1.iloc[r].values if pd.notna(x)])
        if "Earnings" in row_str and "Employee Deds" in row_str:
            row_8_idx = r
            break
            
    cat_ranges = {}
    if row_8_idx != -1:
        row8 = df1.iloc[row_8_idx].values
        categories = ["Earnings", "Employee Deds", "Employee Taxes", "Employer Ded Exp", "Employer Tax Exp"]
        found_cats = []
        for c in range(len(row8)):
            val = str(row8[c]).strip()
            for cat in categories:
                if cat in val:
                    found_cats.append((c, cat))
        found_cats.sort(key=lambda x: x[0])
        for i in range(len(found_cats)):
            start_col = found_cats[i][0]
            end_col = found_cats[i+1][0] - 1 if i < len(found_cats) - 1 else len(row8) - 1
            cat_ranges[found_cats[i][1]] = (start_col, end_col)

    records = []
    current_cc = None
    in_block = False
    dept_er_tax_total = 0.0
    department = ""
    subsidiary = ""
    
    memo_date = f"{je_date.split('/')[2]}.{je_date.split('/')[0].zfill(2)}.{je_date.split('/')[1].zfill(2)}"

    def flush_tax_total():
        nonlocal dept_er_tax_total, department, subsidiary
        if dept_er_tax_total != 0 and department and subsidiary:
             def add_rec(acc, val):
                 records.append({'Date': je_date, 'Subsidiary': subsidiary, 'Department': department, 'Account': acc, 'Journal Entry Memo': je_memo, 'Line Memo': f"{memo_date} {tax_memo}", ' Debit ': val})
             if department in ["COGS : Consulting", "COGS : Professional Services"]:
                 add_rec(tax_cogs or tax_opex, dept_er_tax_total)
             elif department == "Engineering : Infrastructure":
                 amt_opex = round(dept_er_tax_total * 0.7, 2)
                 amt_cogs = round(dept_er_tax_total - amt_opex, 2)
                 if amt_cogs != 0: add_rec(tax_cogs or tax_opex, amt_cogs)
                 if amt_opex != 0: add_rec(tax_opex, amt_opex)
             else:
                 add_rec(tax_opex, dept_er_tax_total)
        dept_er_tax_total = 0.0
        department = ""
        subsidiary = ""
    
    for row_idx in range(row_8_idx + 2, len(df1)):
        row_vals = df1.iloc[row_idx].values
        row_str_joined = " ".join([str(v) for v in row_vals if pd.notna(v)])
        
        if "Group Summary for:" in row_str_joined and "Cost Center:" in row_str_joined:
            flush_tax_total()
            current_cc = row_str_joined.split("Cost Center:")[1].strip() if len(row_str_joined.split("Cost Center:")) > 1 else None
            in_block = True
            
            dept_info = dept_map.get(current_cc) or next((v for k, v in dept_map.items() if current_cc and k in current_cc), None)
            if not dept_info and current_cc == "Revenue":
                dept_info = {'department': 'Sales & Marketing : Revenue', 'subsidiary': 'Acme Holdings : Acme, Inc. : Acme Canada'}
                
            if dept_info:
                department = dept_info['department']
                subsidiary = dept_info['subsidiary']
            else:
                department = ""
                subsidiary = ""
            continue
        elif "Group Totals:" in row_str_joined:
            flush_tax_total()
            in_block = False
            current_cc = None
            continue
            
        if in_block and current_cc and department:
            for c in range(len(row_vals)):
                cell_val = str(row_vals[c]).strip()
                if not cell_val or cell_val == 'nan': continue
                
                cat = None
                end_col = len(row_vals) - 1
                for k, v in cat_ranges.items():
                    if v[0] <= c <= v[1]:
                        cat = k
                        end_col = min(v[1], len(row_vals) - 1)
                        break
                
                if not cat: continue
                
                amt_val = 0.0
                found_num = False
                for c2 in range(end_col, c, -1):
                    val_str = str(row_vals[c2]).strip()
                    if val_str and val_str != 'nan':
                        try:
                            v = val_str.replace(',', '').replace('"', '').replace('$', '')
                            if v.startswith('(') and v.endswith(')'): v = "-" + v[1:-1]
                            amt_val = float(v)
                            found_num = True
                            break
                        except:
                            pass
                            
                if not found_num or amt_val == 0.0: continue
                
                if cat == 'Employee Deds':
                    if cell_val == 'CLTD': 
                        amt = -abs(amt_val)
                    else: continue
                elif cat == 'Employee Taxes':
                    continue
                elif cat == 'Employer Tax Exp':
                    # Instead of validating the precise tax code, sum ALL employer taxes
                    # as requested by user to consolidate taxes and avoid mapping mismatches.
                    dept_er_tax_total += abs(amt_val)
                    continue
                else: # Earnings, Employer Ded Exp
                    if cell_val not in gl_map:
                        continue # If an earning/match isn't exact mapped, skip
                    # Preserve sign so retro reversals (e.g., -823.21 CREG)
                    # become a CR to Salaries instead of being abs'd into a DR.
                    amt = amt_val
                    
                mappings = [m for m in gl_map.get(cell_val, []) if m['type'] == 'Expense']
                if not mappings: continue
                m = mappings[0]
                
                def add_rec(acc, val):
                    records.append({'Date': je_date, 'Subsidiary': subsidiary, 'Department': department, 'Account': acc, 'Journal Entry Memo': je_memo, 'Line Memo': f"{memo_date} {m['memo']}", ' Debit ': val})
                
                if department in ["COGS : Consulting", "COGS : Professional Services"]:
                    add_rec(m['cogs'] or m['opex'], amt)
                elif department == "Engineering : Infrastructure":
                    amt_opex = round(amt * 0.7, 2)
                    amt_cogs = round(amt - amt_opex, 2)
                    if amt_cogs != 0: add_rec(m['cogs'] or m['opex'], amt_cogs)
                    if amt_opex != 0: add_rec(m['opex'], amt_opex)
                else:
                    add_rec(m['opex'], amt)
    
    # Aggregate department expenses by Account to combine multiple Salary lines into one
    expense_df = pd.DataFrame(records)
    if not expense_df.empty:
        expense_df = expense_df.groupby(['Date', 'Subsidiary', 'Department', 'Account', 'Journal Entry Memo', 'Line Memo'], dropna=False, as_index=False)[' Debit '].sum()
        records = expense_df.to_dict('records')

    # Flush memory in case the last block ended at EOF
    flush_tax_total()

    # Read Sheet 2 for Liabilities
    df2 = pd.read_excel(raw_xls, sheet_name='Sheet2', header=None)
    
    total_net = 0.0
    for row_idx in range(len(df2)):
        row_str_joined = " ".join([str(v) for v in df2.iloc[row_idx].values if pd.notna(v)])
        if "Total Net:" in row_str_joined:
            row_vals = df2.iloc[row_idx].values
            for c in range(len(row_vals)):
                if str(row_vals[c]).strip() == "Total Net:":
                    for ptr in range(c + 1, len(row_vals)):
                        if pd.notna(row_vals[ptr]) and str(row_vals[ptr]).strip() != '':
                            total_net = parse_amt(row_vals[ptr])
                            break
                    break
            break

    total_401k = 0.0
    in_report_totals = False
    for row_idx in range(len(df2)):
        row_vals = df2.iloc[row_idx].values
        row_str_joined = " ".join([str(v) for v in row_vals if pd.notna(v)])
        if "Report Totals:" in row_str_joined: in_report_totals = not in_report_totals
        if in_report_totals:
            for c in range(len(row_vals)):
                if str(row_vals[c]).strip() in ["CRRSP", "CRRVF", "CRSPV"]:
                    for ptr in range(c + 1, len(row_vals)):
                        if pd.notna(row_vals[ptr]) and str(row_vals[ptr]).strip() != '':
                            total_401k += parse_amt(row_vals[ptr])
                            break

    ee_taxes = 0.0
    er_taxes = 0.0
    col_ee_amount = -1
    col_er_amount = -1
    for row_idx in range(min(20, len(df2))):
        row_vals = df2.iloc[row_idx].values
        for c in range(len(row_vals)):
            val = str(row_vals[c]).strip()
            # In Sheet 2, the Category headers perfectly align their columns with the Amounts on the Sum line
            if "Employee Taxes" in val:
                col_ee_amount = c
            if "Employer Tax Exp" in val:
                col_er_amount = c

    count_rt = 0
    for row_idx in range(len(df2)):
        if "Report Totals:" in " ".join([str(v) for v in df2.iloc[row_idx].values if pd.notna(v)]):
            count_rt += 1
            if count_rt == 2:
                sum_row = row_idx + 1
                if sum_row < len(df2):
                    next_row = df2.iloc[sum_row].values
                    if col_ee_amount != -1 and col_ee_amount < len(next_row):
                        ee_taxes = parse_amt(next_row[col_ee_amount])
                    if col_er_amount != -1 and col_er_amount < len(next_row):
                        er_taxes = parse_amt(next_row[col_er_amount])
                break

    payroll_tax_liability = -(ee_taxes + er_taxes)
    
    memo_date = f"{je_date.split('/')[2]}.{je_date.split('/')[0].zfill(2)}.{je_date.split('/')[1].zfill(2)}"
    
    pliab = next((m['opex'] for m in sum(gl_map.values(), []) if m['type'] == 'Liability' and 'Payroll' in m['name'] and 'Tax' not in m['name']), "231200 Payroll Liability")
    records.append({'Date': je_date, 'Subsidiary': "Acme Holdings : Acme, Inc. : Acme Canada", 'Department': "", 'Account': pliab, 'Journal Entry Memo': je_memo, 'Line Memo': f"{memo_date} Canada Payroll Liability", ' Debit ': -total_net})
    
    k401liab = next((m['opex'] for m in sum(gl_map.values(), []) if m['type'] == 'Liability' and '401K' in m['name']), "231250 401K payable")
    records.append({'Date': je_date, 'Subsidiary': "Acme Holdings : Acme, Inc. : Acme Canada", 'Department': "", 'Account': k401liab, 'Journal Entry Memo': je_memo, 'Line Memo': f"{memo_date} Canada 401K Liability", ' Debit ': -total_401k})
    
    taxliab = next((m['opex'] for m in sum(gl_map.values(), []) if m['type'] == 'Liability' and 'Tax' in m['name']), "231350 Payroll Tax Liability")
    records.append({'Date': je_date, 'Subsidiary': "Acme Holdings : Acme, Inc. : Acme Canada", 'Department': "", 'Account': taxliab, 'Journal Entry Memo': je_memo, 'Line Memo': f"{memo_date} Canada Payroll Tax Liability", ' Debit ': payroll_tax_liability})
    
    # Convert internal records (single ' Debit ' column with negatives = credits)
    # into NetSuite Import Assistant schema (separate Debit/Credit columns).
    je_rows = []
    for rec in records:
        try:
            amt = float(rec[' Debit '])
        except (TypeError, ValueError):
            continue
        if amt == 0:
            continue
        if amt > 0:
            debit_val = round(amt, 2)
            credit_val = ""
        else:
            debit_val = ""
            credit_val = round(-amt, 2)
        je_rows.append({
            "Date": rec["Date"],
            "Journal Entry Memo": rec["Journal Entry Memo"],
            "Account": rec["Account"],
            "Debit": debit_val,
            "Credit": credit_val,
            "Line Memo": rec["Line Memo"],
            "Subsidiary": rec["Subsidiary"],
            "Department": rec["Department"],
        })

    total_dr = sum(r["Debit"] for r in je_rows if isinstance(r["Debit"], (int, float)))
    total_cr = sum(r["Credit"] for r in je_rows if isinstance(r["Credit"], (int, float)))
    imbalance = round(total_dr - total_cr, 2)
    balanced = abs(imbalance) <= 0.01

    # Filename prefix uses pay date in MM.DD.YYYY (matches the folder convention).
    parts = je_date.split("/")
    file_prefix = f"{parts[0].zfill(2)}.{parts[1].zfill(2)}.{parts[2]}"
    iso_date = f"{parts[2]}-{parts[0].zfill(2)}-{parts[1].zfill(2)}"

    out_dir = os.path.dirname(os.path.abspath(output_csv))
    backup_xlsx = os.path.join(out_dir, f"{file_prefix} Canada Payroll Backup.xlsx")
    csv_path = os.path.join(out_dir, f"{file_prefix} Canada Payroll JE Import.csv")

    wb = Workbook()
    ws_raw = wb.active
    ws_raw.title = "raw_Sheet1"
    for row in dataframe_to_rows(df1, index=False, header=False):
        ws_raw.append(list(row))
    ws_raw2 = wb.create_sheet("raw_Sheet2")
    for row in dataframe_to_rows(df2, index=False, header=False):
        ws_raw2.append(list(row))
    ws_je = wb.create_sheet("JE")
    cols = ["Date", "Journal Entry Memo", "Account", "Debit", "Credit", "Line Memo", "Subsidiary", "Department"]
    ws_je.append(cols)
    for r in je_rows:
        ws_je.append([r[c] for c in cols])
    wb.save(backup_xlsx)

    if balanced:
        write_je_csv(
            rows=je_rows,
            output_path=csv_path,
            currency="CAD",
            default_external_id=make_external_id(iso_date, "CA"),
        )

    print(f"Wrote {backup_xlsx}")
    if balanced:
        print(f"Wrote {csv_path}")
    else:
        print(f"CSV NOT written (imbalanced — fix Backup.xlsx and re-run)")
    print(f"  JE date:   {je_date}")
    print(f"  JE memo:   {je_memo}")
    print(f"  Lines:     {len(je_rows)}")
    print(f"  Total Dr:  {total_dr:,.2f} CAD")
    print(f"  Total Cr:  {total_cr:,.2f} CAD")
    print(f"  Imbalance: {imbalance:,.2f} CAD")
    print(f"  Balanced:  {balanced}")

if __name__ == "__main__":
    script_dir = os.path.dirname(os.path.abspath(__file__))
    dept_map_default = os.path.join(script_dir, "Canada Payroll Department Mapping File.csv")
    gl_map_default = os.path.join(script_dir, "Payroll_Mapping_with_GL_Accounts (final).csv")

    if len(sys.argv) < 2:
        print("Usage: python process_canada_payroll.py <input.xls> [dept_map.csv] [gl_map.csv]")
        sys.exit(1)

    raw_file = sys.argv[1]
    input_dir = os.path.dirname(os.path.abspath(raw_file))

    # output_csv arg below is a placeholder used only to derive the output
    # directory; the actual filenames (Backup.xlsx + JE Import.csv) are
    # generated inside generate_je_csv() using the pay date.
    out_csv = os.path.join(input_dir, "_unused.csv")
    dept_map = sys.argv[2] if len(sys.argv) > 2 else dept_map_default
    gl_map = sys.argv[3] if len(sys.argv) > 3 else gl_map_default

    print(f"Starting Canadian Payroll Generation")
    print(f"  Input:  {raw_file}")
    generate_je_csv(raw_file, dept_map, gl_map, out_csv)
