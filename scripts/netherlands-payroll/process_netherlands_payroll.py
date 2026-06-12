"""
Netherlands Monthly Payroll - JE Generator

Reads the raw Dutch payroll export (CompanyEmployeeWageComponents*.xlsx) and
produces a 'YYYY-MM Netherlands Payroll Backup.xlsx' workbook with three
sheets: Sheet1 (pivot-like aggregation), Page_1 (raw + Department), JE (the
journal entry, balanced, ready for NetSuite import).

Usage:
  python process_netherlands_payroll.py <raw_input.xlsx|folder> [output.xlsx]
"""

import os
import re
import sys
import glob
import calendar
import datetime
from collections import defaultdict
from pathlib import Path

import pandas as pd
import openpyxl
from openpyxl import Workbook

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, os.path.join(SCRIPT_DIR, "..", "_shared"))
from je_csv_writer import write_je_csv, make_external_id  # noqa: E402

EMP_MAP_PATH = os.path.join(SCRIPT_DIR, "Netherlands Payroll Employee Mapping.csv")
COMP_MAP_PATH = os.path.join(SCRIPT_DIR, "Netherlands Payroll Component Mapping.csv")

SUBSIDIARY = "Acme Holdings : Acme, Inc. : Acme Netherlands"
CURRENCY = "EUR"

# Standard JE template -- 16 rows in this exact order (matches existing backups)
JE_TEMPLATE = [
    # (account, side, route_key, dept_key, line_memo_suffix)
    ("611100 Salary and Compensation : Salaries and Wages", "D", "salary_expense", "Product", "BV Payroll - Salary"),
    ("511100 COGS - Salary and Compensation : COGS - Salaries and Wages", "D", "salary_expense", "Professional Services", "BV Payroll - Salary"),
    ("231206 Netherlands Holiday Pay Accrual", "D", "holiday_payout_debit", "Product", "BV Payroll - Holiday Allowance Payout"),
    ("231206 Netherlands Holiday Pay Accrual", "D", "holiday_payout_debit", "Professional Services", "BV Payroll - Holiday Allowance Payout"),
    ("611400 Salary and Compensation : Other Benefits", "D", "zero", "Product", "BV Payroll - Other Benefits"),
    ("511300 COGS - Salary and Compensation : COGS - Other Benefits", "D", "zero", "Professional Services", "BV Payroll - Other Benefits"),
    ("611350 Salary and Compensation : 401k Match", "D", "match_401k_expense", "Product", "BV Payroll - 401K Match"),
    ("511250 COGS - Salary and Compensation : COGS - 401k Match", "D", "match_401k_expense", "Professional Services", "BV Payroll - 401K Match"),
    ("611150 Salary and Compensation : Bonus", "D", "bonus_expense", "Product", "BV Payroll - MIP"),
    ("511150 COGS - Salary and Compensation : COGS - Bonus", "D", "bonus_expense", "Professional Services", "BV Payroll - MIP"),
    ("611450 Salary and Compensation : Payroll Taxes", "D", "tax_expense", "Product", "BV Payroll Taxes"),
    ("511350 COGS - Salary and Compensation : COGS - Payroll Taxes", "D", "tax_expense", "Professional Services", "BV Payroll Taxes"),
    ("231200 Payroll Liability", "C", "payroll_liability", None, "BV Payroll Liability"),
    ("231205 Netherlands Pension Payable", "C", "pension_liability", None, "BV Pension Liability"),
    ("231350 Payroll Tax Liability", "C", "tax_liability", None, "BV Payroll Tax Liability"),
    ("231206 Netherlands Holiday Pay Accrual", "C", "holiday_accrual_liability", None, "BV Holiday Pay Accrual"),
]


def parse_period_from_folder(folder_name: str) -> tuple:
    """'03-2026' -> (2026, 3); '2026-03' -> (2026, 3); '03.31.2026' -> (2026, 3)."""
    m = re.match(r"^(\d{2})-(\d{4})$", folder_name)
    if m:
        return int(m.group(2)), int(m.group(1))
    m = re.match(r"^(\d{4})-(\d{2})$", folder_name)
    if m:
        return int(m.group(1)), int(m.group(2))
    m = re.match(r"^(\d{2})\.\d{2}\.(\d{4})$", folder_name)
    if m:
        return int(m.group(2)), int(m.group(1))
    raise ValueError(f"Cannot parse year/month from folder name: {folder_name!r}")


def find_period_column(ws, month_num: int) -> int:
    """Return 1-based column index of the current month's amount column.
    Accepts headers like 'P3' or 'R2/P3'."""
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    # Prefer plain P{month}
    for i, h in enumerate(headers):
        if h is not None and str(h).strip() == f"P{month_num}":
            return i + 1
    # Fall back to R?/P{month}
    for i, h in enumerate(headers):
        if h is not None and re.match(rf"R\d+/P{month_num}$", str(h).strip()):
            return i + 1
    raise ValueError(f"Could not find period column for month P{month_num}. Headers: {headers}")


def find_raw_file(target: str) -> str:
    """Accept a folder or a direct path. Return the raw NL payroll xlsx."""
    if os.path.isfile(target):
        return target
    if os.path.isdir(target):
        matches = [
            f for f in glob.glob(os.path.join(target, "CompanyEmployeeWageComponents*.xlsx"))
            if "Backup" not in os.path.basename(f) and not os.path.basename(f).startswith("~")
        ]
        if not matches:
            raise FileNotFoundError(f"No raw NL payroll file found in {target}")
        if len(matches) > 1:
            print(f"WARNING: multiple raw files found; using {matches[0]!r}")
        return matches[0]
    raise FileNotFoundError(f"Not a file or folder: {target!r}")


def load_employee_map():
    df = pd.read_csv(EMP_MAP_PATH)
    df["Employee ID"] = df["Employee ID"].astype(int)
    return df.set_index("Employee ID").to_dict("index")


def load_component_map():
    df = pd.read_csv(COMP_MAP_PATH).fillna("")
    return df.set_index("Component Name").to_dict("index")


def read_raw_records(raw_xlsx: str, month_num: int, employee_map: dict):
    """Return a list of {Employee ID, Name, Code, Component, Amount, Department}."""
    wb = openpyxl.load_workbook(raw_xlsx, data_only=True)
    ws = wb["Page_1"]
    period_col = find_period_column(ws, month_num)

    records = []
    for r in range(2, ws.max_row + 1):
        emp_id_raw = ws.cell(row=r, column=1).value
        if emp_id_raw is None:
            continue
        try:
            emp_id = int(emp_id_raw)
        except (ValueError, TypeError):
            continue

        emp_name = ws.cell(row=r, column=2).value
        comp_code_raw = ws.cell(row=r, column=3).value
        comp_code = int(comp_code_raw) if comp_code_raw is not None and str(comp_code_raw).replace(".0", "").lstrip("-").isdigit() else None
        comp_name = ws.cell(row=r, column=4).value
        amount_raw = ws.cell(row=r, column=period_col).value
        try:
            amount = float(amount_raw) if amount_raw is not None else 0.0
        except (ValueError, TypeError):
            amount = 0.0

        if not comp_name or amount == 0.0:
            # Keep zeros? Usually we skip; but non-zero components drive all JE lines
            if amount == 0.0 and comp_name:
                continue

        emp_info = employee_map.get(emp_id, {})
        department = emp_info.get("Department", "UNKNOWN")

        records.append({
            "Employee ID": emp_id,
            "Employee Name": emp_name,
            "Component Code": comp_code,
            "Component Name": str(comp_name).strip() if comp_name else "",
            "Amount": amount,
            "Department": department,
        })
    wb.close()
    return records


def route_components(records: list, component_map: dict):
    """Build routing buckets keyed for the JE template."""
    buckets = defaultdict(float)

    # Department -> short key ("Product" | "Professional Services")
    def dept_short(full_dept: str) -> str:
        if not full_dept:
            return "UNKNOWN"
        if "Product" in full_dept:
            return "Product"
        if "Professional Services" in full_dept:
            return "Professional Services"
        return full_dept

    unmapped_components = []

    for rec in records:
        name = rec["Component Name"]
        dept = dept_short(rec["Department"])
        amt = rec["Amount"]

        if name not in component_map:
            unmapped_components.append((name, rec["Component Code"], rec["Employee ID"], amt))
            continue

        route = component_map[name]
        abs_for_liability = (str(route.get("Liability Absolute Value", "")).strip().lower() == "yes")

        # --- Expense / debit side ---
        expense_opex = str(route.get("Expense Account - OpEx (Product)", "")).strip()
        expense_cogs = str(route.get("Expense Account - COGS (Professional Services)", "")).strip()

        if expense_opex or expense_cogs:
            expense_amount = amt  # expense uses raw value (may be negative but typically positive)
            # Category key for template matching
            cat_key = None
            # Salary expense captures Gross salary + Reservation holiday allowance
            if "Salaries and Wages" in expense_opex or "COGS - Salaries and Wages" in expense_cogs:
                cat_key = "salary_expense"
            elif "401k Match" in expense_opex or "401k Match" in expense_cogs:
                cat_key = "match_401k_expense"
            elif "Bonus" in expense_opex or "COGS - Bonus" in expense_cogs:
                cat_key = "bonus_expense"
            elif "Payroll Taxes" in expense_opex or "COGS - Payroll Taxes" in expense_cogs:
                cat_key = "tax_expense"
            if cat_key:
                buckets[(cat_key, dept)] += expense_amount

        # --- Liability / credit side ---
        liability_acct = str(route.get("Liability Account", "")).strip()
        if liability_acct:
            liability_amount = abs(amt) if abs_for_liability else amt
            lcat_key = None
            bucket_dept_key = None  # most liabilities are aggregate across depts
            if "231200" in liability_acct:
                lcat_key = "payroll_liability"
            elif "231205" in liability_acct:
                lcat_key = "pension_liability"
            elif "231350" in liability_acct:
                lcat_key = "tax_liability"
            elif "231206" in liability_acct:
                # Distinguish accrual (7711 Reservation holiday allowance, monthly
                # credit) from payout (7710 Holiday allowance, May Vakantiegeld
                # debit). Payout is per-dept since the template has separate
                # Product / Professional Services payout lines. Confirmed via
                # NetSuite history JE##### (2024-05) + JE##### (2025-05).
                if "7710" == str(rec.get("Component Code", "")).strip() \
                        or name.strip().lower() == "holiday allowance":
                    lcat_key = "holiday_payout_debit"
                    bucket_dept_key = dept
                else:
                    lcat_key = "holiday_accrual_liability"
            if lcat_key:
                buckets[(lcat_key, bucket_dept_key)] += liability_amount

    return buckets, unmapped_components


def build_je_rows(buckets: dict, year: int, month: int):
    """Emit the 16-row JE template populated with aggregated amounts."""
    je_date = datetime.date(year, month, calendar.monthrange(year, month)[1])
    period_memo = f"{year}-{month:02d}"  # e.g., "2026-03"
    journal_memo = f"{period_memo} NL Payroll"

    # Department full NS paths
    dept_full = {
        "Product": "Research & Development : Product",
        "Professional Services": "COGS : Professional Services",
    }

    rows = []
    for account, side, route_key, dept_key, line_memo_suffix in JE_TEMPLATE:
        amount = buckets.get((route_key, dept_key), 0.0) if route_key != "zero" else 0.0
        line_memo = f"{period_memo} {line_memo_suffix}"
        department_full = dept_full.get(dept_key, "") if dept_key else ""
        row = {
            "Date": je_date,
            "Journal Entry Memo": journal_memo,
            "Subsidiary": SUBSIDIARY,
            "Currency": CURRENCY,
            "Account": account,
            "Debit": round(amount, 2) if side == "D" else None,
            "Credit": round(amount, 2) if side == "C" else None,
            "Line Memo": line_memo,
            "Department": department_full,
        }
        rows.append(row)
    return rows


def write_output(output_path: str, raw_xlsx: str, records: list, buckets: dict, je_rows: list, year: int, month: int):
    """Write the YYYY-MM Netherlands Payroll Backup.xlsx with three sheets."""
    wb = Workbook()

    # --- Sheet1: component aggregation by dept ---
    s1 = wb.active
    s1.title = "Sheet1"
    s1.cell(1, 1).value = f"Netherlands Payroll aggregation - {year}-{month:02d}"
    s1.cell(3, 1).value = "Department"
    s1.cell(3, 2).value = "Component"
    s1.cell(3, 3).value = "Amount"

    agg = defaultdict(float)
    for rec in records:
        agg[(rec["Department"], rec["Component Name"])] += rec["Amount"]
    r = 4
    for (dept, comp), amt in sorted(agg.items()):
        s1.cell(r, 1).value = dept
        s1.cell(r, 2).value = comp
        s1.cell(r, 3).value = round(amt, 2)
        r += 1

    # --- Page_1: raw records with Department added ---
    s2 = wb.create_sheet("Page_1")
    page1_headers = ["Employee ID", "Employee Name", "Component Code", "Component Name", "Amount", "Department"]
    for c, h in enumerate(page1_headers, 1):
        s2.cell(1, c).value = h
    for r, rec in enumerate(records, 2):
        s2.cell(r, 1).value = rec["Employee ID"]
        s2.cell(r, 2).value = rec["Employee Name"]
        s2.cell(r, 3).value = rec["Component Code"]
        s2.cell(r, 4).value = rec["Component Name"]
        s2.cell(r, 5).value = round(rec["Amount"], 2)
        s2.cell(r, 6).value = rec["Department"]

    # --- JE: the journal entry ---
    s3 = wb.create_sheet("JE")
    s3.cell(1, 5).value = "Month"
    s3.cell(1, 6).value = f"{year}-{month:02d}"
    headers = ["Date", "Journal Entry Memo", "Subsidiary", "Currency", "Account", "Debit", "Credit", "Line Memo", "Department"]
    for c, h in enumerate(headers, 1):
        s3.cell(3, c).value = h
    for r, row in enumerate(je_rows, 4):
        s3.cell(r, 1).value = row["Date"]
        s3.cell(r, 2).value = row["Journal Entry Memo"]
        s3.cell(r, 3).value = row["Subsidiary"]
        s3.cell(r, 4).value = row["Currency"]
        s3.cell(r, 5).value = row["Account"]
        s3.cell(r, 6).value = row["Debit"]
        s3.cell(r, 7).value = row["Credit"]
        s3.cell(r, 8).value = row["Line Memo"]
        s3.cell(r, 9).value = row["Department"]

    # Totals row
    total_row = 4 + len(je_rows) + 1
    total_debit = sum((row["Debit"] or 0) for row in je_rows)
    total_credit = sum((row["Credit"] or 0) for row in je_rows)
    s3.cell(total_row, 5).value = "TOTALS"
    s3.cell(total_row, 6).value = round(total_debit, 2)
    s3.cell(total_row, 7).value = round(total_credit, 2)
    s3.cell(total_row, 8).value = f"Balance: {round(total_debit - total_credit, 2):.2f}"

    wb.save(output_path)


def main():
    if len(sys.argv) < 2:
        print("Usage: python process_netherlands_payroll.py <raw.xlsx|folder> [output.xlsx]")
        sys.exit(1)

    target = sys.argv[1]
    raw_xlsx = find_raw_file(target)
    input_dir = os.path.dirname(os.path.abspath(raw_xlsx))
    folder_name = os.path.basename(input_dir)
    year, month = parse_period_from_folder(folder_name)

    default_output = os.path.join(input_dir, f"{year}-{month:02d} Netherlands Payroll Backup.xlsx")
    output_path = sys.argv[2] if len(sys.argv) > 2 else default_output

    print(f"Netherlands Payroll — {year}-{month:02d}")
    print(f"  Input:  {raw_xlsx}")
    print(f"  Output: {output_path}")

    employee_map = load_employee_map()
    component_map = load_component_map()

    records = read_raw_records(raw_xlsx, month, employee_map)
    print(f"  Loaded {len(records)} non-zero raw records")

    buckets, unmapped = route_components(records, component_map)

    if unmapped:
        print("\n  WARNING: Unmapped components found (these were SKIPPED, JE will not balance):")
        for name, code, emp_id, amt in unmapped:
            print(f"    - {name!r} (code {code}, emp {emp_id}, amount {amt})")
        print("  Run check_mappings.py first to surface these before processing.\n")

    je_rows = build_je_rows(buckets, year, month)

    total_debit = sum((row["Debit"] or 0) for row in je_rows)
    total_credit = sum((row["Credit"] or 0) for row in je_rows)
    print(f"  Total Debit:  {total_debit:,.2f} EUR")
    print(f"  Total Credit: {total_credit:,.2f} EUR")
    print(f"  Imbalance:    {total_debit - total_credit:,.2f} EUR")

    # Detect known edge-case: Wg present without Wn (bonus-run months often miss Wn)
    components_present = {rec["Component Name"] for rec in records}
    paired_missing = []
    if "Pension premium Wg" in components_present and "Pension premium Wn" not in components_present:
        paired_missing.append(
            "  NOTE: 'Pension premium Wg' is present but 'Pension premium Wn' is missing "
            "from the raw file. This is common in bonus-run (PerRunPeriod_) exports. "
            "Pension Wn's absolute value is the employee pension contribution that should "
            "credit 231205 Netherlands Pension Payable. Manually add it to the output JE "
            "before posting, or ask the payroll provider for the complete component list."
        )

    write_output(output_path, raw_xlsx, records, buckets, je_rows, year, month)
    print(f"\n  Wrote {output_path}")

    if paired_missing:
        print()
        for msg in paired_missing:
            print(msg)

    balanced = abs(total_debit - total_credit) <= 0.01
    if not balanced:
        print("\n  STATUS: OUT OF BALANCE -- manual adjustment required before posting.")
        print("  CSV not written. Fix imbalance in Backup.xlsx and re-run.")
        sys.exit(1)
    if unmapped:
        print("\n  STATUS: Unmapped components -- JE incomplete.")
        print("  CSV not written. Add component mappings and re-run.")
        sys.exit(1)
    if paired_missing:
        print("\n  STATUS: Pension Wg/Wn pair incomplete -- manual JE adjustment required.")
        print("  CSV not written. Add Pension Wn rows to Backup.xlsx and re-run with the corrected raw, "
              "or create the CSV manually from the corrected Backup.xlsx.")
        sys.exit(1)

    ym_prefix = f"{year}-{month:02d}"
    csv_path = Path(output_path).with_name(f"{ym_prefix} Netherlands Payroll JE Import.csv")
    write_je_csv(
        rows=je_rows,
        output_path=csv_path,
        currency="EUR",
        default_external_id=make_external_id(ym_prefix, "NL"),
    )
    print(f"  Wrote {csv_path}")
    print("\n  STATUS: Balanced. Ready for review.")


if __name__ == "__main__":
    main()
