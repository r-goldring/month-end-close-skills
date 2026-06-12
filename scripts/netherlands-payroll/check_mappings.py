"""
Netherlands Payroll - Pre-flight mapping check.

Scans the raw Dutch payroll export for any new employees or new payroll
component codes/names that aren't in the mapping CSVs. Run this BEFORE
process_netherlands_payroll.py so new items can be added to the knowledge
base instead of being silently dropped.

Usage:
  python check_mappings.py <raw.xlsx|folder>

Exit codes:
  0 = all items mapped, safe to proceed
  1 = unmapped items found, review required
"""

import os
import re
import sys
import glob

import pandas as pd
import openpyxl

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, os.path.join(SCRIPT_DIR, "..", "_shared"))
from preflight_report import print_report  # noqa: E402

EMP_MAP_PATH = os.path.join(SCRIPT_DIR, "Netherlands Payroll Employee Mapping.csv")
COMP_MAP_PATH = os.path.join(SCRIPT_DIR, "Netherlands Payroll Component Mapping.csv")

# Single logical "category" for Netherlands — every component must be mapped
CATEGORY_INFO = {
    "Payroll Components": {
        "note": "every Dutch payroll component name must map to an expense and/or liability account",
        "required": True,
    },
}


def find_raw_file(target: str) -> str:
    if os.path.isfile(target):
        return target
    if os.path.isdir(target):
        matches = [
            f for f in glob.glob(os.path.join(target, "CompanyEmployeeWageComponents*.xlsx"))
            if "Backup" not in os.path.basename(f) and not os.path.basename(f).startswith("~")
        ]
        if not matches:
            raise FileNotFoundError(f"No raw NL payroll file found in {target}")
        return matches[0]
    raise FileNotFoundError(f"Not a file or folder: {target!r}")


def parse_period_from_folder(folder_name: str):
    m = re.match(r"^(\d{2})-(\d{4})$", folder_name)
    if m:
        return int(m.group(2)), int(m.group(1))
    m = re.match(r"^(\d{4})-(\d{2})$", folder_name)
    if m:
        return int(m.group(1)), int(m.group(2))
    return None, None


def scan_input(raw_xlsx: str):
    """Return (employee_ids, component_names) used in the file (any period)."""
    wb = openpyxl.load_workbook(raw_xlsx, data_only=True)
    ws = wb["Page_1"]

    employee_ids = set()
    component_names = set()
    employee_names = {}  # id -> name for nicer display

    for r in range(2, ws.max_row + 1):
        emp_id_raw = ws.cell(row=r, column=1).value
        if emp_id_raw is None:
            continue
        try:
            emp_id = int(emp_id_raw)
        except (ValueError, TypeError):
            continue
        employee_ids.add(emp_id)
        employee_names[emp_id] = ws.cell(row=r, column=2).value

        comp_name = ws.cell(row=r, column=4).value
        if comp_name:
            component_names.add(str(comp_name).strip())

    wb.close()
    return employee_ids, component_names, employee_names


def main():
    if len(sys.argv) < 2:
        print("Usage: python check_mappings.py <raw.xlsx|folder>")
        sys.exit(2)

    raw_xlsx = find_raw_file(sys.argv[1])
    input_dir = os.path.dirname(os.path.abspath(raw_xlsx))
    folder_name = os.path.basename(input_dir)

    # Load mappings
    emp_df = pd.read_csv(EMP_MAP_PATH)
    emp_df["Employee ID"] = emp_df["Employee ID"].astype(int)
    mapped_employees = set(emp_df["Employee ID"].tolist())

    comp_df = pd.read_csv(COMP_MAP_PATH)
    mapped_components = set(comp_df["Component Name"].astype(str).str.strip().tolist())

    # Scan input
    found_employees, found_components, emp_names = scan_input(raw_xlsx)

    # ─── Employee check (repurposing the cost-center slots in the shared report) ───
    # The shared report formatter is designed around "cost centers" and "codes by
    # category". For Netherlands we treat employees as the cost-center list and
    # components as a single category.
    emp_display = {eid: f"{eid} - {emp_names.get(eid) or '?'}" for eid in found_employees}
    cc_found = [emp_display[e] for e in found_employees]
    cc_mapped = [emp_display[e] for e in found_employees if e in mapped_employees]
    cc_unmapped = [emp_display[e] for e in found_employees if e not in mapped_employees]

    unmapped_components = found_components - mapped_components

    exit_code = print_report(
        country="Netherlands",
        pay_date_folder=folder_name,
        input_file=os.path.basename(raw_xlsx),
        dept_map_file=os.path.basename(EMP_MAP_PATH),
        gl_map_file=os.path.basename(COMP_MAP_PATH),
        dept_map_entries=len(emp_df),
        gl_map_entries=len(comp_df),
        cost_centers_found=sorted(cc_found),
        cost_centers_mapped=sorted(cc_mapped),
        cost_centers_unmapped=sorted(cc_unmapped),
        cost_centers_special_case=[],
        codes_by_category={"Payroll Components": found_components},
        codes_mapped=mapped_components,
        codes_unmapped_by_category={"Payroll Components": unmapped_components},
        category_info=CATEGORY_INFO,
    )

    sys.exit(exit_code)


if __name__ == "__main__":
    main()
