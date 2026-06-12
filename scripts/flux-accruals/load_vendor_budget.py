"""
Load FP&A's Vendor Budget v{date}.xlsx into a queryable structure.

The FP&A file is the source-of-truth for "what each vendor SHOULD post each month
and where." Sheet name: "Export Budget (5)". 26 columns; 221 rows (as of 4/20/2026).

Key columns we use:
  - Vendor ID         (NetSuite internal ID, primary join key)
  - Vendor Name
  - Department        (expected cost center)
  - Account           ("671100 - Software Subscriptions" -> parsed to acct_number + name)
  - Account Parent    (Software / Professional Services / COGS / Contractors / ...)
  - Jan..Dec          (per-month USD budget; constant per vendor unless Comment notes a change)
  - Contract End Date, Comment

Refresh policy: the accountant replaces the file when FP&A sends a new version. Loader
auto-discovers the latest by glob pattern.
"""

from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Iterable, List, Optional

from openpyxl import load_workbook


MONTH_COLS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
              "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]


@dataclass
class BudgetRow:
    vendor_id: int
    vendor_name: str
    department: str
    account_number: str       # "671100"
    account_name: str         # "Software Subscriptions"
    account_parent: str       # "Software"
    monthly_budget: Dict[str, float]   # {"Jan": 0.0, ..., "Dec": 0.0}
    contract_end_date: Optional[str]
    comment: str

    def budget_for(self, month_idx: int) -> float:
        """month_idx: 1..12"""
        return self.monthly_budget.get(MONTH_COLS[month_idx - 1], 0.0)

    @property
    def category(self) -> str:
        """Map Account Parent to one of: Software / Contractors / ProfFees / COGS / Other."""
        return classify_category(self.account_parent, self.account_number)


@dataclass
class VendorBudget:
    """In-memory store. Two indices for joining: vendor_id and normalized name."""
    by_vendor_id: Dict[int, BudgetRow]
    by_name_normalized: Dict[str, BudgetRow]
    rows: List[BudgetRow]

    def lookup(self, *, vendor_id: Optional[int] = None,
               vendor_name: Optional[str] = None) -> Optional[BudgetRow]:
        if vendor_id is not None and vendor_id in self.by_vendor_id:
            return self.by_vendor_id[vendor_id]
        if vendor_name:
            return self.by_name_normalized.get(_normalize_name(vendor_name))
        return None

    def all_in_category(self, category: str) -> List[BudgetRow]:
        return [r for r in self.rows if r.category == category]


def classify_category(account_parent: str, account_number: str) -> str:
    """
    Bucket each budget row into the same 4 categories the skill emits accruals for.
    Falls back to GL-account-range matching when Account Parent is missing/unusual.
    """
    parent = (account_parent or "").strip().lower()
    acct = (account_number or "").strip()

    # Account Parent first (most reliable)
    if "software" in parent:
        return "Software"
    if "contractor" in parent:
        return "Contractors"
    if "professional" in parent or "legal" in parent:
        return "ProfFees"
    if "cogs" in parent or "cost of goods" in parent or "cost of revenue" in parent:
        return "COGS"

    # GL fallback
    if acct in ("671100", "671000", "511425"):
        return "Software"
    if acct in ("511370", "611700"):
        return "Contractors"
    if acct in ("651100", "651150", "651101"):
        return "ProfFees"
    if acct.startswith("511"):
        return "COGS"

    return "Other"


def _normalize_name(name: str) -> str:
    if not name:
        return ""
    s = str(name).strip().lower()
    s = re.sub(r"[^\w\s]", "", s)        # strip punctuation
    s = re.sub(r"\s+", " ", s)            # collapse whitespace
    # Trim common corp suffixes that vary between systems
    for suffix in (" inc", " llc", " ltd", " llp", " bv", " gmbh",
                   " corp", " corporation", " co", " plc"):
        if s.endswith(suffix):
            s = s[: -len(suffix)].strip()
            break
    return s


def _parse_account_field(value) -> tuple[str, str]:
    """
    "671100 - Software Subscriptions" -> ("671100", "Software Subscriptions")
    Falls back to (raw, "") if there's no separator.
    """
    if value is None:
        return ("", "")
    s = str(value).strip()
    if not s:
        return ("", "")
    # split on " - " (FP&A standard) or " " (defensive)
    m = re.match(r"^\s*(\d{4,6})\s*[-:]?\s*(.*)$", s)
    if m:
        return (m.group(1), m.group(2).strip())
    return (s, "")


def find_latest_budget_file(flux_dir: Path) -> Path:
    """
    Find the most recent Vendor Budget*.xlsx in the Monthly Flux Analysis folder.
    Returns the highest-version file (by mtime) so a new drop wins automatically.
    """
    candidates = sorted(flux_dir.glob("Vendor Budget*.xlsx"),
                        key=lambda p: p.stat().st_mtime, reverse=True)
    if not candidates:
        raise FileNotFoundError(
            f"No Vendor Budget*.xlsx found in {flux_dir}. FP&A drops the file there."
        )
    return candidates[0]


def load_vendor_budget(path: Path) -> VendorBudget:
    """Load and index the FP&A vendor budget workbook."""
    wb = load_workbook(filename=path, read_only=True, data_only=True)
    sheet_name = next((s for s in wb.sheetnames if s.startswith("Export Budget")), wb.sheetnames[0])
    ws = wb[sheet_name]

    rows_iter = ws.iter_rows(values_only=True)
    headers = [str(h).strip() if h is not None else "" for h in next(rows_iter)]

    def col(name: str) -> int:
        try:
            return headers.index(name)
        except ValueError:
            raise KeyError(f"Column {name!r} not found in {path.name}; headers: {headers}")

    c_vid = col("Vendor ID")
    c_vname = col("Vendor Name")
    c_dept = col("Department")
    c_acct = col("Account")
    c_parent = col("Account Parent")
    c_contract_end = col("Contract End Date") if "Contract End Date" in headers else None
    c_comment = col("Comment") if "Comment" in headers else None
    c_months = {m: col(m) for m in MONTH_COLS}

    rows: List[BudgetRow] = []
    by_vid: Dict[int, BudgetRow] = {}
    by_name: Dict[str, BudgetRow] = {}

    for raw in rows_iter:
        if raw is None or all(v is None for v in raw):
            continue
        vid_raw = raw[c_vid]
        if vid_raw is None or str(vid_raw).strip() == "":
            continue
        try:
            vid = int(float(vid_raw))
        except (TypeError, ValueError):
            continue

        vname = str(raw[c_vname] or "").strip()
        dept = str(raw[c_dept] or "").strip()
        acct_num, acct_name = _parse_account_field(raw[c_acct])
        parent = str(raw[c_parent] or "").strip()
        monthly = {m: _to_float(raw[c_months[m]]) for m in MONTH_COLS}
        contract_end = _to_iso_date(raw[c_contract_end]) if c_contract_end is not None else None
        comment = str(raw[c_comment] or "").strip() if c_comment is not None else ""

        row = BudgetRow(
            vendor_id=vid, vendor_name=vname, department=dept,
            account_number=acct_num, account_name=acct_name,
            account_parent=parent, monthly_budget=monthly,
            contract_end_date=contract_end, comment=comment,
        )
        rows.append(row)
        by_vid[vid] = row
        if vname:
            by_name.setdefault(_normalize_name(vname), row)

    wb.close()
    return VendorBudget(by_vendor_id=by_vid, by_name_normalized=by_name, rows=rows)


def _to_float(v) -> float:
    if v is None or v == "":
        return 0.0
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def _to_iso_date(v) -> Optional[str]:
    if v is None or v == "":
        return None
    if hasattr(v, "isoformat"):
        return v.isoformat()[:10]
    return str(v).strip()
