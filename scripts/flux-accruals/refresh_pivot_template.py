"""
Build a per-month copy of the Flux Pivot Template populated with cached
NetSuite report data for a closing month.

Source (master, never modified by this script):
    Monthly Flux Analysis/Flux Pivot Template.xlsx

Output (per-month, regenerated each run):
    Monthly Flux Analysis/{YYYY}/{YYYY-MM}/Flux Pivot Template {YYYY-MM}.xlsx

Reuses the same _cache/ directory that build_candidates.py reads, so no extra
ns_runReport calls. Writes ALL detail-line rows (not the accrual-filtered
subset) so the accountant has the full GL detail available to drill into via pivots.

After this script runs, the accountant opens the per-month copy in Excel and clicks
Data > Refresh All to rebuild the pivot tables themselves (openpyxl can't
execute the pivot calculation; it only updates the source-range reference).

CLI:
    python refresh_pivot_template.py 2026-04 [--repo-root .]
"""

from __future__ import annotations

import argparse
import datetime as dt
import json
import os
import shutil
import sys
from pathlib import Path
from typing import List, Optional

from openpyxl import load_workbook

SCRIPT_DIR = Path(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, str(SCRIPT_DIR))

from parse_reports import (
    _parse_account_label, _parse_leaf_account, _coalesce, _to_float, _negate, _parse_date,
    _period_yyyymm, _values_to_dict, CATEGORY_BY_REPORT_ID,
)


# Sheet name -> report id mapping
DATA_SHEETS = {
    537: "IncomeStatementDetailCOGS",
    540: "GeneralLedgerdetailContra",
    542: "GeneralLedgerdetailProfes",
    721: "GeneralLedgerdetailSoftwa",
}
# Pivot table tab (visible) -> data sheet that backs it
PIVOT_TO_DATA_SHEET = {
    "COGS": "IncomeStatementDetailCOGS",
    "Contractors": "GeneralLedgerdetailContra",
    "Professional Fees": "GeneralLedgerdetailProfes",
    "Software": "GeneralLedgerdetailSoftwa",
}
HEADER_LAST_ROW = 7   # rows 1-7 are headers; data starts at row 8


def parse_args(argv=None):
    p = argparse.ArgumentParser(description="Refresh Flux Pivot Template from cache.")
    p.add_argument("period", help="Closing period as YYYY-MM (e.g., 2026-04).")
    p.add_argument("--repo-root", default=".", help="Repo root path (default: cwd).")
    return p.parse_args(argv)


def main(argv=None) -> int:
    args = parse_args(argv)
    period = args.period
    if not _is_valid_period(period):
        print(f"ERROR: invalid period {period!r}; expected YYYY-MM", file=sys.stderr)
        return 2

    repo = Path(args.repo_root).resolve()
    flux_dir = repo / "Monthly Flux Analysis"
    master_template = flux_dir / "Flux Pivot Template.xlsx"
    month_dir = flux_dir / period[:4] / period
    target_template = month_dir / f"Flux Pivot Template {period}.xlsx"
    cache_dir = month_dir / "_cache"

    if not master_template.exists():
        print(f"ERROR: master template not found: {master_template}", file=sys.stderr)
        return 2
    if not cache_dir.exists():
        print(f"ERROR: cache dir not found: {cache_dir}", file=sys.stderr)
        return 2

    period_label = _period_label(period)
    period_3mo_label = _three_month_period_label(period)

    month_dir.mkdir(parents=True, exist_ok=True)
    if target_template.exists():
        print(f"[1/3] Per-month copy exists; updating in place "
              f"(preserves any notes on pivot tabs): {target_template.relative_to(repo)}")
    else:
        print(f"[1/3] Creating per-month copy from master: {target_template.relative_to(repo)}")
        shutil.copy2(master_template, target_template)
    wb = load_workbook(target_template, keep_links=False, keep_vba=False)

    print(f"[2/3] Refreshing data sheets from {cache_dir}")
    last_data_row_by_sheet: dict[str, int] = {}
    for report_id, sheet_name in DATA_SHEETS.items():
        cache_path = cache_dir / f"{report_id}.json"
        if not cache_path.exists():
            print(f"  WARN: {cache_path.name} missing - skipping {sheet_name}")
            continue
        with open(cache_path, "r", encoding="utf-8") as f:
            payload = json.load(f)

        rows = _build_rows(report_id, payload)
        ws = wb[sheet_name]

        # Update period text in row 4 (column A)
        ws.cell(row=4, column=1, value=period_3mo_label)

        # Clear existing data rows (8 down)
        existing_max = ws.max_row
        if existing_max >= HEADER_LAST_ROW + 1:
            for r in range(HEADER_LAST_ROW + 1, existing_max + 1):
                for c in range(1, 16):
                    ws.cell(row=r, column=c).value = None

        # Write new rows starting at row 8
        for i, row_vals in enumerate(rows, start=HEADER_LAST_ROW + 1):
            for c, v in enumerate(row_vals, start=1):
                ws.cell(row=i, column=c, value=v)

        last_data_row = HEADER_LAST_ROW + len(rows)
        last_data_row_by_sheet[sheet_name] = last_data_row
        print(f"  {sheet_name}: wrote {len(rows)} data rows (range A{HEADER_LAST_ROW}:O{last_data_row})")

    # Update pivot table source ranges. Pivots live on the visible pivot tabs
    # (COGS, Contractors, Professional Fees, Software), not the hidden data tabs.
    print(f"  Updating pivot source ranges...")
    for pivot_tab, data_sheet in PIVOT_TO_DATA_SHEET.items():
        if pivot_tab not in wb.sheetnames or data_sheet not in last_data_row_by_sheet:
            continue
        ws_pivot = wb[pivot_tab]
        last_row = last_data_row_by_sheet[data_sheet]
        new_ref = f"A{HEADER_LAST_ROW}:O{last_row}"
        for pt in (getattr(ws_pivot, "_pivots", None) or []):
            try:
                pt.cache.cacheSource.worksheetSource.ref = new_ref
                pt.cache.cacheSource.worksheetSource.sheet = data_sheet
                print(f"    {pivot_tab}: pivot source -> {data_sheet}!{new_ref}")
            except Exception as e:
                print(f"    {pivot_tab}: could not update pivot source ({e})")

    print(f"[3/3] Saving: {target_template.relative_to(repo)}")
    wb.save(target_template)

    print()
    print(f"Refreshed for {period_label}.")
    print(f"Open {target_template.relative_to(repo)} in Excel and click Data > Refresh All.")
    print(f"(Excel re-reads the source range and rebuilds the pivot caches.)")
    print(f"Master template at {master_template.relative_to(repo)} is unchanged.")
    return 0


def _is_valid_period(p: str) -> bool:
    if len(p) != 7 or p[4] != "-":
        return False
    try:
        y, m = int(p[:4]), int(p[5:])
        return 1 <= m <= 12 and 2020 <= y <= 2099
    except ValueError:
        return False


def _period_label(period: str) -> str:
    y, m = int(period[:4]), int(period[5:])
    return dt.date(y, m, 1).strftime("%B %Y")


def _three_month_period_label(period: str) -> str:
    """E.g., 2026-04 -> 'Feb 2026, Mar 2026, Apr 2026'."""
    y, m = int(period[:4]), int(period[5:])
    parts = []
    for offset in (-2, -1, 0):
        mm = ((m - 1 + offset) % 12) + 1
        yy = y + (m - 1 + offset) // 12
        parts.append(dt.date(yy, mm, 1).strftime("%b %Y"))
    return ", ".join(parts)


def _build_rows(report_id: int, payload: dict) -> List[list]:
    """Parse a report payload into 15-column rows ready to write."""
    if report_id == 537:
        return _build_cogs_rows(payload)
    return _build_gl_detail_rows(payload)


def _build_cogs_rows(payload: dict) -> List[list]:
    """
    COGS data sheet schema (15 cols A-O):
      Financial Row, Type, Date, Document Number, Name, Entity (Line),
      Final Name, Clr, Split, Amount (NEGATED), Memo, Message,
      Account (Line): Name (cleaned), Department: Name, Accounting Period: Name
    """
    out: List[list] = []
    report_data = payload.get("reportData", {})
    rows_in_order = sorted(report_data.items(), key=lambda kv: int(kv[0]))
    current_section = ""

    for _, row in rows_in_order:
        if not row.get("isDetailLine", False):
            label = row.get("value") or row.get("label") or ""
            num, name = _parse_leaf_account(label)
            if num:
                current_section = f"{num} - {name}".strip(" -")
            continue

        dlv = _values_to_dict(row.get("detailLineValues", []))
        rtype = dlv.get("Type")
        if not rtype or str(rtype).strip().lower() == "none":
            continue

        # Financial Row from section header (leaf account)
        financial_row = current_section

        # Account (Line): Name -> clean \x01 to " : " for pivot display
        acct_name = (dlv.get("Account (Line): Name (GL-style)")
                     or dlv.get("Account (Line): Name")
                     or "").replace("\x01", " : ").strip()

        amount = _negate(_to_float(dlv.get("Amount")))
        date_iso = _parse_date(dlv.get("Date"))
        period = _period_yyyymm(dlv.get("Accounting Period: Name"), date_iso)
        final_name = _coalesce(dlv.get("Entity (Line)"), dlv.get("Name"))

        out.append([
            financial_row,
            rtype,
            date_iso,
            (dlv.get("Document Number") or "").strip(),
            (dlv.get("Name") or "").strip(),
            (dlv.get("Entity (Line)") or "").strip(),
            final_name,
            (dlv.get("Clr") or "").strip(),
            (dlv.get("Split") or "").strip(),
            amount,
            (dlv.get("Memo") or "").strip(),
            (dlv.get("Message") or "").strip(),
            acct_name,
            (dlv.get("Department: Name") or "").strip(),
            period,
        ])
    return out


def _build_gl_detail_rows(payload: dict) -> List[list]:
    """
    GL Detail data sheet schema (15 cols A-O):
      Account, Type, Date, Document Number, Name, Entity (Line): Name,
      Final Name, Debit, Credit, Net, Memo, Message, Department: Name,
      Subsidiary: Name, Accounting Period: Name
    """
    out: List[list] = []
    report_data = payload.get("reportData", {})
    rows_in_order = sorted(report_data.items(), key=lambda kv: int(kv[0]))

    for _, row in rows_in_order:
        if not row.get("isDetailLine", False):
            continue
        dlv = _values_to_dict(row.get("detailLineValues", []))
        rtype = dlv.get("Type")
        if not rtype or str(rtype).strip().lower() == "none":
            continue

        # Account is parent\x01leaf in GL detail; show as 'parent : leaf' for pivot.
        acct_field = (dlv.get("Account") or "").replace("\x01", " : ").strip()
        net = _to_float(dlv.get("_total") or dlv.get("") or dlv.get("Amount"))
        if net == 0:
            continue
        debit = net if net > 0 else ""
        credit = -net if net < 0 else ""
        date_iso = _parse_date(dlv.get("Date"))
        period = _period_yyyymm(dlv.get("Accounting Period: Name"), date_iso)
        final_name = _coalesce(dlv.get("Entity (Line): Name"), dlv.get("Name"))

        out.append([
            acct_field,
            dlv.get("Type") or "",
            date_iso,
            (dlv.get("Document Number") or "").strip(),
            (dlv.get("Name") or "").strip(),
            (dlv.get("Entity (Line): Name") or "").strip(),
            final_name,
            debit,
            credit,
            net,
            (dlv.get("Memo") or "").strip(),
            (dlv.get("Message") or "").strip(),
            (dlv.get("Department: Name") or "").strip(),
            (dlv.get("Subsidiary: Name") or "").strip(),
            period,
        ])
    return out


if __name__ == "__main__":
    sys.exit(main())
