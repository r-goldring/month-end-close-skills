"""
Read the the accountant-marked Accrual_Reclass_Candidates_{YYYY-MM}.xlsx and emit up to 3
NetSuite Import Assistant JE CSVs:

  {YYYY-MM} Accruals JE Import.csv
  {YYYY-MM} Dept Reclass JE Import.csv
  {YYYY-MM} Software Reclass JE Import.csv

Approval rules per row (column A):
  Y      = include using `suggested_amount` (accrual) or `expected_dept` (reclass)
  EDIT   = include using `edited_amount` / `edited_target_dept`
  N or _ = exclude

CLI:
    python generate_je_csvs.py 2026-04 [--repo-root .]
"""

from __future__ import annotations

import argparse
import calendar
import datetime as dt
import json
import os
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Iterable, List, Optional

from openpyxl import load_workbook

SCRIPT_DIR = Path(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, str(SCRIPT_DIR))
sys.path.insert(0, str(SCRIPT_DIR.parent / "_shared"))

from je_csv_writer import write_je_csv  # noqa: E402

ACCRUED_LIAB_ACCOUNT = "231100"
US_SUBSIDIARY = "Acme Holdings : Acme, Inc."
ACCRUAL_TABS = ("Accruals_Software", "Accruals_Contractors",
                "Accruals_ProfFees", "Accruals_COGS")
DEPT_RECLASS_TAB = "Reclass_Dept_Drift"
SW_RECLASS_TAB = "Reclass_Software"


@dataclass
class AccrualRow:
    vendor_name: str
    vendor_id: int
    account: str       # full string from workbook ("671100 - Software Subscriptions")
    department: str
    amount: float
    notes: str = ""


@dataclass
class ReclassRow:
    vendor_name: str
    account: str
    actual_dept: str
    target_dept: str
    amount: float
    notes: str = ""


def parse_args(argv=None):
    p = argparse.ArgumentParser(description="Emit JE Import CSVs from candidate workbook.")
    p.add_argument("period", help="Closing period as YYYY-MM (e.g., 2026-04).")
    p.add_argument("--repo-root", default=".", help="Repo root path (default: cwd).")
    return p.parse_args(argv)


def main(argv=None) -> int:
    args = parse_args(argv)
    period = args.period
    repo = Path(args.repo_root).resolve()
    flux = repo / "Monthly Flux Analysis"
    year = period[:4]
    month_dir = flux / year / period
    candidate_path = month_dir / f"Accrual_Reclass_Candidates_{period}.xlsx"

    if not candidate_path.exists():
        print(f"ERROR: candidate workbook not found: {candidate_path}\n"
              f"  Run build_candidates.py {period} first.", file=sys.stderr)
        return 2

    print(f"[1/3] Reading marked candidates: {candidate_path.name}")
    wb = load_workbook(candidate_path, data_only=True)

    accrual_rows = _collect_accrual_rows(wb)
    dept_rows = _collect_reclass_rows(wb, DEPT_RECLASS_TAB)
    sw_rows = _collect_reclass_rows(wb, SW_RECLASS_TAB)

    print(f"      Accrual rows approved:        {len(accrual_rows)}")
    print(f"      Dept reclass rows approved:   {len(dept_rows)}")
    print(f"      Software reclass approved:    {len(sw_rows)}")

    period_date = _last_day_of_month(period)
    je_label = period_date.strftime("%b-%y")  # "Apr-26"

    written: List[Path] = []

    if accrual_rows:
        out = month_dir / f"{period} Accruals JE Import.csv"
        _write_accruals_csv(accrual_rows, out, period_date, je_label, period)
        print(f"      Wrote {out.name} ({len(accrual_rows)} accruals)")
        written.append(out)

    if dept_rows:
        out = month_dir / f"{period} Dept Reclass JE Import.csv"
        _write_reclass_csv(dept_rows, out, period_date, je_label, period,
                           memo_prefix=f"{je_label} Dept Reclass",
                           external_id=f"{period}-DEPT-RECLASS")
        print(f"      Wrote {out.name} ({len(dept_rows)} dept reclasses)")
        written.append(out)

    if sw_rows:
        out = month_dir / f"{period} Software Reclass JE Import.csv"
        _write_reclass_csv(sw_rows, out, period_date, je_label, period,
                           memo_prefix=f"{je_label} Software Reclass",
                           external_id=f"{period}-SW-RECLASS")
        print(f"      Wrote {out.name} ({len(sw_rows)} software reclasses)")
        written.append(out)

    if not written:
        print("[2/3] No approved rows. Nothing to emit.")
        return 0

    print(f"[3/3] Appending GENERATE_CSV entries to audit_log.json")
    _append_audit_log(repo, period, written)

    print()
    print("Next: upload each CSV via NetSuite UI:")
    print("  Lists -> Import Assistant -> Transactions -> Journal Entry -> upload")
    print("Each JE will land in the controller's Pending Approval queue.")
    return 0


def _last_day_of_month(period: str) -> dt.date:
    y, m = int(period[:4]), int(period[5:])
    last = calendar.monthrange(y, m)[1]
    return dt.date(y, m, last)


def _collect_accrual_rows(wb) -> List[AccrualRow]:
    out: List[AccrualRow] = []
    for tab in ACCRUAL_TABS:
        if tab not in wb.sheetnames:
            continue
        ws = wb[tab]
        # Header at row 3, data row 4+
        headers = [c.value for c in ws[3]]
        idx = {h: i for i, h in enumerate(headers) if h}
        for row in ws.iter_rows(min_row=4, values_only=True):
            if row is None or all(v is None for v in row):
                continue
            approve = (row[idx["approve"]] or "").strip().upper() if isinstance(row[idx["approve"]], str) else ""
            if approve not in ("Y", "EDIT"):
                continue
            suggested = _to_float(row[idx["suggested_amount"]])
            edited = _to_float(row[idx["edited_amount"]])
            amount = edited if approve == "EDIT" and edited > 0 else suggested
            if amount <= 0:
                continue
            out.append(AccrualRow(
                vendor_name=str(row[idx["vendor_name"]] or "").strip(),
                vendor_id=int(row[idx["vendor_id"]] or 0),
                account=str(row[idx["account"]] or "").strip(),
                department=str(row[idx["dept"]] or "").strip(),
                amount=round(amount, 2),
                notes=str(row[idx["notes"]] or "").strip(),
            ))
    return out


def _collect_reclass_rows(wb, tab: str) -> List[ReclassRow]:
    if tab not in wb.sheetnames:
        return []
    ws = wb[tab]
    headers = [c.value for c in ws[4]]
    idx = {h: i for i, h in enumerate(headers) if h}
    out: List[ReclassRow] = []
    for row in ws.iter_rows(min_row=5, values_only=True):
        if row is None or all(v is None for v in row):
            continue
        approve = (row[idx["approve"]] or "").strip().upper() if isinstance(row[idx["approve"]], str) else ""
        if approve not in ("Y", "EDIT"):
            continue
        amount = _to_float(row[idx["actual_amount"]])
        if amount <= 0:
            continue
        target = ""
        if approve == "EDIT":
            target = str(row[idx["edited_target_dept"]] or "").strip()
        if not target:
            target = str(row[idx["expected_dept"]] or "").strip()
        if not target:
            continue
        out.append(ReclassRow(
            vendor_name=str(row[idx["vendor"]] or "").strip(),
            account=str(row[idx["account"]] or "").strip(),
            actual_dept=str(row[idx["actual_dept"]] or "").strip(),
            target_dept=target,
            amount=round(amount, 2),
            notes=str(row[idx.get("notes", -1)] or "").strip() if "notes" in idx else "",
        ))
    return out


def _write_accruals_csv(rows: List[AccrualRow], out_path: Path,
                        period_date: dt.date, je_label: str, period: str) -> None:
    je_memo = f"{je_label} Month-End Accruals"
    csv_rows = []
    for r in rows:
        line_memo = f"{r.vendor_name} - {je_label} accrual"
        csv_rows.append({
            "Date": period_date,
            "Journal Entry Memo": je_memo,
            "Account": r.account,
            "Debit": r.amount,
            "Credit": "",
            "Line Memo": _safe_ascii(line_memo),
            "Subsidiary": US_SUBSIDIARY,
            "Department": r.department,
        })
        csv_rows.append({
            "Date": period_date,
            "Journal Entry Memo": je_memo,
            "Account": ACCRUED_LIAB_ACCOUNT,
            "Debit": "",
            "Credit": r.amount,
            "Line Memo": _safe_ascii(line_memo),
            "Subsidiary": US_SUBSIDIARY,
            "Department": "",
        })
    write_je_csv(csv_rows, out_path, currency="USD",
                 default_external_id=f"{period}-ACCRUALS")


def _write_reclass_csv(rows: List[ReclassRow], out_path: Path,
                       period_date: dt.date, je_label: str, period: str,
                       memo_prefix: str, external_id: str) -> None:
    csv_rows = []
    for r in rows:
        line_memo = f"{r.vendor_name} - {je_label} reclass {r.actual_dept} to {r.target_dept}"
        csv_rows.append({
            "Date": period_date,
            "Journal Entry Memo": memo_prefix,
            "Account": r.account,
            "Debit": "",
            "Credit": r.amount,
            "Line Memo": _safe_ascii(line_memo),
            "Subsidiary": US_SUBSIDIARY,
            "Department": r.actual_dept,
        })
        csv_rows.append({
            "Date": period_date,
            "Journal Entry Memo": memo_prefix,
            "Account": r.account,
            "Debit": r.amount,
            "Credit": "",
            "Line Memo": _safe_ascii(line_memo),
            "Subsidiary": US_SUBSIDIARY,
            "Department": r.target_dept,
        })
    write_je_csv(csv_rows, out_path, currency="USD",
                 default_external_id=external_id)


def _safe_ascii(s: str) -> str:
    """Strip any non-ASCII (em-dash, curly quotes, etc.) from a memo string."""
    return s.encode("ascii", errors="replace").decode("ascii").replace("?", "")


def _to_float(v) -> float:
    if v is None or v == "":
        return 0.0
    try:
        return float(v)
    except (TypeError, ValueError):
        s = str(v).strip().replace(",", "").replace("$", "")
        if s.startswith("(") and s.endswith(")"):
            s = "-" + s[1:-1]
        try:
            return float(s)
        except (TypeError, ValueError):
            return 0.0


def _append_audit_log(repo: Path, period: str, csv_paths: List[Path]) -> None:
    log_path = repo / "audit_log.json"
    if log_path.exists():
        with open(log_path, "r", encoding="utf-8") as f:
            log = json.load(f)
    else:
        log = []

    ts = dt.datetime.now().strftime("%Y-%m-%dT%H:%M:%S")
    for csv_path in csv_paths:
        log.append({
            "timestamp": ts,
            "skill": "flux-accruals",
            "action": "GENERATE_CSV",
            "description": f"{period} {csv_path.stem}",
            "csv_path": str(csv_path.relative_to(repo)),
        })

    with open(log_path, "w", encoding="utf-8") as f:
        json.dump(log, f, indent=2)


if __name__ == "__main__":
    sys.exit(main())
