"""
Build the Accrual_Reclass_Candidates_{YYYY-MM}.xlsx workbook.

Tabs:
  Summary             - dashboard + instructions
  Accruals_Software
  Accruals_Contractors
  Accruals_ProfFees
  Accruals_COGS
  Reclass_Dept_Drift
  Reclass_Software
  Unmatched_Actuals   - vendors in actuals but not in budget
  Billcom_AlreadyInNS - validation log (debugging the accountant's export filter)
  Billcom_Unmatched   - bill.com vendors that didn't match any budget vendor

Column A is a Y/N/EDIT dropdown so the accountant reviews in Excel and we re-read on
generate_je_csvs.py.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF")
SUMMARY_FILL = PatternFill("solid", fgColor="DDEBF7")
INSTRUCTION_FILL = PatternFill("solid", fgColor="FFF2CC")
SIGNAL_FILL = {
    "missing_budgeted":      PatternFill("solid", fgColor="FCE4D6"),  # peach
    "billcom_confirmed":     PatternFill("solid", fgColor="C6EFCE"),  # green
    "partial_below_baseline":PatternFill("solid", fgColor="FFEB9C"),  # yellow
    "over_budget_review":    PatternFill("solid", fgColor="F4CCCC"),  # red-pink
    "unbudgeted_vendor":     PatternFill("solid", fgColor="EAD1DC"),  # lavender
}
THIN_BORDER = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)


@dataclass
class AccrualCandidate:
    vendor_name: str
    vendor_id: int
    account: str            # "671100 - Software Subscriptions"
    department: str
    budget_amount: float
    actual_amount: float
    gap: float
    trailing_3mo: List[float]  # [M-2, M-1] amounts (3-month pull window)
    signal: str
    suggested_amount: float
    confidence: str         # HIGH / MED / LOW
    notes: str = ""


@dataclass
class ReclassCandidate:
    vendor_name: str
    account: str
    actual_dept: str
    expected_dept: str
    actual_amount: float
    reason: str
    notes: str = ""


@dataclass
class CandidateBundle:
    period: str   # "2026-04"
    period_label: str  # "April 2026"
    accruals: Dict[str, List[AccrualCandidate]]   # category -> rows
    dept_drift: List[ReclassCandidate]
    software_reclass: List[ReclassCandidate]
    unmatched_actuals: List[dict]                  # raw rows
    billcom_already_in_ns: List[dict]              # raw rows
    billcom_unmatched: List[dict]                  # raw rows


ACCRUAL_HEADERS = [
    "approve", "vendor_name", "vendor_id", "account", "dept",
    "budget", "actual", "gap",
    "M-2", "M-1",
    "signal", "suggested_amount", "edited_amount", "confidence", "notes",
]
RECLASS_HEADERS = [
    "approve", "vendor", "account", "actual_dept", "expected_dept",
    "actual_amount", "reason", "edited_target_dept", "notes",
]


def write_workbook(bundle: CandidateBundle, out_path: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    wb.remove(ws)

    _write_summary(wb, bundle)
    for cat in ("Software", "Contractors", "ProfFees", "COGS"):
        _write_accrual_tab(wb, f"Accruals_{cat}", bundle.accruals.get(cat, []), bundle.period_label)
    _write_reclass_tab(wb, "Reclass_Dept_Drift", bundle.dept_drift, bundle.period_label,
                       intro="Vendors with current-month activity in a department other than the FP&A budget. "
                             "Excludes software-account amortizations (see Reclass_Software).")
    _write_reclass_tab(wb, "Reclass_Software", bundle.software_reclass, bundle.period_label,
                       intro="Software accounts (671xxx + 511425) where actual dept != budget dept. "
                             "Most of these are monthly prepaid amortizations landing in a default dept.")
    _write_audit_tab(wb, "Unmatched_Actuals", bundle.unmatched_actuals,
                     ["category", "vendor", "account", "department", "actual_amount"],
                     intro="Vendors that posted actuals this period but are NOT in the FP&A Vendor Budget. "
                           "Send to FP&A to update the budget file or confirm they're truly out-of-scope.")
    _write_audit_tab(wb, "Billcom_AlreadyInNS", bundle.billcom_already_in_ns,
                     ["vendor", "invoice_number", "amount", "netsuite_tranid", "approval_status"],
                     intro="BillFlow export rows that DO already exist in NetSuite. If this tab has many "
                           "rows, the export filter ('not yet exported') is misbehaving.")
    _write_audit_tab(wb, "Billcom_Unmatched", bundle.billcom_unmatched,
                     ["vendor", "invoice_number", "amount", "approval_status"],
                     intro="Open BillFlow bills whose vendor name didn't match any vendor in the FP&A "
                           "Vendor Budget (rapidfuzz < 85). Manually reconcile.")

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path


def _write_summary(wb: Workbook, bundle: CandidateBundle) -> None:
    ws = wb.create_sheet("Summary")
    ws.sheet_properties.tabColor = "1F4E78"

    ws["A1"] = f"Accrual & Reclass Candidates - {bundle.period_label}"
    ws["A1"].font = Font(bold=True, size=16)
    ws.merge_cells("A1:F1")

    instructions = [
        "REVIEW WORKFLOW",
        "1. Open this workbook in Excel.",
        "2. (Recommended) Open Flux Pivot Template.xlsx side-by-side; click Data > Refresh All.",
        "3. On each Accruals tab, mark column A:",
        "       Y      = approve as suggested",
        "       N      = reject (do not accrue)",
        "       EDIT   = override; type your number into the 'edited_amount' column",
        "4. Same flow on Reclass_Dept_Drift and Reclass_Software (use 'edited_target_dept' for overrides).",
        "5. Save the workbook (keep this filename).",
        "6. Reply 'build' so the skill emits up to 3 JE Import CSVs from the Y/EDIT rows.",
        "",
        "SIGNALS",
        "  missing_budgeted        = budget > 0 and zero actuals this month",
        "  partial_below_baseline  = actuals < 50% of budget AND gap > category threshold",
        "  billcom_confirmed       = BillFlow has an open bill for this vendor (HIGH confidence)",
        "  over_budget_review      = actuals > 120% of budget (informational, NOT an accrual)",
        "  unbudgeted_vendor       = posted actuals but no FP&A budget row (informational)",
    ]
    for i, line in enumerate(instructions, start=3):
        ws.cell(row=i, column=1, value=line)
        if i == 3 or line.startswith("SIGNALS"):
            ws.cell(row=i, column=1).font = Font(bold=True)
        ws.cell(row=i, column=1).fill = INSTRUCTION_FILL

    # Counts table
    start_row = 3 + len(instructions) + 2
    ws.cell(row=start_row, column=1, value="CATEGORY").font = Font(bold=True)
    ws.cell(row=start_row, column=2, value="# Candidates").font = Font(bold=True)
    ws.cell(row=start_row, column=3, value="Total Suggested").font = Font(bold=True)
    for c in range(1, 4):
        ws.cell(row=start_row, column=c).fill = HEADER_FILL
        ws.cell(row=start_row, column=c).font = Font(bold=True, color="FFFFFF")

    r = start_row + 1
    grand_count, grand_total = 0, 0.0
    for cat in ("Software", "Contractors", "ProfFees", "COGS"):
        rows = bundle.accruals.get(cat, [])
        approve_count = sum(1 for x in rows if x.signal not in ("over_budget_review", "unbudgeted_vendor"))
        total = sum(x.suggested_amount for x in rows
                    if x.signal not in ("over_budget_review", "unbudgeted_vendor"))
        ws.cell(row=r, column=1, value=cat)
        ws.cell(row=r, column=2, value=approve_count)
        ws.cell(row=r, column=3, value=total).number_format = '"$"#,##0.00'
        grand_count += approve_count
        grand_total += total
        r += 1
    ws.cell(row=r, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=r, column=2, value=grand_count).font = Font(bold=True)
    cell = ws.cell(row=r, column=3, value=grand_total)
    cell.number_format = '"$"#,##0.00'
    cell.font = Font(bold=True)

    r += 2
    ws.cell(row=r, column=1, value="Reclass_Dept_Drift candidates").font = Font(bold=True)
    ws.cell(row=r, column=2, value=len(bundle.dept_drift))
    r += 1
    ws.cell(row=r, column=1, value="Reclass_Software candidates").font = Font(bold=True)
    ws.cell(row=r, column=2, value=len(bundle.software_reclass))

    ws.column_dimensions["A"].width = 80
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 22


def _write_accrual_tab(wb: Workbook, sheet_name: str,
                      candidates: List[AccrualCandidate], period_label: str) -> None:
    ws = wb.create_sheet(sheet_name)
    ws.cell(row=1, column=1, value=f"{sheet_name} - {period_label}").font = Font(bold=True, size=12)

    for c, h in enumerate(ACCRUAL_HEADERS, start=1):
        cell = ws.cell(row=3, column=c, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    if not candidates:
        ws.cell(row=4, column=1, value="(no candidates this month)").font = Font(italic=True, color="808080")
    else:
        for i, cand in enumerate(candidates, start=4):
            tr = list(cand.trailing_3mo) + [0.0] * (2 - len(cand.trailing_3mo))
            row_vals = [
                "",                                # approve (the accountant fills)
                cand.vendor_name,
                cand.vendor_id,
                cand.account,
                cand.department,
                cand.budget_amount,
                cand.actual_amount,
                cand.gap,
                tr[0], tr[1],
                cand.signal,
                cand.suggested_amount,
                "",                                # edited_amount (the accountant fills)
                cand.confidence,
                cand.notes,
            ]
            for c, v in enumerate(row_vals, start=1):
                cell = ws.cell(row=i, column=c, value=v)
                cell.border = THIN_BORDER
                if c in (6, 7, 8, 9, 10, 12, 13):
                    cell.number_format = '"$"#,##0.00'
            # Highlight the suggested amount column based on signal
            sig_fill = SIGNAL_FILL.get(cand.signal)
            if sig_fill:
                ws.cell(row=i, column=11).fill = sig_fill

    # Y/N/EDIT validation on column A starting row 4
    dv = DataValidation(type="list", formula1='"Y,N,EDIT"', allow_blank=True)
    dv.add(f"A4:A{max(4, 3 + len(candidates))}")
    ws.add_data_validation(dv)

    _apply_widths(ws, {
        "A": 9, "B": 36, "C": 11, "D": 36, "E": 22,
        "F": 13, "G": 13, "H": 13, "I": 12, "J": 12,
        "K": 22, "L": 16, "M": 14, "N": 12, "O": 40,
    })
    ws.freeze_panes = "B4"


def _write_reclass_tab(wb: Workbook, sheet_name: str,
                       candidates: List[ReclassCandidate], period_label: str,
                       intro: str) -> None:
    ws = wb.create_sheet(sheet_name)
    ws.cell(row=1, column=1, value=f"{sheet_name} - {period_label}").font = Font(bold=True, size=12)
    ws.cell(row=2, column=1, value=intro).fill = INSTRUCTION_FILL
    ws.merge_cells("A2:I2")

    for c, h in enumerate(RECLASS_HEADERS, start=1):
        cell = ws.cell(row=4, column=c, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    if not candidates:
        ws.cell(row=5, column=1, value="(no candidates this month)").font = Font(italic=True, color="808080")
    else:
        for i, cand in enumerate(candidates, start=5):
            row_vals = [
                "",  # approve
                cand.vendor_name,
                cand.account,
                cand.actual_dept,
                cand.expected_dept,
                cand.actual_amount,
                cand.reason,
                "",  # edited_target_dept
                cand.notes,
            ]
            for c, v in enumerate(row_vals, start=1):
                cell = ws.cell(row=i, column=c, value=v)
                cell.border = THIN_BORDER
                if c == 6:
                    cell.number_format = '"$"#,##0.00'

    dv = DataValidation(type="list", formula1='"Y,N,EDIT"', allow_blank=True)
    dv.add(f"A5:A{max(5, 4 + len(candidates))}")
    ws.add_data_validation(dv)

    _apply_widths(ws, {
        "A": 9, "B": 32, "C": 36, "D": 22, "E": 22,
        "F": 14, "G": 32, "H": 22, "I": 36,
    })
    ws.freeze_panes = "B5"


def _write_audit_tab(wb: Workbook, sheet_name: str, rows: List[dict],
                     columns: List[str], intro: str) -> None:
    ws = wb.create_sheet(sheet_name)
    ws.cell(row=1, column=1, value=sheet_name).font = Font(bold=True, size=12)
    ws.cell(row=2, column=1, value=intro).fill = INSTRUCTION_FILL
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max(len(columns), 4))

    for c, h in enumerate(columns, start=1):
        cell = ws.cell(row=4, column=c, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    if not rows:
        ws.cell(row=5, column=1, value="(empty - good)").font = Font(italic=True, color="808080")
    else:
        for i, row in enumerate(rows, start=5):
            for c, key in enumerate(columns, start=1):
                cell = ws.cell(row=i, column=c, value=row.get(key, ""))
                if key in ("amount", "actual_amount"):
                    cell.number_format = '"$"#,##0.00'

    widths = {get_column_letter(i + 1): 22 for i in range(len(columns))}
    widths[get_column_letter(1)] = 32
    _apply_widths(ws, widths)


def _apply_widths(ws, widths: Dict[str, int]) -> None:
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
