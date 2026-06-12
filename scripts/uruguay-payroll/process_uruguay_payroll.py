"""
Uruguay Payroll -> NetSuite JE generator.

Reads 1-N raw provider files (SGN_892_*) from a payroll month folder. Each
raw file is one payrun. For each payrun, generates TWO balanced NetSuite JEs:

  Block 1 (Main Payroll):
    DR Salaries & Wages (per dept, COGS/OpEx split)
    DR Payroll Taxes (per dept, COGS/OpEx split)
    DR Special Bonus (e.g., MIP - flagged for review at runtime)
    CR 231200 Payroll Liability (Net Pay)
    CR 231350 Payroll Tax Liability (balance)

  Block 2 (Aguinaldo Accrual):
    DR Bonus expense (per dept, COGS/OpEx split) - from MONTHLY BONUS PROVISION
    DR Bonus PR Tax expense (per dept, COGS/OpEx) - LICENSE CS + SS BONUS provisions
    CR 231207 Accrued PTO
    CR 231171 Accrued Bonus Payroll Tax Liability

Multi-payrun classification (from filename):
  - Mensual / MARZO / FEBRERO / ABRIL / etc. -> 'regular'
  - Espinoza / Pando / similar -> 'offcycle'
  - Liquidacion*ExtraRun -> 'extra-run' (e.g., new hire)
  - Egreso* -> 'egreso' (severance) - routed to 511175/611250 + EBITDA Adjustments

Infrastructure dept (cc=3) and DevOps (cc=4) employees split
COGS/OpEx (default 30/70 from Apr 2026; use --infra-cogs-pct=20 for Jan-Mar regression).

Usage:
    python process_uruguay_payroll.py <folder_path> [--infra-cogs-pct=30]

Folder path is the YYYY-MM folder. Script auto-discovers all SGN_892_*.xls* files.
"""

import os
import re
import sys
import glob
import calendar
import datetime as dt
import argparse
from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, os.path.join(SCRIPT_DIR, "..", "_shared"))
from je_csv_writer import write_je_csv, make_external_id  # noqa: E402

try:
    sys.stdout.reconfigure(encoding="utf-8")
except Exception:
    pass

SUBSIDIARY = "Acme Holdings : Acme, Inc. : Acme Uruguay"
LIAB_NET_ACCT = "231200 Payroll Liability"
LIAB_TAX_ACCT = "231350 Payroll Tax Liability"
ACCRUED_PTO_ACCT = "231207 Accrued PTO"
ACCRUED_BONUS_TAX_ACCT = "231171 Accrued Bonus Payroll Tax Liability"

SEVERANCE_DEPT = "EBITDA Adjustments"
SEVERANCE_COGS_ACCT = "511175 COGS - Salary and Compensation : COGS - Severance"
SEVERANCE_OPEX_ACCT = "611250 Severance"


def _to_float(v):
    if pd.isna(v):
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", "").replace("$", "")
    if not s:
        return 0.0
    try:
        return float(s)
    except ValueError:
        return 0.0


def _last_day_of_month(d: dt.date) -> dt.date:
    return d.replace(day=calendar.monthrange(d.year, d.month)[1])


def _classify_payrun(filename):
    """Return (payrun_type, payrun_label) from filename."""
    f_orig = os.path.basename(filename)
    f = f_orig.lower()
    if "egreso" in f:
        # Extract employee name after 'Egreso'
        m = re.search(r"egreso([a-z]+)_", f_orig, re.IGNORECASE)
        return ("egreso", m.group(1) if m else "Egreso")
    # Specific-employee off-cycle (correction or one-off)
    if any(k in f for k in ["pando", "espinoza"]):
        return ("offcycle", "Off-Cycle")
    # New hire / liquidation extra-runs.
    # Match "extrarun" substring specifically (not just "liquidacion"), because the
    # provider sometimes names regular monthly files "LiquidacionMonthName" too
    # (e.g., LiquidacionMayo2026). Only "ExtraRun" (or "extra run"/"extra_run")
    # in the filename signals an actual extra-run.
    if "extrarun" in f or "extra run" in f or "extra_run" in f:
        return ("extra-run", "Extra Run")
    # Regular monthly (Spanish month name or "Mensual"). Catches BOTH
    # "SGN_892_ABRIL_..." style AND "LiquidacionMayo2026_..." style.
    if "mensual" in f or any(m in f for m in ["enero", "febrero", "marzo", "abril", "mayo", "junio",
                                                "julio", "agosto", "septiembre", "octubre", "noviembre", "diciembre"]):
        return ("regular", "")
    return ("regular", "")


def _load_employee_map(csv_path):
    df = pd.read_csv(csv_path)
    m = {}
    for _, r in df.iterrows():
        eid = int(r["Employee ID"])
        m[eid] = {
            "name": str(r["Employee Name"]).strip(),
            "dept": str(r["Department"]).strip(),
            "cogs_pct": float(r["COGS Split %"]),
        }
    return m


def _load_gl_map(csv_path):
    df = pd.read_csv(csv_path).fillna("")
    m = {}
    for _, r in df.iterrows():
        try:
            code = int(r["Pay Code"])
        except ValueError:
            continue
        m[code] = {
            "role": str(r["Role"]).strip(),
            "cogs": str(r["GL Account - COGS"]).strip(),
            "opex": str(r["GL Account - OpEx"]).strip(),
            "memo": str(r["Line Memo"]).strip(),
            "desc": str(r["Description"]).strip(),
        }
    return m


def _read_raw(raw_file):
    """Read SGN sheet, returning DataFrame and the metadata dict."""
    df = pd.read_excel(raw_file, sheet_name="SGN", header=None)
    # Find pay code row (col 0 == 'Pay code')
    code_row = None
    for r in range(min(15, len(df))):
        v = df.iloc[r, 0]
        if pd.notna(v) and str(v).strip() == "Pay code":
            code_row = r
            break
    if code_row is None:
        raise RuntimeError(f"Could not find 'Pay code' header row in {raw_file}")

    header_row = code_row - 1  # Spanish/English header is one row above pay code
    while header_row > 0 and df.iloc[header_row, 0] != "Employee ID":
        header_row -= 1
    if df.iloc[header_row, 0] != "Employee ID":
        raise RuntimeError(f"Could not find 'Employee ID' header row in {raw_file}")

    # Build {pay_code: col_idx}
    code_to_col = {}
    for c in range(6, df.shape[1]):
        v = df.iloc[code_row, c]
        if pd.notna(v):
            try:
                code_to_col[int(v)] = c
            except (ValueError, TypeError):
                pass

    # Pay cycle dates from row 7 (e.g., '2026/01/01-2026/01/31')
    pay_cycle = None
    for r in range(min(10, len(df))):
        v0 = df.iloc[r, 0]
        v1 = df.iloc[r, 1]
        if pd.notna(v0) and str(v0).strip() == "Pay Cycle Dates" and pd.notna(v1):
            pay_cycle = str(v1)
            break

    return df, {"header_row": header_row, "code_row": code_row,
                "data_start_row": code_row + 1,
                "code_to_col": code_to_col, "pay_cycle": pay_cycle}


def _emit_split_lines(amount, dept, cogs_pct, cogs_acct, opex_acct, memo, je_date, je_memo, lines):
    """Emit DR lines split between COGS and OpEx accounts based on cogs_pct."""
    if amount == 0.0:
        return
    cogs_amt = round(amount * cogs_pct / 100, 2)
    opex_amt = round(amount - cogs_amt, 2)
    if cogs_amt != 0 and cogs_acct:
        lines.append({
            "Date": je_date.strftime("%Y-%m-%d"), "Journal Entry Memo": je_memo,
            "Account": cogs_acct, "Debit": cogs_amt, "Credit": "",
            "Line Memo": memo, "Subsidiary": SUBSIDIARY, "Department": dept,
        })
    if opex_amt != 0 and opex_acct:
        lines.append({
            "Date": je_date.strftime("%Y-%m-%d"), "Journal Entry Memo": je_memo,
            "Account": opex_acct, "Debit": opex_amt, "Credit": "",
            "Line Memo": memo, "Subsidiary": SUBSIDIARY, "Department": dept,
        })


def process_payrun(raw_file, emp_map, gl_map, infra_cogs_pct, ym_prefix):
    """Process one raw payrun file, return (block1_lines, block2_lines, payrun_type, summary)."""
    payrun_type, payrun_label = _classify_payrun(raw_file)
    df, meta = _read_raw(raw_file)
    code_to_col = meta["code_to_col"]
    header_row = meta["header_row"]

    # Validate every pay code in the raw is in our GL mapping
    unknown = [c for c in code_to_col if c not in gl_map]
    if unknown:
        raise RuntimeError(
            f"Unknown pay codes in {os.path.basename(raw_file)}: {unknown}. "
            "Add rows to Uruguay Payroll GL Mapping.csv and re-run."
        )

    # Determine JE date: last day of pay cycle's month
    if meta["pay_cycle"]:
        # e.g. '2026/01/01-2026/01/31'
        end_date_str = meta["pay_cycle"].split("-")[1].strip()
        end_date = dt.datetime.strptime(end_date_str, "%Y/%m/%d").date()
    else:
        end_date = dt.date.today()
    je_date = _last_day_of_month(end_date)

    # JE memo per payrun type
    if payrun_type == "regular":
        je_memo_main = f"{ym_prefix} Uruguay Payroll"
        je_memo_accrual = f"{ym_prefix} Uruguay Payroll - Aguinaldo Accrual"
    elif payrun_type == "offcycle":
        je_memo_main = f"{ym_prefix} Uruguay Off-Cycle Payroll"
        je_memo_accrual = f"{ym_prefix} Uruguay Off-Cycle Payroll - Aguinaldo Accrual"
    elif payrun_type == "extra-run":
        je_memo_main = f"{ym_prefix} Uruguay Payroll - Extra Run"
        je_memo_accrual = f"{ym_prefix} Uruguay Payroll - Extra Run Aguinaldo Accrual"
    elif payrun_type == "egreso":
        je_memo_main = f"{ym_prefix} Uruguay Payroll - Egreso ({payrun_label})"
        je_memo_accrual = f"{ym_prefix} Uruguay Payroll - Egreso Aguinaldo Accrual"
    else:
        je_memo_main = f"{ym_prefix} Uruguay Payroll"
        je_memo_accrual = f"{ym_prefix} Uruguay Payroll - Aguinaldo Accrual"

    block1_lines = []
    block2_lines = []
    flagged_special = []  # for SpecialBonus (MIP) review

    # Iterate employee rows
    net_pay_total = 0.0
    medical_coverage_total = 0.0
    medical_deduction_total = 0.0

    employee_rows = []
    data_start = meta["data_start_row"]
    for r in range(data_start, len(df)):
        eid = df.iloc[r, 0]
        if not isinstance(eid, (int, float)) or pd.isna(eid):
            # Could be 'Company total' row or blank — stop on first non-numeric ID
            break
        employee_rows.append(r)

    if not employee_rows:
        raise RuntimeError(f"No employee rows in {raw_file}")

    # Validate every employee is in mapping
    unknown_emps = []
    for r in employee_rows:
        eid = int(df.iloc[r, 0])
        if eid not in emp_map:
            unknown_emps.append((eid, str(df.iloc[r, 1]).strip()))
    if unknown_emps:
        raise RuntimeError(
            f"Unknown employees in {os.path.basename(raw_file)} (not in Uruguay Payroll Employee Mapping.csv): "
            + "; ".join(f"{eid} {n}" for eid, n in unknown_emps)
        )

    # Per-employee processing
    for r in employee_rows:
        eid = int(df.iloc[r, 0])
        emp = emp_map[eid]
        dept = emp["dept"] if payrun_type != "egreso" else SEVERANCE_DEPT
        cogs_pct = emp["cogs_pct"]
        # Override cogs_pct for Infrastructure with the runtime flag
        if "Infrastructure" in dept and cogs_pct != 0:
            cogs_pct = infra_cogs_pct

        # Get values for each pay code
        def val(code):
            c = code_to_col.get(code)
            if c is None:
                return 0.0
            return _to_float(df.iloc[r, c])

        gross_pay = val(2000)              # Total Gross Pay
        medical = val(67)                   # MEDICAL COVERAGE (offsets via 667)
        medical_ded = val(667)              # MEDICAL COVERAGE DEDUCTION
        net = val(2030)
        special_bonus = val(50)             # MIP/BONO

        # ER taxes (Block 1)
        er_taxes = val(7001) + val(7002) + val(7003) + val(7004) + val(7007)

        # Aguinaldo accrual (Block 2)
        aguinaldo = val(6300)
        aguinaldo_pr_tax = val(6303) + val(6304)

        # DR Salaries & Wages = Gross - Medical Coverage - Special Bonus
        # (Medical Coverage offset by 667 deduction; Special Bonus emitted as separate line)
        salary_dr = gross_pay - medical - special_bonus
        if salary_dr != 0:
            if payrun_type == "egreso":
                cogs_acct = SEVERANCE_COGS_ACCT
                opex_acct = SEVERANCE_OPEX_ACCT
                memo = je_memo_main
            else:
                cogs_acct = "511100 COGS - Salary and Compensation : COGS - Salaries and Wages"
                opex_acct = "611100 Salary and Compensation : Salaries and Wages"
                memo = f"{ym_prefix} Uruguay Payroll - Salaries & Wages"
                if payrun_type != "regular":
                    memo = f"{je_memo_main} - Salaries & Wages"
            _emit_split_lines(salary_dr, dept, cogs_pct, cogs_acct, opex_acct, memo,
                              je_date, je_memo_main, block1_lines)

        # DR Special Bonus (MIP) — flag for review
        if special_bonus != 0:
            flagged_special.append((eid, emp["name"], special_bonus, dept))
            # Default to standard Bonus accounts; line memo includes "Special Bonus"
            cogs_acct = "511150 COGS - Salary and Compensation : COGS - Bonus"
            opex_acct = "611150 Salary and Compensation : Bonus"
            memo = f"{ym_prefix} Uruguay Payroll - Special Bonus"
            _emit_split_lines(special_bonus, dept, cogs_pct, cogs_acct, opex_acct, memo,
                              je_date, je_memo_main, block1_lines)

        # DR Payroll Taxes (Block 1)
        if er_taxes != 0:
            cogs_acct = "511350 COGS - Salary and Compensation : COGS - Payroll Taxes"
            opex_acct = "611450 Salary and Compensation : Payroll Taxes"
            if payrun_type == "regular":
                memo = f"{ym_prefix} Uruguay Payroll Taxes"
            else:
                memo = f"{je_memo_main} Taxes"
            _emit_split_lines(er_taxes, dept, cogs_pct, cogs_acct, opex_acct, memo,
                              je_date, je_memo_main, block1_lines)

        # Block 2: Aguinaldo Accrual (skip for egreso)
        if payrun_type != "egreso":
            if aguinaldo != 0:
                cogs_acct = "511150 COGS - Salary and Compensation : COGS - Bonus"
                opex_acct = "611150 Salary and Compensation : Bonus"
                memo = f"{ym_prefix} Uruguay Payroll - Aguinaldo"
                if payrun_type != "regular":
                    memo = f"{je_memo_main} - Aguinaldo"
                _emit_split_lines(aguinaldo, dept, cogs_pct, cogs_acct, opex_acct, memo,
                                  je_date, je_memo_accrual, block2_lines)
            if aguinaldo_pr_tax != 0:
                cogs_acct = "511350 COGS - Salary and Compensation : COGS - Payroll Taxes"
                opex_acct = "611450 Salary and Compensation : Payroll Taxes"
                memo = f"{ym_prefix} Uruguay Payroll - Aguinaldo PR Tax"
                if payrun_type != "regular":
                    memo = f"{je_memo_main} - Aguinaldo PR Tax"
                _emit_split_lines(aguinaldo_pr_tax, dept, cogs_pct, cogs_acct, opex_acct, memo,
                                  je_date, je_memo_accrual, block2_lines)

        net_pay_total += net
        medical_coverage_total += medical
        medical_deduction_total += medical_ded

    # Block 1 CR: Net Pay + Tax Liability (balance)
    if net_pay_total != 0:
        net_memo = f"{ym_prefix} Uruguay Payroll Liability" if payrun_type == "regular" else f"{je_memo_main} Liability"
        block1_lines.append({
            "Date": je_date.strftime("%Y-%m-%d"), "Journal Entry Memo": je_memo_main,
            "Account": LIAB_NET_ACCT, "Debit": "", "Credit": round(net_pay_total, 2),
            "Line Memo": net_memo, "Subsidiary": SUBSIDIARY, "Department": "",
        })

    # Compute Block 1 DR total to derive Tax Liability
    block1_dr = sum(float(l["Debit"]) for l in block1_lines if l["Debit"] != "")
    block1_cr = sum(float(l["Credit"]) for l in block1_lines if l["Credit"] != "")
    tax_liab = round(block1_dr - block1_cr, 2)
    if abs(tax_liab) > 0.005:
        tax_memo = f"{ym_prefix} Uruguay Payroll - Payroll Tax Liability" if payrun_type == "regular" else f"{je_memo_main} - Payroll Tax Liability"
        block1_lines.append({
            "Date": je_date.strftime("%Y-%m-%d"), "Journal Entry Memo": je_memo_main,
            "Account": LIAB_TAX_ACCT, "Debit": "", "Credit": tax_liab,
            "Line Memo": tax_memo, "Subsidiary": SUBSIDIARY, "Department": "",
        })

    # Block 2 CR: Accrued PTO + Accrued Bonus PR Tax Liability
    block2_dr = sum(float(l["Debit"]) for l in block2_lines if l["Debit"] != "")
    if block2_dr != 0:
        # Aguinaldo CR = sum of AguinaldoBonus lines DR
        # Aguinaldo PR Tax CR = sum of AguinaldoBonusTax lines DR
        aguinaldo_dr = sum(float(l["Debit"]) for l in block2_lines if l["Debit"] != "" and "PR Tax" not in l["Line Memo"])
        aguinaldo_tax_dr = sum(float(l["Debit"]) for l in block2_lines if l["Debit"] != "" and "PR Tax" in l["Line Memo"])
        if aguinaldo_dr != 0:
            block2_lines.append({
                "Date": je_date.strftime("%Y-%m-%d"), "Journal Entry Memo": je_memo_accrual,
                "Account": ACCRUED_PTO_ACCT, "Debit": "", "Credit": round(aguinaldo_dr, 2),
                "Line Memo": f"{ym_prefix} Uruguay Payroll - Aguinaldo" if payrun_type == "regular"
                              else f"{je_memo_main} - Aguinaldo",
                "Subsidiary": SUBSIDIARY, "Department": "",
            })
        if aguinaldo_tax_dr != 0:
            block2_lines.append({
                "Date": je_date.strftime("%Y-%m-%d"), "Journal Entry Memo": je_memo_accrual,
                "Account": ACCRUED_BONUS_TAX_ACCT, "Debit": "", "Credit": round(aguinaldo_tax_dr, 2),
                "Line Memo": f"{ym_prefix} Uruguay Payroll - Aguinaldo PR Tax" if payrun_type == "regular"
                              else f"{je_memo_main} - Aguinaldo PR Tax",
                "Subsidiary": SUBSIDIARY, "Department": "",
            })

    summary = {
        "raw_file": os.path.basename(raw_file),
        "payrun_type": payrun_type,
        "payrun_label": payrun_label,
        "je_date": je_date,
        "je_memo_main": je_memo_main,
        "je_memo_accrual": je_memo_accrual,
        "employee_count": len(employee_rows),
        "net_pay": net_pay_total,
        "block1_dr": block1_dr,
        "block1_cr": sum(float(l["Credit"]) for l in block1_lines if l["Credit"] != ""),
        "block2_dr": sum(float(l["Debit"]) for l in block2_lines if l["Debit"] != ""),
        "block2_cr": sum(float(l["Credit"]) for l in block2_lines if l["Credit"] != ""),
        "flagged_special": flagged_special,
    }
    return block1_lines, block2_lines, summary


def aggregate_lines(lines):
    """Aggregate lines by (Account, Department, Line Memo)."""
    if not lines:
        return []
    df = pd.DataFrame(lines)
    df["Debit_n"] = pd.to_numeric(df["Debit"], errors="coerce").fillna(0.0)
    df["Credit_n"] = pd.to_numeric(df["Credit"], errors="coerce").fillna(0.0)
    agg = (df.groupby(["Date", "Journal Entry Memo", "Account", "Line Memo", "Subsidiary", "Department"],
                       dropna=False, as_index=False, sort=False)[["Debit_n", "Credit_n"]].sum())
    rows = []
    for _, r in agg.iterrows():
        rows.append({
            "Date": r["Date"], "Journal Entry Memo": r["Journal Entry Memo"],
            "Account": r["Account"],
            "Debit": round(r["Debit_n"], 2) if r["Debit_n"] else "",
            "Credit": round(r["Credit_n"], 2) if r["Credit_n"] else "",
            "Line Memo": r["Line Memo"], "Subsidiary": r["Subsidiary"],
            "Department": r["Department"],
        })
    return rows


def generate_jes(folder_path, infra_cogs_pct, output_xlsx=None):
    script_dir = os.path.dirname(os.path.abspath(__file__))
    emp_map = _load_employee_map(os.path.join(script_dir, "Uruguay Payroll Employee Mapping.csv"))
    gl_map = _load_gl_map(os.path.join(script_dir, "Uruguay Payroll GL Mapping.csv"))

    # Discover raw files
    raw_files = sorted(glob.glob(os.path.join(folder_path, "SGN_892_*.xls*")))
    raw_files = [f for f in raw_files if not os.path.basename(f).startswith("~$")]
    if not raw_files:
        raise RuntimeError(f"No SGN_892_*.xls* files found in {folder_path}")

    # Derive YYYY-MM from folder name
    folder_name = os.path.basename(folder_path.rstrip("/\\"))
    m = re.match(r"(\d{4})-(\d{2})", folder_name)
    if m:
        ym_prefix = f"{m.group(1)}-{m.group(2)}"
    else:
        ym_prefix = folder_name

    if output_xlsx is None:
        output_xlsx = os.path.join(folder_path, f"{ym_prefix} Uruguay Payroll Backup.xlsx")

    print(f"Folder: {folder_path}")
    print(f"Output: {output_xlsx}")
    print(f"Infra COGS Split: {infra_cogs_pct}%")
    print(f"Found {len(raw_files)} raw payrun file(s):")
    for f in raw_files:
        print(f"  - {os.path.basename(f)}")

    all_jes = []  # list of (je_label, lines)
    summaries = []
    raw_dfs = {}
    for raw_file in raw_files:
        print(f"\n--- Processing {os.path.basename(raw_file)} ---")
        block1, block2, summary = process_payrun(raw_file, emp_map, gl_map, infra_cogs_pct, ym_prefix)
        summaries.append(summary)
        block1_agg = aggregate_lines(block1)
        block2_agg = aggregate_lines(block2)

        b1_dr = sum(float(l["Debit"]) for l in block1_agg if l["Debit"] != "")
        b1_cr = sum(float(l["Credit"]) for l in block1_agg if l["Credit"] != "")
        b2_dr = sum(float(l["Debit"]) for l in block2_agg if l["Debit"] != "")
        b2_cr = sum(float(l["Credit"]) for l in block2_agg if l["Credit"] != "")

        print(f"  Type:       {summary['payrun_type']}")
        print(f"  Date/memo:  {summary['je_date']} / {summary['je_memo_main']}")
        print(f"  Employees:  {summary['employee_count']}")
        print(f"  Block 1:    DR {b1_dr:>14,.2f} | CR {b1_cr:>14,.2f}  {'BALANCED' if abs(b1_dr-b1_cr)<0.01 else 'IMBALANCED'}")
        print(f"  Block 2:    DR {b2_dr:>14,.2f} | CR {b2_cr:>14,.2f}  {'BALANCED' if abs(b2_dr-b2_cr)<0.01 else 'IMBALANCED'}")
        if summary['flagged_special']:
            print(f"  [FLAG] Special bonus (code 50/MIP):")
            for eid, name, amt, dept in summary['flagged_special']:
                print(f"    {eid} {name} {amt:,.2f} -> {dept}")

        if abs(b1_dr - b1_cr) > 0.01:
            raise RuntimeError(f"Block 1 imbalanced for {os.path.basename(raw_file)}: DR={b1_dr:.2f}, CR={b1_cr:.2f}")
        if abs(b2_dr - b2_cr) > 0.01:
            raise RuntimeError(f"Block 2 imbalanced for {os.path.basename(raw_file)}: DR={b2_dr:.2f}, CR={b2_cr:.2f}")

        if block1_agg:
            all_jes.append((summary['je_memo_main'], block1_agg))
        if block2_agg:
            all_jes.append((summary['je_memo_accrual'], block2_agg))

        # Read the raw for output workbook
        try:
            raw_dfs[os.path.basename(raw_file)] = pd.read_excel(raw_file, sheet_name="SGN", header=None)
        except Exception:
            pass

    # Write output workbook: raw tabs + JE tabs
    wb = Workbook()
    # Remove default sheet
    default = wb.active
    wb.remove(default)

    # Add raw tabs
    for fname, df in raw_dfs.items():
        # Sanitize sheet name (max 31 chars, no special chars)
        sheet_name = re.sub(r"[\\/:*?\"<>|]", "_", os.path.splitext(fname)[0])[:31]
        ws = wb.create_sheet(f"raw_{sheet_name}"[:31])
        for row in dataframe_to_rows(df, index=False, header=False):
            ws.append(list(row))

    # Add JE tabs
    for i, (label, lines) in enumerate(all_jes, start=1):
        sheet_name = re.sub(r"[\\/:*?\"<>|]", "_", label)[:30]
        sheet_name = f"JE{i:02d}_{sheet_name}"[:31]
        ws = wb.create_sheet(sheet_name)
        cols = ["Date", "Journal Entry Memo", "Account", "Debit", "Credit",
                "Line Memo", "Subsidiary", "Department"]
        ws.append(cols)
        for row in lines:
            ws.append([row[c] for c in cols])

    wb.save(output_xlsx)

    # Stack all JEs into one CSV; tag each row with a per-JE External ID so
    # NetSuite Import Assistant bundles rows from the same JE together.
    combined_rows = []
    for je_label, lines in all_jes:
        suffix = je_label.replace(ym_prefix, "", 1).strip()
        suffix = re.sub(r"^Uruguay\s*", "", suffix, flags=re.IGNORECASE)
        eid = make_external_id(ym_prefix, "UY", suffix)
        for line in lines:
            combined_rows.append({**line, "_external_id": eid})

    csv_path = Path(output_xlsx).with_name(f"{ym_prefix} Uruguay Payroll JE Import.csv")
    write_je_csv(
        rows=combined_rows,
        output_path=csv_path,
        currency="UYU",
        external_id_fn=lambda row: row["_external_id"],
    )

    print(f"\n=== Wrote {output_xlsx} with {len(raw_dfs)} raw tabs + {len(all_jes)} JE tabs ===")
    print(f"Wrote {csv_path}")
    print(f"Total JEs to post: {len(all_jes)}")
    grand_dr = sum(float(l["Debit"]) for _, lines in all_jes for l in lines if l["Debit"] != "")
    grand_cr = sum(float(l["Credit"]) for _, lines in all_jes for l in lines if l["Credit"] != "")
    print(f"Grand DR: {grand_dr:,.2f} UYU | Grand CR: {grand_cr:,.2f} UYU")
    return all_jes, summaries


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Process Uruguay payroll into NetSuite JEs.")
    parser.add_argument("folder", help="Path to YYYY-MM payroll folder containing SGN_892_* files")
    parser.add_argument("--infra-cogs-pct", type=float, default=30.0,
                        help="Infrastructure dept COGS split %% (default 30; use 20 for Jan-Mar 2026 regression)")
    parser.add_argument("--output", help="Output xlsx path (default: <folder>/<YYYY-MM> Uruguay Payroll Backup.xlsx)")
    args = parser.parse_args()
    generate_jes(args.folder, args.infra_cogs_pct, args.output)
