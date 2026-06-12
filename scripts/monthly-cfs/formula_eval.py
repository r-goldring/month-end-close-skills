"""Tiny Excel formula evaluator for the CFS workbooks.

Supports exactly the formula subset observed in the Monthly CFS files:
numbers, A1 refs (with optional $), cross-sheet refs ('Sheet Name'!B3 or
Sheet!B3), SUM(arg, range, ...), + - * / and unary minus, parentheses.

Evaluates an openpyxl workbook loaded with data_only=False (formula strings).
"""

import re
from openpyxl.utils import column_index_from_string, get_column_letter


TOKEN = re.compile(r"""
    (?P<sheet>'[^']+'!|[A-Za-z0-9_. ]+?!)?
    (?P<ref>\$?[A-Z]{1,3}\$?\d+(?::\$?[A-Z]{1,3}\$?\d+)?)
  | (?P<num>\d+\.?\d*(?:[eE][+-]?\d+)?)
  | (?P<func>SUM\()
  | (?P<op>[-+*/(),])
""", re.VERBOSE)


class Evaluator:
    def __init__(self, wb):
        self.wb = wb
        self.memo = {}
        self.overrides = {}   # (sheet, "B12") -> value

    def set_override(self, sheet, addr, value):
        self.overrides[(sheet, addr.replace("$", ""))] = value
        self.memo.clear()

    def cell(self, sheet, addr):
        addr = addr.replace("$", "")
        key = (sheet, addr)
        if key in self.overrides:
            return self.overrides[key]
        if key in self.memo:
            return self.memo[key]
        self.memo[key] = 0.0  # cycle guard
        ws = self.wb[sheet]
        v = ws[addr].value
        if isinstance(v, str) and v.startswith("="):
            out = self.formula(sheet, v)
        elif isinstance(v, (int, float)):
            out = float(v)
        else:
            out = 0.0
        self.memo[key] = out
        return out

    def expand_range(self, sheet, ref):
        if ":" not in ref:
            return [self.cell(sheet, ref)]
        a, b = ref.replace("$", "").split(":")
        m1 = re.match(r"([A-Z]+)(\d+)", a)
        m2 = re.match(r"([A-Z]+)(\d+)", b)
        c1, r1 = column_index_from_string(m1.group(1)), int(m1.group(2))
        c2, r2 = column_index_from_string(m2.group(1)), int(m2.group(2))
        out = []
        for rr in range(min(r1, r2), max(r1, r2) + 1):
            for cc in range(min(c1, c2), max(c1, c2) + 1):
                out.append(self.cell(sheet, f"{get_column_letter(cc)}{rr}"))
        return out

    def formula(self, home_sheet, text):
        toks = []
        i = 0
        body = text[1:] if text.startswith("=") else text
        while i < len(body):
            if body[i] == " ":
                i += 1
                continue
            m = TOKEN.match(body, i)
            if not m:
                raise ValueError(f"cannot tokenize {body!r} at {i}")
            i = m.end()
            if m.group("func"):
                toks.append(("func", "SUM"))
                toks.append(("op", "("))
            elif m.group("ref"):
                sheet = m.group("sheet")
                if sheet:
                    sheet = sheet[:-1].strip("'")
                toks.append(("ref", (sheet or home_sheet, m.group("ref"))))
            elif m.group("num"):
                toks.append(("num", float(m.group("num"))))
            else:
                toks.append(("op", m.group("op")))
        # re-entrant: evaluating a ref may recurse into another formula
        saved = (getattr(self, "_toks", None), getattr(self, "_pos", 0))
        self._toks = toks
        self._pos = 0
        try:
            return self._expr()
        finally:
            self._toks, self._pos = saved

    # recursive descent: expr := term (('+'|'-') term)*
    def _peek(self):
        return self._toks[self._pos] if self._pos < len(self._toks) else (None, None)

    def _next(self):
        t = self._peek()
        self._pos += 1
        return t

    def _expr(self):
        v = self._term()
        while self._peek() == ("op", "+") or self._peek() == ("op", "-"):
            op = self._next()[1]
            rhs = self._term()
            v = v + rhs if op == "+" else v - rhs
        return v

    def _term(self):
        v = self._factor()
        while self._peek() == ("op", "*") or self._peek() == ("op", "/"):
            op = self._next()[1]
            rhs = self._factor()
            v = v * rhs if op == "*" else (v / rhs if rhs else 0.0)
        return v

    def _factor(self):
        kind, val = self._peek()
        if (kind, val) == ("op", "-"):
            self._next()
            return -self._factor()
        if (kind, val) == ("op", "+"):
            self._next()
            return self._factor()
        if (kind, val) == ("op", "("):
            self._next()
            v = self._expr()
            assert self._next() == ("op", ")"), "expected )"
            return v
        if kind == "func":
            self._next()           # SUM
            self._next()           # (
            total = 0.0
            while True:
                k2, v2 = self._peek()
                if (k2, v2) == ("op", ")"):
                    self._next()
                    break
                if (k2, v2) == ("op", ","):
                    self._next()
                    continue
                if k2 == "ref":
                    self._next()
                    total += sum(self.expand_range(v2[0], v2[1]))
                else:
                    total += self._expr()
            return total
        if kind == "ref":
            self._next()
            vals = self.expand_range(val[0], val[1])
            return vals[0] if len(vals) == 1 else sum(vals)
        if kind == "num":
            self._next()
            return val
        raise ValueError(f"unexpected token {kind} {val}")
