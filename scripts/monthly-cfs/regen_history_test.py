"""Regression test: rebuild IS/BS tab content for historical months from the
pulled NetSuite data and diff against the actual workbook tabs.

Usage: python regen_history_test.py [months...] [entities...]
Defaults: 2026-02 2026-03 2026-04 x US BV CAD PL UK UY
"""

import sys
import openpyxl

import cfs_config as cfg
from tab_builders import (EntityData, build_is_rows, build_bs_rows,
                          parse_bs_template, find_tab)

TOL = 0.011

# Rows where the NetSuite UI export applies display-layer adjustments that do
# not exist in the GL (verified during skill development; see SKILL.md):
#   - 121900/211900: open intercompany AR/AP gross-up + dynamic revaluation
#   - 111100/241xxx (US): undeposited-funds netting into deferred revenue
#   - Unrealized Gain/Loss + Net Income + Retained Earnings + 300xxx equity:
#     dynamic revaluation display and late-posted equity restatements
KNOWN_VARIANCE_PREFIXES = (
    "121900 -", "211900 -", "111100 -", "241100 -", "241150 -", "241160 -",
    "Unrealized Gain/Loss", "Net Income", "Retained Earnings", "300",
)
# AR/AP rows can differ by a cent or two (per-line vs per-transaction rounding)
ARAP_PREFIXES = ("121100 -", "121101 -", "211100 -")


def classify(label):
    s = str(label)
    if any(s.startswith(p) for p in KNOWN_VARIANCE_PREFIXES):
        return "known-variance"
    if any(s.startswith(p) for p in ARAP_PREFIXES):
        return "arap-rounding"
    return "value"


def norm_formula(f):
    return str(f).replace(" ", "") if f is not None else None


def cell_repr(v):
    if isinstance(v, float):
        return round(v, 2)
    return v


def compare(entity, ym, tab_kind, actual_ws, rows, report):
    n_diffs = {"structure": 0, "value": 0, "known-variance": 0, "arap-rounding": 0}
    max_r = max(actual_ws.max_row, len(rows))
    for i in range(1, max_r + 1):
        a_lbl = actual_ws.cell(row=i, column=1).value
        a_b = actual_ws.cell(row=i, column=2).value
        a_c = actual_ws.cell(row=i, column=3).value
        g = rows[i - 1] if i <= len(rows) else None
        g_lbl = g.label if g else None
        gb = gc = None
        if g is not None:
            if getattr(g, "formulas", None):
                gb, gc = g.formulas
            elif g.values is not None:
                gb, gc = g.values
        a_lbl_s = "" if a_lbl is None else str(a_lbl)
        g_lbl_s = "" if g_lbl is None else str(g_lbl)
        if a_lbl_s != g_lbl_s:
            report.append(f"  [{entity} {ym} {tab_kind}] row {i} LABEL: actual={a_lbl_s!r} built={g_lbl_s!r}")
            n_diffs["structure"] += 1
            continue
        for col, av, gv in (("B", a_b, gb), ("C", a_c, gc)):
            a_is_f = isinstance(av, str) and av.startswith("=")
            g_is_f = isinstance(gv, str) and gv.startswith("=")
            if a_is_f or g_is_f:
                if norm_formula(av) != norm_formula(gv):
                    report.append(f"  [{entity} {ym} {tab_kind}] {col}{i} FORMULA: actual={av!r} built={gv!r}")
                    n_diffs["structure"] += 1
            else:
                an = float(av) if isinstance(av, (int, float)) else None
                gn = float(gv) if isinstance(gv, (int, float)) else None
                if an is None and gn is None:
                    continue
                if (an is None) != (gn is None) or abs((an or 0) - (gn or 0)) > TOL:
                    cls = classify(a_lbl_s)
                    if cls == "arap-rounding" and abs((an or 0) - (gn or 0)) > 0.06:
                        cls = "value"
                    n_diffs[cls] += 1
                    tag = "" if cls == "value" else f" [{cls}]"
                    report.append(f"  [{entity} {ym} {tab_kind}] {col}{i} ({a_lbl_s[:40]}): actual={cell_repr(av)} built={cell_repr(gv)}{tag}")
    return n_diffs


def run(months, entities):
    total = 0
    for ym in months:
        for ek in entities:
            path = cfg.find_workbook(ek, ym)
            if not path:
                print(f"[{ek} {ym}] workbook not found, skipped")
                continue
            wb = openpyxl.load_workbook(path, data_only=False)
            e = cfg.ENTITIES[ek]
            data = EntityData(ek)
            prior = cfg.prev_ym(ym)
            report = []

            is_ws = find_tab(wb, "is")
            bs_ws = find_tab(wb, "bs")
            is_rows = build_is_rows(data, prior, ym)
            d1 = compare(ek, ym, "IS", is_ws, is_rows, report)

            flags = []
            tpl = parse_bs_template(bs_ws)
            bs_rows = build_bs_rows(data, tpl, prior, ym, flags)
            d2 = compare(ek, ym, "BS", bs_ws, bs_rows, report)

            agg = {k: d1[k] + d2[k] for k in d1}
            hard = agg["structure"] + agg["value"]
            status = "OK  " if hard == 0 else "DIFF"
            print(f"[{status}] {ek} {ym}: structure={agg['structure']} value={agg['value']} "
                  f"known-variance={agg['known-variance']} arap={agg['arap-rounding']}"
                  + (f", flags={flags}" if flags else ""))
            for line in report[:40]:
                print(line)
            if len(report) > 40:
                print(f"  ... +{len(report) - 40} more")
            total += hard
    print(f"\nTOTAL HARD DIFFS (structure + unexplained values): {total}")
    return total


if __name__ == "__main__":
    args = sys.argv[1:]
    months = [a for a in args if a[:2] == "20"] or ["2026-02", "2026-03", "2026-04"]
    entities = [a for a in args if a[:2] != "20"] or list(cfg.ENTITIES)
    sys.exit(0 if run(months, entities) == 0 else 1)
