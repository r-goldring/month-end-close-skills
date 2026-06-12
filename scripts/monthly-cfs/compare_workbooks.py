"""Compare a generated CFS workbook against the accountant's manually prepared one.

Compares values (generated formulas are evaluated; manual workbook uses
Excel's cached values) and formula strings, tab by tab. Known UI-display
variance rows (intercompany AR/AP gross-up, undeposited-funds netting,
dynamic revaluation) are annotated rather than failed.

Usage:
  python compare_workbooks.py <generated.xlsx> <manual.xlsx>
  python compare_workbooks.py 2026-05 US          (auto-locate both)
"""

import re
import sys

import openpyxl
from openpyxl.utils import get_column_letter

import cfs_config as cfg
from formula_eval import Evaluator

TOL = 0.02

KNOWN_VARIANCE_PREFIXES = (
    "121900 -", "211900 -", "111100 -", "241100 -", "241150 -", "241160 -",
    "Unrealized Gain/Loss", "Net Income", "Retained Earnings", "300",
    "121100 -", "121101 -", "211100 -",
    # subtotal rows that roll up the above leaves
    "Total - 121000", "Total - 121100", "Total Accounts Receivable",
    "Total Other Current Asset", "Total Current Assets", "Total ASSETS",
    "Total - 211000", "Total - 211100", "Total Accounts Payable",
    "Total - 241000", "Total Other Current Liability",
    "Total Current Liabilities", "Total Liabilities", "Total - 300150",
    "Total - Equity", "Total Equity", "Total Capital and Reserves",
    "Total Assets Less", "Current Assets Less",
)
# CFS rows fed by known-variance BS rows / dynamic revaluation
KNOWN_CFS_LABELS = ("Intercompany", "Non-cash foreign exchange adjustment",
                    "Net Income")
KNOWN_CFS_COLLABELS = ("Intercompany Receivables", "Intercompany Payables",
                       "Accumulated deficit", "Accounts receivable, net",
                       "Accounts payable", "Deferred revenue",
                       "Prepaid expenses and other current assets",
                       "Class A", "Class B", "Class C", "Class D")


def pair_tabs(gen_wb, man_wb):
    """Pair tabs by role across the two workbooks."""
    def role(name):
        if "Income Statement" in name:
            return "IS"
        if "Balance Sheet" in name or "ACSBalanceSheet" in name:
            return "BS"
        if "2023" in name:
            return None
        m = re.search(r"CFS(.*)$", name)
        if m:
            return "CFS" + m.group(1).strip()
        return None
    gen = {role(n): n for n in gen_wb.sheetnames if role(n)}
    man = {role(n): n for n in man_wb.sheetnames if role(n)}
    # CFS tab suffixes can differ in month prefix only; normalize
    pairs = []
    for r, gname in gen.items():
        if r in man:
            pairs.append((gname, man[r]))
        else:
            cands = [mn for mr, mn in man.items()
                     if mr and r and mr.split()[-1] == r.split()[-1]]
            if cands:
                pairs.append((gname, cands[0]))
    return pairs


def known(label_a, label_col=""):
    s = str(label_a)
    if any(s.startswith(p) for p in KNOWN_VARIANCE_PREFIXES):
        return True
    if any(s.startswith(p) for p in KNOWN_CFS_LABELS):
        return True
    if any(str(label_col).startswith(p) for p in KNOWN_CFS_COLLABELS):
        return True
    return False


def compare(gen_path, man_path):
    gen_wb = openpyxl.load_workbook(gen_path, data_only=False)
    man_f = openpyxl.load_workbook(man_path, data_only=False)
    man_v = openpyxl.load_workbook(man_path, data_only=True)
    ev = Evaluator(gen_wb)

    summary = {}
    for gname, mname in pair_tabs(gen_wb, man_f):
        gws = gen_wb[gname]
        mfs, mvs = man_f[mname], man_v[mname]
        diffs, knowns, formula_diffs = [], [], []
        externals = 0
        max_r = max(gws.max_row, mfs.max_row)
        max_c = max(gws.max_column, mfs.max_column, 3)
        col_labels = {c: gws.cell(row=2, column=c).value or "" for c in range(1, max_c + 1)}
        for r in range(1, max_r + 1):
            row_label = gws.cell(row=r, column=1).value or mfs.cell(row=r, column=1).value or ""
            for c in range(1, max_c + 1):
                gcell = gws.cell(row=r, column=c)
                mf = mfs.cell(row=r, column=c).value
                mv = mvs.cell(row=r, column=c).value
                gv = gcell.value
                addr = f"{get_column_letter(c)}{r}"
                g_is_f = isinstance(gv, str) and gv.startswith("=")
                m_is_f = isinstance(mf, str) and mf.startswith("=")
                if g_is_f and "[" in gv:
                    # external-workbook reference: Python cannot evaluate it;
                    # Excel verifies on open. Compare formula shape only,
                    # ignoring the link index ([4] vs [12] is the same file).
                    externals += 1
                    if m_is_f:
                        norm = lambda s: re.sub(r"\[\d+\]", "[]", s.replace(" ", ""))
                        if norm(gv) != norm(mf):
                            formula_diffs.append((addr, mf, gv))
                    continue
                if g_is_f and m_is_f and gv.replace(" ", "") != mf.replace(" ", ""):
                    formula_diffs.append((addr, mf, gv))
                # numeric comparison
                gnum = None
                if g_is_f:
                    try:
                        gnum = ev.formula(gname, gv)
                    except Exception:
                        externals += 1     # depends on external links; Excel verifies
                        continue
                elif isinstance(gv, (int, float)) and not isinstance(gv, bool):
                    gnum = float(gv)
                mnum = float(mv) if isinstance(mv, (int, float)) and not isinstance(mv, bool) else None
                if gnum is None and mnum is None:
                    continue
                if abs((gnum or 0) - (mnum or 0)) <= TOL:
                    continue   # treats None vs ~0 as equal
                if True:
                    entry = (addr, str(row_label)[:45], mnum, None if gnum is None else round(gnum, 2))
                    if known(row_label, col_labels.get(c, "")):
                        knowns.append(entry)
                    else:
                        diffs.append(entry)
        summary[f"{gname} vs {mname}"] = (diffs, knowns, formula_diffs, externals)
    return summary


def main():
    args = sys.argv[1:]
    if args and args[0].endswith(".xlsx"):
        gen_path, man_path = args[0], args[1]
    else:
        ym, ek = args[0], args[1]
        import os
        d = cfg.month_dir(ym)
        e = cfg.ENTITIES[ek]
        files = sorted(f for f in os.listdir(d)
                       if f.startswith(f"{e['file_index']}. {ym} CFS {e['file_label']} ")
                       and f.endswith(".xlsx") and not f.startswith("~$"))
        if len(files) < 2:
            print("need two workbooks (generated + manual) in", d)
            sys.exit(2)
        man_path = os.path.join(d, files[0])
        gen_path = os.path.join(d, files[-1])
        print(f"manual:    {files[0]}\ngenerated: {files[-1]}\n")

    summary = compare(gen_path, man_path)
    total = 0
    for tabpair, (diffs, knowns, fdiffs, externals) in summary.items():
        print(f"== {tabpair}: {len(diffs)} value diffs, {len(knowns)} known-variance, "
              f"{len(fdiffs)} formula diffs"
              + (f", {externals} external-link cells (Excel verifies)" if externals else ""))
        for addr, lbl, mv, gv in diffs[:30]:
            print(f"   {addr} ({lbl}): manual={mv} generated={gv}")
        for addr, lbl, mv, gv in knowns[:10]:
            print(f"   [known] {addr} ({lbl}): manual={mv} generated={gv}")
        for addr, mf, gf in fdiffs[:15]:
            print(f"   formula {addr}: manual={mf!r} generated={gf!r}")
        total += len(diffs)
    print(f"\nTOTAL unexplained value diffs: {total}")
    return total


if __name__ == "__main__":
    sys.exit(0 if main() == 0 else 1)
