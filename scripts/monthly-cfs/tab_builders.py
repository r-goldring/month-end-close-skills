"""Build NetSuite-export-identical Income Statement and Balance Sheet tabs.

IS tabs are regenerated from scratch each month (NetSuite only includes
accounts with activity in the displayed months, so the row set shifts).
BS tabs are regenerated from the prior month's tab as a row template
(NetSuite's BS row set is stable; new accounts are inserted + flagged).

All amounts in the _data JSON are raw GL (debit - credit). Display signs
are applied per layout section (see cfs_config).
"""

import json
import os
import re

import cfs_config as cfg

ACCT_ROW_RE = re.compile(r"^(\d{3,6}) - ")


# ---------------------------------------------------------------------------
# Data loading
# ---------------------------------------------------------------------------

def load_coa():
    with open(os.path.join(cfg.DATA_DIR, "coa.json"), encoding="utf-8") as f:
        rows = json.load(f)
    by_id = {int(r["id"]): r for r in rows}
    return by_id


MONTH_KEYS = {"2025-12": "m2512", "2026-01": "m2601", "2026-02": "m2602",
              "2026-03": "m2603", "2026-04": "m2604", "2026-05": "m2605"}


class EntityData:
    def __init__(self, entity_key):
        self.entity_key = entity_key
        self.cfg = cfg.ENTITIES[entity_key]
        self.coa = load_coa()
        with open(os.path.join(cfg.DATA_DIR, f"{entity_key}_balances.json"),
                  encoding="utf-8") as f:
            rows = json.load(f)
        self.accounts = {}          # key -> record
        for r in rows:
            aid = int(r["id"])
            coa = self.coa.get(aid, {})
            num = r.get("acctnumber")
            full = coa.get("fullname") or ""
            leaf = full.split(" : ")[-1] if full else ""
            key = num if num else "__" + leaf
            self.accounts[key] = {
                "id": aid, "number": num, "leaf_name": leaf,
                "accttype": r["accttype"],
                "parent_num": coa.get("parent_num"),
                "bal": {ym: float(r[mk] or 0) for ym, mk in MONTH_KEYS.items()},
            }

    def bal(self, key, ym):
        a = self.accounts.get(key)
        return a["bal"].get(ym, 0.0) if a else 0.0

    def activity(self, key, ym):
        return self.bal(key, ym) - self.bal(key, cfg.prev_ym(ym))

    def label(self, key):
        a = self.accounts[key]
        if a["number"]:
            return f"{a['number']} - {a['leaf_name']}"
        return a["leaf_name"]

    def parent_label(self, parent_num):
        """Label for a parent account that may have no posting history here."""
        for r in self.coa.values():
            if r.get("acctnumber") == parent_num:
                leaf = (r.get("fullname") or "").split(" : ")[-1]
                return f"{parent_num} - {leaf}"
        return parent_num

    def is_sum(self, ym):
        return sum(a["bal"].get(ym, 0.0) for a in self.accounts.values()
                   if a["accttype"] in cfg.IS_ACCTTYPES)

    def retained_earnings(self, ym):
        """Display value of the computed Retained Earnings row."""
        base = f"{int(ym[:4]) - 1}-12"
        re_acct = sum(a["bal"].get(ym, 0.0) for a in self.accounts.values()
                      if a["accttype"] == "Equity" and a["leaf_name"] == "Retained Earnings")
        return -(self.is_sum(base) + re_acct)

    def net_income(self, ym):
        base = f"{int(ym[:4]) - 1}-12"
        return -(self.is_sum(ym) - self.is_sum(base))


def find_tab(wb, kind, entity_cfg=None):
    """Fuzzy tab resolution -- workbook tab names drift month to month."""
    names = wb.sheetnames
    if kind == "is":
        cands = [n for n in names if "Income Statement" in n]
    elif kind == "bs":
        cands = [n for n in names if "Balance Sheet" in n or "ACSBalanceSheet" in n]
    elif kind == "cfs_local":
        cur = entity_cfg["currency"]
        if cur == "USD":
            cands = [n for n in names if "CFS" in n and "2023" not in n]
        else:
            cands = [n for n in names if f"CFS {cur}" in n and "2023" not in n]
    elif kind == "cfs_usd":
        cands = [n for n in names if "CFS USD" in n and "2023" not in n]
        if not cands:  # US single-tab workbook
            cands = [n for n in names if "CFS" in n and "2023" not in n]
    else:
        cands = []
    if not cands:
        raise KeyError(f"no {kind} tab among {names}")
    return wb[cands[0]]


# ---------------------------------------------------------------------------
# Row model
# ---------------------------------------------------------------------------

class Row:
    def __init__(self, label, kind, indent=0, bold=False, values=None,
                 formula_slots=None, acct_key=None):
        self.label = label
        self.kind = kind          # title|header|account|parent|total|computed|blank
        self.indent = indent
        self.bold = bold
        self.values = values      # (prior, current) numbers, or None
        self.formula_slots = formula_slots  # callable(row_index_map) -> (fB, fC)
        self.acct_key = acct_key
        self.row_idx = None       # assigned at emit time


def _ranges(rows_idx):
    """Compress sorted row indices into Excel range pieces: [44], [102..105]."""
    pieces = []
    for i in sorted(rows_idx):
        if pieces and i == pieces[-1][1] + 1:
            pieces[-1][1] = i
        else:
            pieces.append([i, i])
    return pieces


def sum_formula(col, rows_idx):
    parts = []
    for a, b in _ranges(rows_idx):
        parts.append(f"{col}{a}" if a == b else f"{col}{a}:{col}{b}")
    return f"=SUM({','.join(parts)})"


# ---------------------------------------------------------------------------
# Hierarchy assembly (shared by IS sections and BS subsections)
# ---------------------------------------------------------------------------

_OBS_PARENTS = None

def _observed_parents():
    global _OBS_PARENTS
    if _OBS_PARENTS is None:
        p = os.path.join(cfg.DATA_DIR, "observed_parents.json")
        with open(p, encoding="utf-8") as f:
            _OBS_PARENTS = json.load(f)
    return _OBS_PARENTS


def build_account_tree(data, keys):
    """Group account keys by parent chains. Returns ordered node list.

    Node: {"kind": "leaf", "key": k} or
          {"kind": "parent", "num": n, "label": str, "own": key|None,
           "children": [nodes]}
    Ordering: ascending account number; unnumbered last (alphabetical);
    a parent sorts by its own number among siblings.
    """
    keys = list(keys)
    present = set(keys)
    nodes_by_parent = {}

    def sort_key(num_or_name, is_num):
        return (0, num_or_name) if is_num else (1, num_or_name)

    # Determine each key's chain of ancestors (parent numbers).
    info = {}
    for k in keys:
        a = data.accounts[k]
        info[k] = a

    # Build parent groups bottom-up. We only nest one or two levels deep in
    # practice (e.g. 660000 > 661000 > leaves); handle arbitrary depth.
    def ancestors(k):
        chain = []
        p = info[k]["parent_num"]
        seen = set()
        while p and p not in seen:
            seen.add(p)
            chain.append(p)
            row = next((r for r in data.coa.values() if r.get("acctnumber") == p), None)
            p = row.get("parent_num") if row else None
        return chain  # immediate parent first

    roots = {}

    def get_parent_node(chain):
        """chain: list of parent numbers from outermost to innermost."""
        level = roots
        node = None
        for num in chain:
            container = level.setdefault(num, {
                "kind": "parent", "num": num,
                "label": data.parent_label(num),
                "own": None, "children_map": {}, "leaves": []})
            node = container
            level = container["children_map"]
        return node

    # Whether an account with children renders as parent-header + own-posting
    # row is a property of the Acme report definition, not derivable from the
    # COA (e.g. 681000 renders as a parent even with no child displayed, but
    # 651100 renders as a plain leaf despite child 651101 having history).
    # Use the set observed in historical exports, plus any number whose child
    # is actually displayed this month.
    coa_parents = set(_observed_parents()) | {
        info[k]["parent_num"] for k in keys if info[k]["parent_num"]}

    plain = []
    for k in keys:
        chain = list(reversed(ancestors(k)))  # outermost first
        num = info[k]["number"]
        is_parent_of_present = bool(num) and num in coa_parents
        if is_parent_of_present:
            node = get_parent_node(chain + [num])
            node["own"] = k
        elif chain:
            node = get_parent_node(chain)
            node["leaves"].append(k)
        else:
            plain.append(k)

    def finalize(container_map, leaves=None):
        out = []
        items = []
        for num, node in container_map.items():
            items.append((sort_key(num, True), ("parent", node)))
        for k in (leaves or []):
            n = info[k]["number"]
            items.append((sort_key(n if n else info[k]["leaf_name"], bool(n)),
                          ("leaf", k)))
        items.sort(key=lambda t: t[0])
        for _, (kind, obj) in items:
            if kind == "leaf":
                out.append({"kind": "leaf", "key": obj})
            else:
                children = finalize(obj["children_map"], obj["leaves"])
                out.append({"kind": "parent", "num": obj["num"],
                            "label": obj["label"], "own": obj["own"],
                            "children": children})
        return out

    top = finalize(roots, plain)
    return top


def emit_tree(rows, data, nodes, ym_pair, sign, indent, currency_total_rows):
    """Emit account/parent/total rows; returns list of member row objects
    (the rows whose values roll into the enclosing total)."""
    members = []
    for node in nodes:
        if node["kind"] == "leaf":
            k = node["key"]
            vals = tuple(sign * data.activity(k, ym) if ym_pair[2] == "activity"
                         else sign * data.bal(k, ym) for ym in ym_pair[:2])
            r = Row(data.label(k), "account", indent=indent, values=vals, acct_key=k)
            rows.append(r)
            members.append(r)
        else:
            hdr = Row(node["label"], "parent", indent=indent, bold=False)
            rows.append(hdr)
            inner = []
            if node["own"]:
                k = node["own"]
                vals = tuple(sign * data.activity(k, ym) if ym_pair[2] == "activity"
                             else sign * data.bal(k, ym) for ym in ym_pair[:2])
                r = Row(data.label(k), "account", indent=indent + 1, values=vals,
                        acct_key=k)
                rows.append(r)
                inner.append(r)
            inner += emit_tree(rows, data, node["children"], ym_pair, sign,
                               indent + 1, currency_total_rows)
            tot = Row(f"Total - {node['label']}", "total", indent=indent,
                      bold=True)
            tot.sum_members = inner
            rows.append(tot)
            members.append(tot)
    return members


# ---------------------------------------------------------------------------
# Income Statement builder
# ---------------------------------------------------------------------------

def build_is_rows(data, prior_ym, current_ym):
    """Return ordered list of Row objects for the IS tab."""
    layout = cfg.IS_LAYOUTS[data.cfg["is_layout"]]
    e = data.cfg
    ym_pair = (prior_ym, current_ym, "activity")

    def active(k):
        return (abs(data.activity(k, prior_ym)) > 0.005
                or abs(data.activity(k, current_ym)) > 0.005)

    by_type = {}
    for k, a in data.accounts.items():
        if a["accttype"] in cfg.IS_ACCTTYPES and active(k):
            by_type.setdefault(a["accttype"], []).append(k)

    rows = [
        Row(e["entity_name"], "title", bold=True),
        Row(e["entity_path"], "title"),
        Row("Acme Income Statement", "report_title", bold=True),
        Row(f"{cfg.period_name(prior_ym)}, {cfg.period_name(current_ym)}", "title"),
        Row("", "title"),
        Row("Options: Activity Only", "title"),
        Row("Financial Row", "colhead", values=(cfg.period_name(prior_ym),
                                                cfg.period_name(current_ym))),
        Row("\xa0", "colhead", values=("Amount", "Amount")),
    ]

    sec_indent = 1 if layout["preamble"] else 0
    if layout["preamble"]:
        rows.append(Row(layout["preamble"], "header", indent=0, bold=True))

    section_totals = {}

    def emit_section(label, accttypes, sign, indent):
        keys = [k for t in accttypes for k in by_type.get(t, [])]
        if not keys:
            return None
        rows.append(Row(label, "header", indent=indent, bold=True))
        tree = build_account_tree(data, keys)
        members = emit_tree(rows, data, tree, ym_pair, sign, indent + 1, rows)
        tot = Row(f"Total - {label}", "total", indent=indent, bold=True)
        tot.sum_members = members
        rows.append(tot)
        section_totals[label] = tot
        return tot

    main_totals = []
    for label, types, sign in layout["sections"]:
        main_totals.append(emit_section(label, types, sign, sec_indent))

    # Gross Profit after the second section (Income/Sales, COS/Purchases)
    gp = Row(layout["gross_profit_label"], "computed", indent=sec_indent, bold=True)
    gp.compute = ("sub", main_totals[0], main_totals[1])
    # insert right after section 2's total (which is current end of rows)
    idx_after_cos = rows.index(main_totals[1]) + 1 if main_totals[1] else len(rows)
    rows.insert(idx_after_cos, gp)

    op = Row(layout["operating_label"], "computed", indent=sec_indent, bold=True)
    op.compute = ("sub", gp, main_totals[2])
    rows.append(op)

    if layout["other_preamble"]:
        rows.append(Row(layout["other_preamble"], "header", indent=0, bold=True))
    other_totals = []
    for label, types, sign in layout["other_sections"]:
        other_totals.append(emit_section(label, types, sign, sec_indent))

    if layout["net_other_label"] is not None:
        no = Row(layout["net_other_label"], "computed", indent=sec_indent, bold=True)
        no.compute = ("sub_or0", other_totals[0], other_totals[1])
        rows.append(no)
        net = Row(layout["net_label"], "computed", indent=0, bold=True)
        net.compute = ("add", op, no)
        rows.append(net)
    else:
        net = Row(layout["net_label"], "computed", indent=0, bold=True)
        net.compute = ("add_or0", op, other_totals[0], other_totals[1])
        rows.append(net)

    assign_formulas(rows)
    return rows


def assign_formulas(rows, value_cols=("B", "C")):
    """Assign row indices and build formula strings + computed values."""
    for i, r in enumerate(rows, start=1):
        r.row_idx = i
    for r in rows:
        if r.kind == "total":
            idxs = [m.row_idx for m in r.sum_members]
            r.formulas = tuple(sum_formula(c, idxs) for c in value_cols)
            r.calc = tuple(sum((m.values[j] if m.values else m.calc[j])
                               for m in r.sum_members) for j in range(2))
        elif r.kind == "computed" and hasattr(r, "compute"):
            op = r.compute[0]
            args = r.compute[1:]

            def ref(x, col):
                return f"{col}{x.row_idx}" if x is not None else "0"

            def val(x, j):
                if x is None:
                    return 0.0
                return x.values[j] if x.values else x.calc[j]

            if op == "sub":
                r.formulas = tuple(f"={ref(args[0], c)}-{ref(args[1], c)}"
                                   for c in value_cols)
                r.calc = tuple(val(args[0], j) - val(args[1], j) for j in range(2))
            elif op == "sub_or0":
                r.formulas = tuple(f"={ref(args[0], c)}-{ref(args[1], c)}"
                                   for c in value_cols)
                r.calc = tuple(val(args[0], j) - val(args[1], j) for j in range(2))
            elif op == "add":
                r.formulas = tuple(f"={ref(args[0], c)}+{ref(args[1], c)}"
                                   for c in value_cols)
                r.calc = tuple(val(args[0], j) + val(args[1], j) for j in range(2))
            elif op == "add_or0":
                r.formulas = tuple(
                    f"={ref(args[0], c)}+{ref(args[1], c)}+{ref(args[2], c)}"
                    for c in value_cols)
                r.calc = tuple(val(args[0], j) + val(args[1], j) + val(args[2], j)
                               for j in range(2))


# ---------------------------------------------------------------------------
# Balance Sheet builder (template-driven)
# ---------------------------------------------------------------------------

BS_COMPUTED_LABELS = {
    "Retained Earnings", "Net Income", "Cumulative Translation Adjustment",
}

def parse_bs_template(ws):
    """Parse a prior-month BS tab into an ordered list of template rows."""
    tpl = []
    for r in range(1, ws.max_row + 1):
        a = ws.cell(row=r, column=1).value
        b = ws.cell(row=r, column=2).value
        c = ws.cell(row=r, column=3).value
        tpl.append({"row": r, "label": "" if a is None else str(a),
                    "b": b, "c": c,
                    "b_is_formula": isinstance(b, str) and b.startswith("="),
                    })
    return tpl


def build_bs_rows(data, template_rows, prior_ym, current_ym, flags):
    """Rebuild the BS tab from the template's row sequence with fresh values.

    New accounts (nonzero balance, not in template) are inserted in account-
    number order within the subsection holding their accttype, and flagged.
    Returns ordered Row list.
    """
    e = data.cfg
    bs_sign = {t: s for _, t, s in cfg.BS_SUBSECTIONS}
    bs_sign["DeferRevenue"] = -1
    bs_sign["CredCard"] = -1

    # --- figure out which accounts the template displays --------------------
    tpl_accounts = []
    for t in template_rows:
        m = ACCT_ROW_RE.match(t["label"])
        if m and not t["label"].startswith("Total"):
            # parent header rows have no values; account rows do (incl. 0)
            tpl_accounts.append((t, m.group(1)))

    tpl_numbers = set()
    for t in template_rows:
        m = ACCT_ROW_RE.match(t["label"])
        if m:
            tpl_numbers.add(m.group(1))

    # new accounts: BS-type, nonzero balance either month, not in template
    new_accts = []
    for k, a in data.accounts.items():
        if a["accttype"] in cfg.IS_ACCTTYPES or not a["number"]:
            continue
        if a["leaf_name"] == "Retained Earnings":
            continue  # folds into the computed Retained Earnings row
        if a["number"] in tpl_numbers:
            continue
        if abs(data.bal(k, prior_ym)) > 0.005 or abs(data.bal(k, current_ym)) > 0.005:
            new_accts.append(k)

    # --- rebuild row list ----------------------------------------------------
    rows = []
    title_replacements = {
        3: None,  # report title row stays
        4: f"End of {cfg.period_name(current_ym)}",
    }
    out_of = {"prior": prior_ym, "current": current_ym}

    # helper: display value for an account label
    def acct_values(number_or_label):
        m = ACCT_ROW_RE.match(number_or_label)
        num = m.group(1) if m else None
        key = num
        if key is None or key not in data.accounts:
            # unnumbered account rows (e.g. CTA-Elimination): match by leaf name
            for k2, a2 in data.accounts.items():
                if a2["leaf_name"] == number_or_label:
                    key = k2
                    break
        if key is None or key not in data.accounts:
            return (0.0, 0.0), None
        sign = bs_sign.get(data.accounts[key]["accttype"], 1)
        return (sign * data.bal(key, prior_ym), sign * data.bal(key, current_ym)), key

    inserted = set()
    i = 0
    n = len(template_rows)
    while i < n:
        t = template_rows[i]
        lbl = t["label"]
        rownum = t["row"]
        if rownum <= 9:
            # title block: rewrite the period labels
            if rownum == 4:
                rows.append(Row(f"End of {cfg.period_name(current_ym)}", "title"))
            elif rownum == 8:
                rows.append(Row("\xa0", "colhead",
                                values=(f"As of {cfg.period_name(prior_ym)}",
                                        f"As of {cfg.period_name(current_ym)}")))
            elif rownum == 7:
                rows.append(Row("Financial Row", "colhead", values=(None, None)))
            elif rownum == 9:
                rows.append(Row("\xa0", "colhead", values=("\xa0", "\xa0")))
            else:
                kind = "report_title" if rownum == 3 else "title"
                rows.append(Row(lbl, kind, bold=rownum in (1, 3)))
            i += 1
            continue

        m = ACCT_ROW_RE.match(lbl)
        is_total = lbl.startswith("Total")
        if m and not is_total:
            num = m.group(1)
            # insert any pending new accounts that sort before this one and
            # share its accttype neighborhood
            for k in sorted(new_accts, key=lambda x: data.accounts[x]["number"]):
                if k in inserted:
                    continue
                a = data.accounts[k]
                here_type = data.accounts.get(num, {}).get("accttype")
                if here_type == a["accttype"] and a["number"] < num:
                    sign = bs_sign.get(a["accttype"], 1)
                    r = Row(data.label(k), "account",
                            values=(sign * data.bal(k, prior_ym),
                                    sign * data.bal(k, current_ym)),
                            acct_key=k)
                    r.is_new = True
                    rows.append(r)
                    inserted.add(k)
                    flags.append({"type": "new_bs_account", "entity": data.entity_key,
                                  "account": data.label(k),
                                  "balance_current": round(data.bal(k, current_ym), 2),
                                  "inserted_before": lbl})
            if t["b"] is None and t["c"] is None and not t["b_is_formula"]:
                rows.append(Row(lbl, "parent"))
            else:
                vals, key = acct_values(lbl)
                rows.append(Row(lbl, "account", values=vals, acct_key=key))
            i += 1
            continue

        if lbl in BS_COMPUTED_LABELS and not t["b_is_formula"]:
            if lbl == "Retained Earnings":
                v = data.retained_earnings(current_ym)
                vp = data.retained_earnings(prior_ym)
                rows.append(Row(lbl, "re", values=(vp, v)))
            elif lbl == "Net Income":
                rows.append(Row(lbl, "ni", values=(data.net_income(prior_ym),
                                                   data.net_income(current_ym))))
            else:  # CTA computed row (entity statements: 0)
                rows.append(Row(lbl, "cta", values=(0.0, 0.0)))
            i += 1
            continue

        if is_total or (t["b_is_formula"]):
            r = Row(lbl, "tpl_total", bold=True)
            r.tpl_formula = t["b"]
            r.tpl_row = rownum
            rows.append(r)
            i += 1
            continue

        # unnumbered account-like rows with values (e.g. CTA-Elimination acct)
        if not m and (isinstance(t["b"], (int, float)) or isinstance(t["c"], (int, float))):
            vals, key = acct_values(lbl)
            rows.append(Row(lbl, "account", values=vals, acct_key=key))
            i += 1
            continue

        # plain section header
        rows.append(Row(lbl, "header", bold=True))
        i += 1

    # any new accounts not yet inserted (sorted after all peers): flag only
    for k in new_accts:
        if k not in inserted:
            flags.append({"type": "new_bs_account_unplaced", "entity": data.entity_key,
                          "account": data.label(k),
                          "balance_current": round(data.bal(k, current_ym), 2),
                          "note": "could not find insertion point; add manually"})

    # assign row indices and remap template formulas to new row numbers
    offset_map = {}
    for new_idx, r in enumerate(rows, start=1):
        r.row_idx = new_idx
    # map old template row -> new row: walk both lists in lockstep
    new_iter = [r for r in rows if not getattr(r, "is_new", False)]
    for t, r in zip(template_rows, new_iter):
        offset_map[t["row"]] = r.row_idx

    ref_re = re.compile(r"([B-C])(\d+)")

    def remap(formula):
        def sub(mm):
            col, old = mm.group(1), int(mm.group(2))
            return f"{col}{offset_map.get(old, old)}"
        return ref_re.sub(sub, formula)

    for r in rows:
        if r.kind == "tpl_total":
            fb = remap(r.tpl_formula)
            fc = ref_re.sub(lambda mm: f"C{mm.group(2)}", fb)
            # widen SUM ranges that should absorb adjacent inserted accounts
            r.formulas = (fb, fc)
    # expand SUM ranges to cover inserted rows that fall inside/adjacent
    if any(getattr(r, "is_new", False) for r in rows):
        _expand_ranges_for_inserts(rows)
    return rows


def _expand_ranges_for_inserts(rows):
    new_rows = [r.row_idx for r in rows if getattr(r, "is_new", False)]
    rng = re.compile(r"([BC])(\d+):([BC])(\d+)")
    for r in rows:
        if r.kind != "tpl_total":
            continue
        fixed = []
        for f in r.formulas:
            def widen(mm):
                col, a, _, b = mm.group(1), int(mm.group(2)), mm.group(3), int(mm.group(4))
                for nr in new_rows:
                    if a - 1 <= nr <= b + 1:
                        a, b = min(a, nr), max(b, nr)
                return f"{col}{a}:{col}{b}"
            fixed.append(rng.sub(widen, f))
        r.formulas = tuple(fixed)


# ---------------------------------------------------------------------------
# Write rows into an openpyxl worksheet
# ---------------------------------------------------------------------------

def write_rows(ws, rows, currency_symbol, tab_kind):
    from openpyxl.styles import Font, PatternFill, Alignment

    body = Font(name="Arial", size=8)
    bold = Font(name="Arial", size=8, bold=True)
    title = Font(name="Arial", size=12, bold=True)
    rpt = Font(name="Arial", size=14, bold=True)
    head = Font(name="Arial", size=7, bold=True)
    gray = PatternFill("solid", fgColor="FFD0D0D0")
    numfmt = (cfg.is_number_format(currency_symbol) if tab_kind == "is"
              else cfg.bs_number_format(currency_symbol))

    # clear columns A:C only -- columns D+ may hold the accountant's side-calculations
    # (e.g. the US BS fixed-asset delta column) and must survive the refresh
    for rng in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(rng))
    none_fill = PatternFill()
    for r in range(1, ws.max_row + 1):
        for c in range(1, 4):
            cell = ws.cell(row=r, column=c)
            cell.value = None
            cell.fill = none_fill
            cell.number_format = "General"

    for r in rows:
        i = r.row_idx
        ca = ws.cell(row=i, column=1, value=r.label if r.label != "" else "")
        cb = ws.cell(row=i, column=2)
        cc = ws.cell(row=i, column=3)
        if r.kind in ("title", "report_title"):
            ca.font = rpt if r.kind == "report_title" else (title if r.bold else body)
        elif r.kind == "colhead":
            for cell, v in ((ca, None), (cb, r.values[0] if r.values else None),
                            (cc, r.values[1] if r.values else None)):
                cell.font = head
                cell.fill = gray
                if v is not None:
                    cell.value = v
        else:
            f = bold if r.bold or r.kind in ("header", "parent", "total",
                                             "tpl_total", "computed",
                                             "re", "ni", "cta") else body
            ca.font = f
            if r.indent:
                ca.alignment = Alignment(indent=r.indent)
            if getattr(r, "formulas", None):
                cb.value, cc.value = r.formulas
            elif r.values is not None:
                cb.value = round(r.values[0], 2) if isinstance(r.values[0], float) else r.values[0]
                cc.value = round(r.values[1], 2) if isinstance(r.values[1], float) else r.values[1]
            for cell in (cb, cc):
                cell.font = f
                cell.number_format = numfmt
            if getattr(r, "is_new", False):
                fill = PatternFill("solid", fgColor=cfg.YELLOW_FILL)
                ca.fill = fill
                cb.fill = fill
                cc.fill = fill

    ws.column_dimensions["A"].width = (cfg.STYLE["col_a_width_is"] if tab_kind == "is"
                                       else cfg.STYLE["col_a_width_bs"])
    w = (cfg.STYLE["amount_col_width_usd"] if currency_symbol == "$"
         else cfg.STYLE["amount_col_width_foreign"])
    ws.column_dimensions["B"].width = w
    ws.column_dimensions["C"].width = w
