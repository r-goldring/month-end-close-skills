"""Build the consolidated CFS workbook for a target month.

Copies the prior month's consolidated workbook, repoints its external links
(zip-level: relationship targets + source sheet names) at the new month's
subsidiary workbooks on the Finance Team shared drive, renames tabs, rewrites
formula month references, drops in FX rates, and mirrors the manual cells
from the freshly built subsidiary workbooks.

Run AFTER build_subsidiary_cfs.py for the same month.

Usage: python build_consolidated_cfs.py 2026-05
"""

import datetime
import json
import os
import re
import shutil
import sys

import openpyxl
from openpyxl.styles import PatternFill

import cfs_config as cfg
from tab_builders import find_tab

YELLOW = PatternFill("solid", fgColor=cfg.YELLOW_FILL)
YM_RE = re.compile(r"20(2[5-9]|[3-9]\d)-\d\d")


def swap_ym(text, target_ym):
    return YM_RE.sub(target_ym, text)


def build(target_ym, prep_date=None, report=None):
    prep_date = prep_date or datetime.date.today()
    report = report if report is not None else {}
    prior_ym = cfg.prev_ym(target_ym)
    tpl_path = cfg.find_workbook("CONS", prior_ym)
    if not tpl_path:
        raise FileNotFoundError(f"no consolidated workbook for {prior_ym}")
    out_dir = cfg.month_dir(target_ym)
    out_path = os.path.join(out_dir, cfg.workbook_name("CONS", target_ym, prep_date))
    shutil.copy2(tpl_path, out_path)
    rep = report.setdefault("CONS", {"manual_cells": [], "flags": [], "checks": [],
                                     "output": out_path})
    try:
        # locate the freshly built subsidiary workbooks
        sub_paths, new_names = {}, {}
        for ek in cfg.ENTITIES:
            p = cfg.find_workbook(ek, target_ym)
            if not p:
                raise FileNotFoundError(f"build subsidiary {ek} {target_ym} first")
            sub_paths[ek] = p
            new_names[ek] = os.path.basename(p)

        wb = openpyxl.load_workbook(out_path)

        # rename tabs + rewrite every formula's month references
        renames = {}
        for name in list(wb.sheetnames):
            new = swap_ym(name, target_ym)
            if new != name:
                renames[name] = new
                wb[name].title = new
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for c in row:
                    if isinstance(c.value, str) and (c.value.startswith("=")):
                        nv = swap_ym(c.value, target_ym)
                        if nv != c.value:
                            c.value = nv
            # period dates in column A
            date_cells = [ws.cell(row=r, column=1) for r in range(2, 8)
                          if isinstance(ws.cell(row=r, column=1).value, datetime.datetime)]
            if len(date_cells) >= 2:
                date_cells[0].value = datetime.datetime.combine(
                    cfg.month_end(prior_ym), datetime.time())
                date_cells[1].value = datetime.datetime.combine(
                    cfg.month_end(target_ym), datetime.time())
            for r in range(1, ws.max_row + 1):
                a = ws.cell(row=r, column=1)
                if isinstance(a.value, str) and "Note:" in a.value:
                    a.value = re.sub(r"(January|February|March|April|May|June|July|"
                                     r"August|September|October|November|December) 20\d\d",
                                     cfg.month_end(target_ym).strftime("%B %Y"), a.value)

        # FX rates on the per-subsidiary USD tabs
        with open(os.path.join(cfg.DATA_DIR, "fx.json"), encoding="utf-8") as f:
            fx = json.load(f)
        from build_subsidiary_cfs import fx_target_cells
        for ek, (tab_pat, prior_cell, cur_cell, _off) in cfg.CONSOLIDATED["sub_tabs"].items():
            ws = wb[tab_pat.format(ym=target_ym)]
            cur = cfg.ENTITIES[ek]["currency"]
            spec = fx_target_cells(ws) or (prior_cell, cur_cell, ek == "BV")
            p_cell, c_cell, with_labels = spec
            ws[p_cell] = fx[prior_ym][cur]
            ws[c_cell] = fx[target_ym][cur]
            if with_labels:
                pe, ce = cfg.month_end(prior_ym), cfg.month_end(target_ym)
                ws[p_cell.replace("2", "1")] = f"{pe.month}/{pe.day}/{pe.year} FX Rate"
                ws[c_cell.replace("2", "1")] = f"{ce.month}/{ce.day}/{ce.year} FX Rate"
        rep["checks"].append({"check": "fx_rates",
                              "rates": {c: fx[target_ym][c] for c in cfg.FX_SUBSIDIARY}})

        # refresh / mirror the per-subsidiary tabs from the built workbooks
        from formula_eval import Evaluator
        tab_for_entity = {"US": f"{target_ym} CFS US"}
        for ek, (tab_pat, *_rest) in cfg.CONSOLIDATED["sub_tabs"].items():
            tab_for_entity[ek] = tab_pat.format(ym=target_ym)
        for ek, cons_tab in tab_for_entity.items():
            sub_wb = openpyxl.load_workbook(sub_paths[ek])
            sub_ws = find_tab(sub_wb, "cfs_usd", cfg.ENTITIES[ek])
            cons_ws = wb[cons_tab]
            has_external = any(
                isinstance(c.value, str) and c.value.startswith("=") and "[" in c.value
                for row in cons_ws.iter_rows(min_row=3, max_row=60, min_col=2, max_col=32)
                for c in row)
            if not has_external:
                # tab carries pasted values (the accountant's pattern varies by month):
                # recompute every cell from the subsidiary workbook
                ev = Evaluator(sub_wb)
                n = 0
                for row in sub_ws.iter_rows(min_row=1, max_row=60, min_col=2, max_col=32):
                    for sc in row:
                        v = sc.value
                        if isinstance(v, str) and v.startswith("="):
                            try:
                                v = round(ev.formula(sub_ws.title, v), 2)
                            except Exception:
                                continue
                        cc = cons_ws[sc.coordinate]
                        if cc.value != v and not (cc.value is None and v in (0, 0.0)):
                            cc.value = v
                            n += 1
                rep["flags"].append({
                    "type": "consolidated_tab_values_refreshed", "entity": ek,
                    "tab": cons_tab, "cells": n,
                    "note": "tab had no live external links (pasted values); "
                            "values recomputed from the subsidiary workbook"})
                continue
            # live-link tab: mirror just the manual constants
            for row in sub_ws.iter_rows(min_row=3, max_row=60, min_col=2, max_col=32):
                for sc in row:
                    if isinstance(sc.value, (int, float)) and not isinstance(sc.value, bool):
                        cc = cons_ws[sc.coordinate]
                        if isinstance(cc.value, str) and cc.value.startswith("="):
                            continue   # consolidated uses a formula here; leave it
                        if cc.value == sc.value or (cc.value is None and sc.value == 0):
                            continue
                        rep["manual_cells"].append({
                            "tab": cons_tab, "cell": sc.coordinate,
                            "old": cc.value, "new": sc.value,
                            "derivation": f"mirrored from {ek} workbook {sub_ws.title}"})
                        cc.value = sc.value
                        cc.fill = YELLOW

        wb.save(out_path)

        # repoint subsidiary-file external links + restore the pristine
        # Excel-authored external-link parts (openpyxl mangles them on save).
        # Done last, as zip surgery, so nothing re-touches them afterwards.
        from external_links import finalize
        repointed = finalize(out_path, tpl_path, prior_ym, target_ym, new_names)
        rep["checks"].append({"check": "external_links_repointed", "links": repointed})
        if len(repointed) != 6:
            rep["flags"].append({"type": "external_links", "entity": "CONS",
                                 "note": f"expected 6 repointed links, got {len(repointed)}"})
        return out_path, rep
    except Exception:
        if os.path.exists(out_path):
            os.remove(out_path)
        raise


if __name__ == "__main__":
    target = sys.argv[1]
    report = {}
    path, rep = build(target, report=report)
    print(f"[CONS] -> {path}")
    for c in rep["checks"]:
        print(f"    check: {c}")
    for m in rep["manual_cells"]:
        print(f"    manual {m['tab']}!{m['cell']}: {m['old']} -> {m['new']}")
    for f in rep["flags"]:
        print(f"    FLAG: {f}")
    rp = os.path.join(cfg.month_dir(target), f"_build_report_{target}.json")
    existing = {}
    if os.path.exists(rp):
        with open(rp, encoding="utf-8") as f:
            existing = json.load(f)
    existing.update(report)
    with open(rp, "w", encoding="utf-8") as f:
        json.dump(existing, f, indent=1)
