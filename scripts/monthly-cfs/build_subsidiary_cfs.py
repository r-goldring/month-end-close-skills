"""Build one subsidiary CFS workbook for a target month.

Copies the prior month's workbook, regenerates the IS and BS tabs from the
pulled NetSuite data, rolls the CFS tab(s) (tab names, period dates, IS
reference re-pointing, FX rates), derives manual cells by zeroing the check
formulas (the accountant's rule), highlights every derived/carried manual cell yellow,
and emits a review report.

Usage: python build_subsidiary_cfs.py 2026-05 [US BV CAD PL UK UY]
"""

import datetime
import json
import os
import re
import shutil
import sys

import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

import cfs_config as cfg
from tab_builders import (EntityData, build_is_rows, build_bs_rows,
                          parse_bs_template, write_rows, find_tab)
from formula_eval import Evaluator

YELLOW = PatternFill("solid", fgColor=cfg.YELLOW_FILL)
MONTH_TAB_RE = re.compile(r"(202[5-9]-\d\d)")


def month_name(ym):
    return cfg.month_end(ym).strftime("%B %Y")     # e.g. "May 2026"


def rewrite_formulas(wb, replacements):
    """String-replace sheet references across all formula cells."""
    if not replacements:
        return
    items = sorted(replacements.items(), key=lambda kv: -len(kv[0]))
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                if isinstance(c.value, str) and c.value.startswith("="):
                    v = c.value
                    for old, new in items:
                        v = v.replace(f"'{old}'!", f"'{new}'!")
                        v = v.replace(f"{old}!", f"'{new}'!")
                    c.value = v


def fx_target_cells(ws):
    """Detect where the FX rates live on a USD CFS tab.

    Two observed layouts: rates in B1/C1 (no labels), or date labels in B1/C1
    with rates in B2/C2. Returns (prior_cell, current_cell, with_labels).
    """
    b1, b2 = ws["B1"].value, ws["B2"].value
    if isinstance(b2, (int, float)) and isinstance(b1, str):
        return "B2", "C2", True
    if isinstance(b1, (int, float)):
        return "B1", "C1", False
    return None


def find_label_row(ws, label, col=1):
    for r in range(1, ws.max_row + 1):
        if str(ws.cell(row=r, column=col).value or "").strip() == label:
            return r
    return None


def build(entity_key, target_ym, prep_date=None, report=None):
    prep_date = prep_date or datetime.date.today()
    report = report if report is not None else {}
    prior_ym = cfg.prev_ym(target_ym)
    e = cfg.ENTITIES[entity_key]
    tpl_path = cfg.find_workbook(entity_key, prior_ym)
    if not tpl_path:
        raise FileNotFoundError(f"no {entity_key} workbook for {prior_ym}")
    out_dir = cfg.month_dir(target_ym)
    os.makedirs(out_dir, exist_ok=True)
    out_path = os.path.join(out_dir, cfg.workbook_name(entity_key, target_ym, prep_date))
    shutil.copy2(tpl_path, out_path)
    try:
        result = _build_inner(entity_key, target_ym, prior_ym, e, out_path, report)
        # restore the legacy external-link parts Excel authored (openpyxl
        # rewrites them on save and Excel then flags a repair) -- targets
        # are dead 2023 refs that never change, so restore verbatim
        from external_links import finalize
        finalize(out_path, tpl_path)
        return result
    except Exception:
        if os.path.exists(out_path):
            os.remove(out_path)     # never leave a half-built workbook behind
        raise


def _build_inner(entity_key, target_ym, prior_ym, e, out_path, report):
    wb = openpyxl.load_workbook(out_path)
    data = EntityData(entity_key)
    rep = report.setdefault(entity_key, {"manual_cells": [], "flags": [],
                                         "checks": [], "output": out_path})

    # ---- capture old IS labels before rewrite (for ref re-pointing) -------
    is_ws = find_tab(wb, "is")
    bs_ws = find_tab(wb, "bs")
    old_is_labels = {r: str(is_ws.cell(row=r, column=1).value or "")
                     for r in range(1, is_ws.max_row + 1)}

    # ---- rename CFS tabs to canonical names, rewrite references ----------
    renames = {}
    local_ws = find_tab(wb, "cfs_local", e)
    canon_local = e["cfs_tabs"][0][0].format(ym=target_ym)
    if local_ws.title != canon_local:
        renames[local_ws.title] = canon_local
    usd_ws = None
    if e["currency"] != "USD":
        usd_ws = find_tab(wb, "cfs_usd", e)
        canon_usd = e["usd_cfs_tab"].format(ym=target_ym)
        if usd_ws.title != canon_usd:
            renames[usd_ws.title] = canon_usd
    for old, new in renames.items():
        wb[old].title = new
    rewrite_formulas(wb, renames)
    local_ws = wb[canon_local]
    if usd_ws is not None:
        usd_ws = wb[e["usd_cfs_tab"].format(ym=target_ym)]
    cfs_tabs = [local_ws] + ([usd_ws] if usd_ws is not None else [])

    # ---- rewrite IS tab ----------------------------------------------------
    is_rows = build_is_rows(data, prior_ym, target_ym)
    write_rows(is_ws, is_rows, e["currency_symbol"], "is")
    new_is_row = {}
    for r in is_rows:
        if r.label and r.label not in new_is_row:
            new_is_row[r.label] = r.row_idx

    # ---- rewrite BS tab ----------------------------------------------------
    tpl = parse_bs_template(bs_ws)
    flags = []
    bs_rows = build_bs_rows(data, tpl, prior_ym, target_ym, flags)
    write_rows(bs_ws, bs_rows, e["currency_symbol"], "bs")
    rep["flags"].extend(flags)
    # row offset map for CFS formulas if BS rows were inserted
    bs_offsets = {}
    new_iter = [r for r in bs_rows if not getattr(r, "is_new", False)]
    for t, r in zip(tpl, new_iter):
        if t["row"] != r.row_idx:
            bs_offsets[t["row"]] = r.row_idx

    # ---- roll the CFS tabs --------------------------------------------------
    is_tab_names = {is_ws.title}
    bs_tab_names = {bs_ws.title}
    ref_re = re.compile(r"(?P<sheet>'[^']+'!|[A-Za-z0-9_. ]+?!)(?P<col>\$?[A-Z]{1,2})(?P<row>\$?\d+)")

    def repoint(cell):
        v = cell.value
        if not (isinstance(v, str) and v.startswith("=")):
            return
        changed = False

        def sub(m):
            nonlocal changed
            sheet = m.group("sheet")[:-1].strip("'")
            col = m.group("col")
            row = int(m.group("row").replace("$", ""))
            if sheet in is_tab_names:
                old_label = old_is_labels.get(row, "")
                new_row = new_is_row.get(old_label)
                if new_row is None:
                    changed = True
                    rep["flags"].append({
                        "type": "is_ref_lost", "entity": entity_key,
                        "cell": f"{cell.parent.title}!{cell.coordinate}",
                        "label": old_label,
                        "note": "account not on new IS (no activity); ref replaced with 0"})
                    return "0"
                if new_row != row:
                    changed = True
                return f"'{sheet}'!{col}{new_row}"
            if sheet in bs_tab_names and row in bs_offsets:
                changed = True
                return f"'{sheet}'!{col}{bs_offsets[row]}"
            return m.group(0)

        nv = ref_re.sub(sub, v)
        if changed:
            cell.value = nv

    for ws in cfs_tabs:
        for row in ws.iter_rows():
            for c in row:
                repoint(c)
        # period dates in column A (datetime cells in the top rows)
        date_cells = [ws.cell(row=r, column=1) for r in range(2, 8)
                      if isinstance(ws.cell(row=r, column=1).value, datetime.datetime)]
        if len(date_cells) >= 2:
            date_cells[0].value = datetime.datetime.combine(
                cfg.month_end(prior_ym), datetime.time())
            date_cells[1].value = datetime.datetime.combine(
                cfg.month_end(target_ym), datetime.time())
        # note rows: swap the month name in the label text
        for r in range(1, ws.max_row + 1):
            a = ws.cell(row=r, column=1)
            if isinstance(a.value, str) and "Note:" in a.value:
                a.value = re.sub(r"(January|February|March|April|May|June|July|"
                                 r"August|September|October|November|December) 20\d\d",
                                 month_name(target_ym), a.value)

    # ---- FX rates -----------------------------------------------------------
    if e["fx_cells"]:
        with open(os.path.join(cfg.DATA_DIR, "fx.json"), encoding="utf-8") as f:
            fx = json.load(f)
        cur = e["currency"]
        prior_rate = fx[prior_ym][cur]
        cur_rate = fx[target_ym][cur]
        tgt = usd_ws if usd_ws is not None else local_ws
        spec = fx_target_cells(tgt) or (e["fx_cells"]["prior"], e["fx_cells"]["current"],
                                        "prior_label" in e["fx_cells"])
        prior_cell, cur_cell, with_labels = spec
        tgt[prior_cell] = prior_rate
        tgt[cur_cell] = cur_rate
        if with_labels:
            pe, ce = cfg.month_end(prior_ym), cfg.month_end(target_ym)
            tgt[prior_cell.replace("2", "1")] = f"{pe.month}/{pe.day}/{pe.year} FX Rate"
            tgt[cur_cell.replace("2", "1")] = f"{ce.month}/{ce.day}/{ce.year} FX Rate"
        rep["checks"].append({"check": "fx_rates", "prior": prior_rate,
                              "current": cur_rate, "currency": cur})

    # ---- manual cells + plugs ----------------------------------------------
    ev = Evaluator(wb)
    rules = cfg.MANUAL_CELL_RULES.get(entity_key, {})

    def set_cell(ws, addr, value, note, old):
        ws[addr] = round(value, 2)
        ws[addr].fill = YELLOW
        ev.set_override(ws.title, addr, round(value, 2))
        rep["manual_cells"].append({
            "tab": ws.title, "cell": addr, "old": old,
            "new": round(value, 2), "derivation": note})

    for ws in cfs_tabs:
        check_row = find_label_row(ws, "Check Figure")
        fx_row = find_label_row(ws, "Effect of Exchange Rate on Cash")
        if check_row is None:
            continue
        # collect constant (manual) cells inside the allocation matrix
        first_data_row = check_row - 46 + 8   # matrix rows are 8..48 in US frame
        manual = {}
        for r in range(first_data_row, check_row):
            for cidx in range(2, 32):
                cell = ws.cell(row=r, column=cidx)
                if isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool):
                    manual.setdefault(get_column_letter(cidx), []).append(cell)

        # 1) specific rules first (cells addressed in the US frame; shift by
        #    the tab's offset relative to check row 49)
        offset = check_row - 49
        for addr, rule in rules.items():
            col = re.match(r"([A-Z]+)(\d+)", addr)
            tgt_addr = f"{col.group(1)}{int(col.group(2)) + offset}"
            cell = ws[tgt_addr]
            if not isinstance(cell.value, (int, float)):
                continue        # rule cell not a constant on this tab
            old = cell.value
            if rule["rule"] == "bs_value":
                v = rule["sign"] * data.bal(rule["acct"],
                                            target_ym if rule["which"] == "current" else prior_ym)
            elif rule["rule"] == "bs_delta":
                v = rule["sign"] * (data.bal(rule["acct"], target_ym)
                                    - data.bal(rule["acct"], prior_ym))
            elif rule["rule"] == "is_value":
                v = rule["sign"] * data.activity(rule["acct"],
                                                 target_ym if rule["which"] == "current" else prior_ym)
            else:   # carry
                rep["manual_cells"].append({"tab": ws.title, "cell": tgt_addr,
                                            "old": old, "new": old,
                                            "derivation": rule["note"] + " (carried, review)"})
                ws[tgt_addr].fill = YELLOW
                continue
            set_cell(ws, tgt_addr, v, rule["note"], old)

        # 2) per-column plugs to zero the check figures
        for cidx in range(3, 32):           # C..AE
            col = get_column_letter(cidx)
            chk = ws.cell(row=check_row, column=cidx)
            if not (isinstance(chk.value, str) and chk.value.startswith("=")):
                continue
            residual = ev.formula(ws.title, chk.value)
            if abs(residual) <= 0.01:
                continue
            rule_rows = {int(re.match(r"[A-Z]+(\d+)", a).group(1)) + offset
                         for a in rules}
            free = [c for c in manual.get(col, [])
                    if c.row not in rule_rows and c.row != check_row]
            if len(free) == 1:
                c = free[0]
                base = c.value or 0
            else:
                plug_row = (cfg.DEFAULT_PLUGS.get(col, cfg.FX_EFFECT_ROW) + offset
                            if col in cfg.DEFAULT_PLUGS else (fx_row or cfg.FX_EFFECT_ROW + offset))
                c = ws.cell(row=plug_row, column=cidx)
                if isinstance(c.value, str) and c.value.startswith("="):
                    rep["flags"].append({"type": "check_residual", "entity": entity_key,
                                         "tab": ws.title, "column": col,
                                         "residual": round(residual, 2),
                                         "note": "plug row holds a formula; review manually"})
                    continue
                base = c.value if isinstance(c.value, (int, float)) else 0
            # the "Effect of Exchange Rate on Cash" plug (row 47) is keyed to
            # whole dollars to match the accountant's style -- this leaves a sub-dollar
            # check residual exactly like his manual files; all other plugs
            # stay exact so their column check is precisely zero.
            is_fx_effect = (fx_row is not None and c.row == fx_row)
            new_val = base + residual
            if is_fx_effect:
                new_val = round(new_val)
            set_cell(ws, c.coordinate, new_val,
                     (f"column {col} FX-effect-on-cash plug, rounded to whole dollars"
                      if is_fx_effect else
                      f"column {col} plug derived to zero Check Figure (was {base})"),
                     base)

        # 3) record final check values
        ev.memo.clear()
        b_chk = ws.cell(row=check_row, column=2).value
        b_val = ev.formula(ws.title, b_chk) if isinstance(b_chk, str) and b_chk.startswith("=") else b_chk
        af_chk = ws.cell(row=check_row, column=32).value
        af_val = (ev.formula(ws.title, af_chk)
                  if isinstance(af_chk, str) and af_chk.startswith("=") else af_chk)
        rep["checks"].append({"check": "cfs_check_figures", "tab": ws.title,
                              "B_check": round(b_val or 0, 2),
                              "AF_check": round(af_val or 0, 2)})

    wb.save(out_path)
    return out_path, rep


if __name__ == "__main__":
    args = sys.argv[1:]
    target = next(a for a in args if a[:2] == "20")
    entities = [a for a in args if a[:2] != "20"] or list(cfg.ENTITIES)
    report = {}
    for ek in entities:
        path, rep = build(ek, target, report=report)
        print(f"[{ek}] -> {path}")
        for m in rep["manual_cells"]:
            print(f"    manual {m['tab']}!{m['cell']}: {m['old']} -> {m['new']}  ({m['derivation']})")
        for f in rep["flags"]:
            print(f"    FLAG: {f}")
        for c in rep["checks"]:
            print(f"    check: {c}")
    with open(os.path.join(cfg.month_dir(target), f"_build_report_{target}.json"),
              "w", encoding="utf-8") as f:
        json.dump(report, f, indent=1)
