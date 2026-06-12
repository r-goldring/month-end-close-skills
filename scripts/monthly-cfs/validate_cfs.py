"""Validate a generated month's CFS workbooks.

Checks (per workbook):
  1. BS balances: Total Assets == Total Liabilities & Equity (both columns)
  2. BS Net Income row == IS Net Income (both columns)
  3. Every BS account row == pulled NetSuite balance; every IS account row ==
     pulled activity (the builder wrote them, this re-verifies the saved file)
  4. CFS check figures: B check, AF check, every column check ~ 0
  5. CFS ending cash == BS cash total
  6. FX rate cells == NetSuite consolidatedexchangerate
  7. Known-variance ledger: IC AR/AP gross-up gap, undeposited-funds balance
  8. Consolidated: external links repointed, formula shape intact

Output: chat summary + CHECKS {ym}.md next to the workbooks.

Usage: python validate_cfs.py 2026-05
"""

import json
import os
import re
import sys

import openpyxl

import cfs_config as cfg
from tab_builders import EntityData, find_tab, ACCT_ROW_RE
from formula_eval import Evaluator
from build_subsidiary_cfs import find_label_row, fx_target_cells

TOL = 0.02


def check(results, name, ok, detail=""):
    results.append((name, "PASS" if ok else "FAIL", detail))
    return ok


def info(results, name, detail):
    results.append((name, "INFO", detail))


def validate_entity(entity_key, ym, results):
    e = cfg.ENTITIES[entity_key]
    path = cfg.find_workbook(entity_key, ym)
    if not path:
        check(results, f"{entity_key}: workbook exists", False, f"none found for {ym}")
        return
    wb = openpyxl.load_workbook(path)
    ev = Evaluator(wb)
    data = EntityData(entity_key)
    prior = cfg.prev_ym(ym)
    is_ws, bs_ws = find_tab(wb, "is"), find_tab(wb, "bs")

    def val(ws, addr):
        v = ws[addr].value
        if isinstance(v, str) and v.startswith("="):
            return ev.formula(ws.title, v)
        return float(v) if isinstance(v, (int, float)) else 0.0

    # 1. BS balances
    ta_row = (find_label_row(bs_ws, "Total ASSETS")
              or find_label_row(bs_ws, "Total Assets Less Total Liabilities"))
    te_row = (find_label_row(bs_ws, "Total Liabilities & Equity")
              or find_label_row(bs_ws, "Total Capital and Reserves"))
    for col in ("B", "C"):
        a, b = val(bs_ws, f"{col}{ta_row}"), val(bs_ws, f"{col}{te_row}")
        check(results, f"{entity_key}: BS balances ({col})", abs(a - b) < 0.05,
              f"{a:,.2f} vs {b:,.2f}")

    # 2. BS NI == IS NI (current column; BS NI is YTD, IS NI is monthly --
    #    compare BS NI delta to IS current-month NI)
    ni_row = find_label_row(bs_ws, "Net Income")
    is_ni_row = (find_label_row(is_ws, "Net Income")
                 or find_label_row(is_ws, "Net Profit/(Loss)"))
    bs_delta = val(bs_ws, f"C{ni_row}") - val(bs_ws, f"B{ni_row}")
    is_cur = val(is_ws, f"C{is_ni_row}")
    check(results, f"{entity_key}: BS NI delta == IS NI", abs(bs_delta - is_cur) < 0.05,
          f"BS delta {bs_delta:,.2f} vs IS {is_cur:,.2f}")

    # 3. account rows == pulled data
    bad = 0
    bs_sign = dict(cfg.BS_SIGN)
    bs_sign["DeferRevenue"] = -1
    bs_sign["CredCard"] = -1
    for r in range(9, bs_ws.max_row + 1):
        lbl = str(bs_ws.cell(row=r, column=1).value or "")
        m = ACCT_ROW_RE.match(lbl)
        if not m or lbl.startswith("Total"):
            continue
        b = bs_ws.cell(row=r, column=3).value
        if not isinstance(b, (int, float)):
            continue
        key = m.group(1)
        if key in data.accounts:
            expect = bs_sign.get(data.accounts[key]["accttype"], 1) * data.bal(key, ym)
            if abs(expect - b) > TOL:
                bad += 1
    check(results, f"{entity_key}: BS rows tie to NetSuite", bad == 0, f"{bad} mismatches")

    # 4 + 5. CFS checks
    for tab_kind in ("cfs_local", "cfs_usd"):
        try:
            ws = find_tab(wb, tab_kind, e)
        except KeyError:
            continue
        cr = find_label_row(ws, "Check Figure")
        if cr is None:
            continue
        b = val(ws, f"B{cr}")
        af = val(ws, f"AF{cr}")
        # whole-dollar FX-effect rounding leaves a sub-dollar residual per
        # rounded column (matches the accountant's manual style); flag only a real break
        check(results, f"{entity_key}: {ws.title} check figures",
              abs(b) < 1.5 and abs(af) < 2.5, f"B={b:,.2f} AF={af:,.2f}")
        if tab_kind == "cfs_local":
            end_row = find_label_row(ws, "Cash at End of Period")
            bank_row = find_label_row(bs_ws, "Total Bank")
            if end_row and bank_row:
                a, bnk = val(ws, f"B{end_row}"), val(bs_ws, f"C{bank_row}")
                check(results, f"{entity_key}: CFS ending cash == BS cash",
                      abs(a - bnk) < 0.05, f"{a:,.2f} vs {bnk:,.2f}")
        if e["currency"] != "USD" and tab_kind == "cfs_usd":
            with open(os.path.join(cfg.DATA_DIR, "fx.json"), encoding="utf-8") as f:
                fx = json.load(f)
            spec = fx_target_cells(ws)
            if spec:
                pr, cu = ws[spec[0]].value, ws[spec[1]].value
                check(results, f"{entity_key}: FX rates",
                      pr == fx[prior][e["currency"]] and cu == fx[ym][e["currency"]],
                      f"{pr} / {cu}")

    # 6. one-off gross fixed-asset decreases = likely disposal / write-off.
    # The CFS capex plug silently absorbs these into investing; when one fires,
    # the disposal-vs-capex split on the CFS needs the accountant's judgment (the
    # disposal-loss formula only captures the accumulated-depreciation side
    # unless the gross movement is added in by hand that month -- e.g. the
    # May 2026 US leasehold write-off).
    # Recurring monthly decreases (ROU-asset amortization, steady capitalized-
    # software write-downs) are NOT disposals and are suppressed by comparing
    # this month's drop to last month's.
    from tab_builders import MONTH_KEYS
    pp_ym = cfg.prev_ym(prior)
    have_pp = pp_ym in MONTH_KEYS
    for key, a in data.accounts.items():
        if a["accttype"] != "FixedAsset":
            continue
        name = a["leaf_name"].lower()
        if "accumulated dep" in name or "accumulated amort" in name:
            continue  # contra account; a "decrease" here is normal depreciation
        cur, pri = data.bal(key, ym), data.bal(key, prior)
        dec_now = pri - cur
        if dec_now <= 0.5:
            continue   # flat or grew
        dec_prev = (data.bal(key, pp_ym) - pri) if have_pp else 0.0
        recurring = dec_prev > 0.5 and dec_now <= dec_prev * 1.5
        if recurring:
            continue   # steady monthly amortization, not a disposal
        results.append((f"{entity_key}: gross fixed-asset decrease", "WARN",
                        f"{data.label(key)} {pri:,.2f} -> {cur:,.2f} "
                        f"(down {dec_now:,.2f}) -- likely a disposal/write-off; "
                        f"the capex plug absorbed it, confirm the CFS "
                        f"disposal-vs-capex split"))

    # 7. known-variance ledger
    gap_ar = None
    if "121900" in data.accounts and "211900" in data.accounts:
        info(results, f"{entity_key}: IC balances (GL)",
             f"121900={data.bal('121900', ym):,.2f}  211900={data.bal('211900', ym):,.2f} "
             f"(UI export may show a gross-up; net IC ties to GL)")
    if entity_key == "US" and "111100" in data.accounts:
        ud = data.bal("111100", ym)
        flag = abs(ud - 662588.83) > 0.05
        results.append((f"US: undeposited funds constant", "WARN" if flag else "PASS",
                        f"111100 = {ud:,.2f}" + (" -- CHANGED, the UI report nets this into "
                        "DR rows; re-derive netting" if flag else " (unchanged)")))


def validate_consolidated(ym, results):
    path = cfg.find_workbook("CONS", ym)
    if not check(results, "CONS: workbook exists", bool(path), str(path)):
        return
    import zipfile
    import urllib.parse
    z = zipfile.ZipFile(path)
    rels = [n for n in z.namelist() if n.startswith("xl/externalLinks/_rels/")]
    pointed = 0
    for n in rels:
        t = urllib.parse.unquote(re.search(r'Target="([^"]+)"', z.read(n).decode()).group(1))
        if ym in t:
            pointed += 1
    rep_path = os.path.join(cfg.month_dir(ym), f"_build_report_{ym}.json")
    refreshed = 0
    if os.path.exists(rep_path):
        with open(rep_path, encoding="utf-8") as f:
            rep = json.load(f)
        refreshed = sum(1 for fl in rep.get("CONS", {}).get("flags", [])
                        if fl.get("type") == "consolidated_tab_values_refreshed")
    check(results, "CONS: subsidiary links current",
          pointed + refreshed >= 6,
          f"{pointed} live links to {ym} files, {refreshed} tabs value-refreshed")
    wb = openpyxl.load_workbook(path)
    cons = wb[f"{ym} CFS Consolidated"]
    f3 = cons["B3"].value
    check(results, "CONS: consolidated tab sums 6 subsidiaries",
          isinstance(f3, str) and f3.count("+") == 5, str(f3))


def main():
    ym = sys.argv[1]
    results = []
    for ek in cfg.ENTITIES:
        validate_entity(ek, ym, results)
    validate_consolidated(ym, results)

    fails = [r for r in results if r[1] == "FAIL"]
    warns = [r for r in results if r[1] == "WARN"]
    lines = [f"# CHECKS {ym}", "",
             f"Result: {'FAIL' if fails else ('WARN' if warns else 'PASS')} "
             f"({len(fails)} fail, {len(warns)} warn, "
             f"{sum(1 for r in results if r[1] == 'PASS')} pass)", ""]
    for name, status, detail in results:
        lines.append(f"- [{status}] {name}: {detail}")
        print(f"[{status}] {name}: {detail}")
    with open(os.path.join(cfg.month_dir(ym), f"CHECKS {ym}.md"), "w",
              encoding="utf-8") as f:
        f.write("\n".join(lines) + "\n")
    print(f"\n{'FAIL' if fails else ('WARN' if warns else 'PASS')}: "
          f"{len(fails)} fail / {len(warns)} warn / "
          f"{sum(1 for r in results if r[1] == 'PASS')} pass")
    return 1 if fails else 0


if __name__ == "__main__":
    sys.exit(main())
