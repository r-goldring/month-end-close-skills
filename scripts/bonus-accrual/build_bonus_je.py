#!/usr/bin/env python3
"""
Acme Corp Monthly Bonus Accrual JE Builder.

Reads FP&A's monthly bonus accrual Excel workbook and produces a NetSuite-ready
Journal Entry CSV file covering all five subsidiaries (US, Canada, Netherlands,
UK, Uruguay). Each subsidiary becomes a separate External ID / JE, balanced
within itself, all reversing on 12/31 of the same year.

Two invocation modes:

1) Folder mode (preferred — matches the accountant's drop-and-run workflow):
     python build_bonus_je.py --folder "Monthly Bonus Accrual/2026-05 May 2026"

   The script looks for one `*_Bonus Accrual_*.xlsx` file in that folder,
   infers the period from the folder name (2026-05 -> 2605), and writes the
   output CSV alongside the input file.

2) Explicit mode (for ad-hoc / re-runs):
     python build_bonus_je.py --file "/path/to/Bonus Accrual.xlsx" \
                              --period 2605 \
                              --tab with-csm \
                              --output "/path/to/output.csv"
"""

import argparse
import calendar
import csv
import glob
import json
import os
import re
import sys
from decimal import Decimal, ROUND_HALF_UP
from itertools import groupby
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl is required.  Run: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


def r2(x):
    """Round to 2 decimal places using ROUND_HALF_UP."""
    if x is None:
        return 0.0
    return float(Decimal(str(x)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))


# Entity master data — column index is 0-based into the workbook row.
ENTITIES = {
    "US":  {"prefix": "USB",  "subsidiary": "Acme Holdings : Acme, Inc.",                          "currency": "USD", "pt_rate": 0.09,    "col": 21},
    "CAD": {"prefix": "CANB", "subsidiary": "Acme Holdings : Acme, Inc. : Acme Canada",       "currency": "CAD", "pt_rate": 0.10,    "col": 17},
    "NL":  {"prefix": "NLB",  "subsidiary": "Acme Holdings : Acme, Inc. : Acme Netherlands",  "currency": "EUR", "pt_rate": 0.0,     "col": 18},
    "UK":  {"prefix": "UKB",  "subsidiary": "Acme Holdings : Acme, Inc. : Acme UK Ltd",       "currency": "GBP", "pt_rate": 0.138,   "col": 20},
    "URY": {"prefix": "URYB", "subsidiary": "Acme Holdings : Acme, Inc. : Acme Uruguay",      "currency": "UYU", "pt_rate": 0.12625, "col": 22},
}
ENTITY_ORDER = ["US", "CAD", "NL", "UK", "URY"]

ACCTS = {
    "511150": "511150 COGS - Salary and Compensation : COGS - Bonus",
    "611150": "611150 Salary and Compensation : Bonus",
    "511350": "511350 COGS - Salary and Compensation : COGS - Payroll Taxes",
    "611450": "611450 Salary and Compensation : Payroll Taxes",
    "231170": "231170 Accrued Bonus",
    "231171": "231171 Accrued Bonus Payroll Tax Liability",
}

# Department rows for the "Bonus - with CSM" tab.
# Each tuple: (workbook_row_1indexed, ns_dept_path, bonus_acct, pt_acct)
DEPT_ROWS_WITH_CSM = [
    ( 6, "Engineering : Infrastructure",        "511150", "511350"),
    ( 7, "COGS : Professional Services",                                "511150", "511350"),
    ( 8, "COGS : Consulting",                                           "511150", "511350"),
    ( 9, "COGS : SRE",                                                  "511150", "511350"),
    (10, "COGS : Managed Survey QA",                                    "511150", "511350"),
    (13, "Engineering : Infrastructure",        "611150", "611450"),
    (14, "Sales & Marketing : Marketing",                               "611150", "611450"),
    (15, "Sales & Marketing : Revenue",                                 "611150", "611450"),
    (16, "Research & Development : Technology : SW Quality Engineering","611150", "611450"),
    (17, "Sales & Marketing : Revenue : New Business",                  "611150", "611450"),
    (18, "Research & Development : Product",                            "611150", "611450"),
    (20, "General & Administrative : General Mgmt and Growth",          "611150", "611450"),
    (21, "General & Administrative : People",                           "611150", "611450"),
    (22, "Research & Development : Technology : Engineering",           "611150", "611450"),
    (23, "Sales & Marketing : Revenue : Solutions Consulting",          "611150", "611450"),
    (24, "Sales & Marketing : Customer Enablement Ops",                 "611150", "611450"),
    (25, "Sales & Marketing : Customer Success",                        "611150", "611450"),
    (26, "Sales & Marketing : Revenue : Customer Success Mgmt",         "611150", "611450"),
    (27, "Research & Development : Technology : Data Science Analytics","611150", "611450"),
    (28, "Research & Development : Technology : Info Security Privacy", "611150", "611450"),
    (29, "Sales & Marketing : Marketing : Workforce Transformation",    "611150", "611450"),
    (30, "Sales & Marketing : Revenue : Business Operations",           "611150", "611450"),
    (31, "Sales & Marketing : Revenue : Strategic Expansion",           "611150", "611450"),
    (32, "General & Administrative : Information Technology",           "611150", "611450"),
    (33, "Sales & Marketing : Marketing : Sales Development",           "611150", "611450"),
    (34, "Sales & Marketing : Revenue Enablement",                      "611150", "611450"),
    (35, "General & Administrative : GA",                               "611150", "611450"),
    (36, "General & Administrative : Legal",                            "611150", "611450"),
    (37, "Research & Development : Technology",                         "611150", "611450"),
    (38, "EBITDA Adjustments",                                          "611150", "611450"),
]

# "Bonus - excluding CSM" tab has no CSM row; rows after Customer Success
# shift up by 1.
DEPT_ROWS_WITHOUT_CSM = [
    (row if row <= 25 else row - 1, dept, bonus_acct, pt_acct)
    for (row, dept, bonus_acct, pt_acct) in DEPT_ROWS_WITH_CSM
    if dept != "Sales & Marketing : Revenue : Customer Success Mgmt"
]

CSV_HEADERS = [
    "External ID", "Journal Entry Memo", "Line Memo", "Date", "Reversal Date",
    "Subsidiary", "Department", "Account", "Currency", "Debit",
]

ENTITY_MEMO_NAMES = {
    "US": "US", "CAD": "Canada", "NL": "Netherlands", "UK": "UK", "URY": "URY",
}


def period_to_dates(period):
    """YYMM ('2605') -> ('5/31/2026', '12/31/2026')."""
    yy = int(period[:2])
    mm = int(period[2:])
    yyyy = 2000 + yy
    last_day = calendar.monthrange(yyyy, mm)[1]
    return f"{mm}/{last_day}/{yyyy}", f"12/31/{yyyy}"


def folder_name_to_period(folder_path):
    """Parse 'Monthly Bonus Accrual/2026-05 May 2026' -> '2605'."""
    name = Path(folder_path).name
    m = re.match(r"^(\d{4})-(\d{2})\b", name)
    if not m:
        return None
    yyyy, mm = m.group(1), m.group(2)
    return f"{yyyy[2:]}{mm}"


def autodiscover_input(folder_path):
    """Find a single Bonus Accrual xlsx file in the folder."""
    candidates = [p for p in glob.glob(str(Path(folder_path) / "*.xlsx"))
                  if "bonus accrual" in os.path.basename(p).lower()
                  and not os.path.basename(p).startswith("~")]
    if not candidates:
        return None
    if len(candidates) > 1:
        # Prefer the one with 'vF' (final) in the name; otherwise newest mtime.
        vf = [p for p in candidates if "vf" in os.path.basename(p).lower()]
        if vf:
            return sorted(vf, key=os.path.getmtime)[-1]
        return sorted(candidates, key=os.path.getmtime)[-1]
    return candidates[0]


def build_je(excel_path, period, tab, output_path, skip_zeros=True):
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    tab_name = "Bonus - with CSM" if tab == "with-csm" else "Bonus - excluding CSM"
    if tab_name not in wb.sheetnames:
        print(f"ERROR: Tab '{tab_name}' not found.  Available: {wb.sheetnames}", file=sys.stderr)
        sys.exit(1)

    ws = wb[tab_name]
    sheet = list(ws.iter_rows(values_only=True))
    dept_rows = DEPT_ROWS_WITH_CSM if tab == "with-csm" else DEPT_ROWS_WITHOUT_CSM
    je_date, reversal_date = period_to_dates(period)

    all_rows = []
    verification = []

    for entity_key in ENTITY_ORDER:
        ent = ENTITIES[entity_key]
        eid = f"{ent['prefix']}{period}"
        yy, mm = period[:2], period[2:]
        memo = f"20{yy}-{mm} {ENTITY_MEMO_NAMES[entity_key]} Bonus Accrual"

        # Build (dept, bonus_acct, pt_acct, amount) for every dept row.
        bonus_lines = []
        for (row_1, dept, bonus_acct, pt_acct) in dept_rows:
            raw = sheet[row_1 - 1][ent["col"]]
            amt = r2(float(raw)) if raw is not None else 0.0
            bonus_lines.append((dept, bonus_acct, pt_acct, amt))

        # If every line is zero, skip the entity entirely.
        if all(amt == 0.0 for (_, _, _, amt) in bonus_lines):
            continue

        bonus_total = r2(sum(amt for (_, _, _, amt) in bonus_lines))

        def make_row(dept, acct_key, debit):
            return {
                "External ID": eid,
                "Journal Entry Memo": memo,
                "Line Memo": memo,
                "Date": je_date,
                "Reversal Date": reversal_date,
                "Subsidiary": ent["subsidiary"],
                "Department": dept,
                "Account": ACCTS[acct_key],
                "Currency": ent["currency"],
                "Debit": debit,
            }

        # Bonus DR lines (one per dept, including zeros for audit clarity).
        for (dept, bonus_acct, _, amt) in bonus_lines:
            if skip_zeros and amt == 0.0:
                continue
            all_rows.append(make_row(dept, bonus_acct, amt))

        # Bonus CR line (231170).
        all_rows.append(make_row("", "231170", r2(-bonus_total)))

        # Payroll-tax lines (skip if entity has 0% rate).
        pt_total = 0.0
        if ent["pt_rate"] > 0:
            for (dept, _, pt_acct, bonus_amt) in bonus_lines:
                pt_amt = r2(bonus_amt * ent["pt_rate"])
                if skip_zeros and pt_amt == 0.0:
                    continue
                all_rows.append(make_row(dept, pt_acct, pt_amt))
                pt_total = r2(pt_total + pt_amt)
            if pt_total != 0.0:
                all_rows.append(make_row("", "231171", r2(-pt_total)))

        verification.append({
            "external_id": eid,
            "currency": ent["currency"],
            "bonus_total": bonus_total,
            "pt_total": pt_total,
            "subsidiary": ent["subsidiary"],
        })

    # Write CSV
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=CSV_HEADERS)
        writer.writeheader()
        writer.writerows(all_rows)

    # Verify each entity balances to zero.
    balance_errors = []
    for eid_key, grp in groupby(all_rows, key=lambda r: r["External ID"]):
        net = sum(float(r["Debit"]) for r in grp)
        if abs(net) > 0.02:
            balance_errors.append({"external_id": eid_key, "net": net})

    return {
        "output_path": str(output_path),
        "period": period,
        "tab": tab,
        "je_date": je_date,
        "reversal_date": reversal_date,
        "row_count": len(all_rows),
        "entities": verification,
        "balance_errors": balance_errors,
        "balanced": len(balance_errors) == 0,
    }


def print_summary(result):
    print(f"\n{'='*60}")
    print(f"  Bonus Accrual JE — Period {result['period']} — Tab: {result['tab']}")
    print(f"{'='*60}")
    print(f"  Output:   {result['output_path']}")
    print(f"  Date:     {result['je_date']}   Reversal: {result['reversal_date']}")
    print(f"  Rows:     {result['row_count']}")
    print(f"\n  Entity summary:")
    for e in result['entities']:
        print(f"  {e['external_id']:12s} ({e['currency']})  bonus={e['bonus_total']:>14,.2f}  PT={e['pt_total']:>12,.2f}")
    if result['balance_errors']:
        print("\n  BALANCE ERRORS:")
        for e in result['balance_errors']:
            print(f"  !! {e['external_id']}: net = {e['net']:.4f}")
    else:
        print("\n  OK — all entities balance to zero.")
    print()


def main():
    p = argparse.ArgumentParser(description="Build Acme Corp Bonus Accrual JE CSV")
    p.add_argument("--folder", help="Month folder containing the .xlsx (auto-detects file and period)")
    p.add_argument("--file",   help="Explicit path to bonus accrual .xlsx (overrides --folder)")
    p.add_argument("--period", help="YYMM (e.g. 2605). Required unless --folder is supplied.")
    p.add_argument("--tab",    default="with-csm", choices=["with-csm", "without-csm"])
    p.add_argument("--output", help="Output CSV path (default: same folder as input)")
    p.add_argument("--keep-zeros", action="store_true",
                   help="Keep zero-amount lines in the CSV (default omits them for a cleaner upload)")
    p.add_argument("--json", action="store_true", help="Emit machine-readable summary JSON on stdout")
    args = p.parse_args()

    # Resolve input
    if args.file:
        excel_path = args.file
    elif args.folder:
        excel_path = autodiscover_input(args.folder)
        if not excel_path:
            print(f"ERROR: no *Bonus Accrual*.xlsx found in {args.folder}", file=sys.stderr)
            sys.exit(1)
    else:
        print("ERROR: provide either --folder or --file", file=sys.stderr)
        sys.exit(1)

    if not os.path.isfile(excel_path):
        print(f"ERROR: file not found: {excel_path}", file=sys.stderr)
        sys.exit(1)

    # Resolve period
    period = args.period
    if not period and args.folder:
        period = folder_name_to_period(args.folder)
    if not period:
        print("ERROR: could not determine period; pass --period YYMM", file=sys.stderr)
        sys.exit(1)

    # Resolve output
    if args.output:
        output_path = args.output
    else:
        yy, mm = period[:2], period[2:]
        default_name = f"20{yy}-{mm} Bonus Accrual JE Import.csv"
        output_dir = Path(args.folder) if args.folder else Path(excel_path).parent
        output_path = str(output_dir / default_name)

    result = build_je(excel_path, period, args.tab, output_path,
                      skip_zeros=not args.keep_zeros)
    result["input_path"] = excel_path

    if args.json:
        print(json.dumps(result, indent=2))
    else:
        print_summary(result)


if __name__ == "__main__":
    main()
