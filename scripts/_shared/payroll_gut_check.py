"""
Payroll Gut-Check Review — pre-upload validation of a generated JE Import CSV
against the prior 2 same-skill JEs in NetSuite.

Architecture: this module contains all Python-side logic (CSV parse, audit-log
walk, line classification, variance computation, special checks, output
formatting). NetSuite SuiteQL execution is delegated back to the caller
(the gut-check skill), which uses ns_runCustomSuiteQL via MCP. The skill
orchestrates the two-phase flow:

  Phase 1 (this module):     parse CSV, find prior candidates, build SQL strings
  Caller (skill):            execute validate_sql via MCP
  Phase 2a (this module):    filter to valid priors, build line-fetch SQL list
  Caller (skill):            execute each line-fetch SQL via MCP
  Phase 2b (this module):    run comparison, format report, write workbook tab

Public API:
  parse_csv(csv_path)                                   -> list[LineDict]
  find_prior_candidates(audit_log_path, skill, n=8,
                        exclude_csv_path=None)           -> list[dict]
  build_validation_sql(candidate_tranids)                -> str
  filter_validated_priors(candidates, validation_rows,
                          cadence_days, n=2)             -> {priors, warnings}
  build_line_fetch_sql(je_number, subsidiary)            -> str
  analyze(current_lines, prior_lines_by_je, config)      -> Findings
  format_chat_report(findings, header_meta, csv_path)    -> str
  write_workbook_tab(workbook_path, findings, header_meta) -> tab_name

LineDict shape: {date, account, account_num, debit, credit, signed_amount,
                 line_memo, subsidiary, department, dept_normalized, family}
"""

from __future__ import annotations

import csv
import datetime as _dt
import json
import os
import re
from collections import defaultdict
from dataclasses import dataclass, field
from typing import Optional

import openpyxl

# Local imports
import sys
_THIS_DIR = os.path.dirname(os.path.abspath(__file__))
if _THIS_DIR not in sys.path:
    sys.path.insert(0, _THIS_DIR)

from skill_configs import (  # noqa: E402
    SKILL_CONFIGS, THRESHOLDS, JB_THRESHOLDS,
    classify_account, canonical_sign, normalize_dept,
    detect_skill_from_folder,
)


# ============================================================================
# Data classes
# ============================================================================

@dataclass
class Finding:
    tier: str               # "T1" | "T2" | "SPECIAL"
    severity: str           # "FAIL" | "WARN" | "PASS" | "INFO"
    check: str              # short identifier (e.g., "variance", "sign_flip")
    department: Optional[str]
    account: Optional[str]
    current: Optional[float]
    prior_1: Optional[float] = None
    prior_2: Optional[float] = None
    mean_prior: Optional[float] = None
    delta: Optional[float] = None
    delta_pct: Optional[float] = None
    threshold: Optional[str] = None
    message: str = ""
    suggested_action: str = ""


@dataclass
class GutCheckResult:
    findings: list[Finding] = field(default_factory=list)
    counts: dict = field(default_factory=lambda: {"FAIL": 0, "WARN": 0, "PASS": 0, "INFO": 0})
    has_fail: bool = False

    def add(self, f: Finding):
        self.findings.append(f)
        self.counts[f.severity] = self.counts.get(f.severity, 0) + 1
        if f.severity == "FAIL":
            self.has_fail = True


# ============================================================================
# Phase 1 — CSV parsing
# ============================================================================

def parse_csv(csv_path: str) -> list[dict]:
    """Read a JE Import CSV and return a list of line dicts.

    Expected columns: Date, Journal Entry Memo, Account, Debit, Credit,
    Line Memo, Subsidiary, Department.

    Each output line gets:
      - signed_amount: positive for debit, negative for credit
      - account_num: leading numeric token from Account
      - dept_normalized: lowercased + Mgmt-collapsed dept key
      - family: classified account family (e.g., 'salaries', 'liab_payroll')
    """
    rows: list[dict] = []
    with open(csv_path, encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        for raw in reader:
            account = (raw.get("Account") or "").strip()
            debit_str = (raw.get("Debit") or "").strip()
            credit_str = (raw.get("Credit") or "").strip()
            debit = float(debit_str) if debit_str else 0.0
            credit = float(credit_str) if credit_str else 0.0
            signed = round(debit - credit, 2)
            dept = (raw.get("Department") or "").strip()
            rows.append({
                "date": (raw.get("Date") or "").strip(),
                "memo_header": (raw.get("Journal Entry Memo") or "").strip(),
                "account": account,
                "account_num": account.split()[0] if account else "",
                "debit": debit,
                "credit": credit,
                "signed_amount": signed,
                "line_memo": (raw.get("Line Memo") or "").strip(),
                "subsidiary": (raw.get("Subsidiary") or "").strip(),
                "department": dept,
                "dept_normalized": normalize_dept(dept),
                "family": classify_account(account),
            })
    return rows


# ============================================================================
# Phase 1 — Audit-log walker
# ============================================================================

_TRANID_RE = re.compile(r"^JE\d+$")


def _iter_audit_je_entries(audit_log: list, skill: str):
    """Yield (tranid, internal_id, timestamp, source_csv_path, parent_entry)
    for every JE# attribution to this skill in the audit log.

    Walks both top-level je_number/netsuite_internal_id AND je_numbers[*] arrays
    (for POST_JE_BATCH entries from Uruguay). Does NOT filter on `action`,
    because entries are heterogeneous (POST_JE, GENERATE_CSV+follow-up,
    POST_JE_BATCH, legacy netsuite_id field).
    """
    for entry in audit_log:
        if entry.get("skill") != skill:
            continue
        ts = entry.get("timestamp", "")
        sources = entry.get("source_files") or {}
        csv_path = sources.get("output_csv") if isinstance(sources, dict) else None
        # Top-level
        je = entry.get("je_number") or entry.get("netsuite_id")
        if je and _TRANID_RE.match(str(je)):
            yield {
                "tranid": str(je),
                "internal_id": entry.get("netsuite_internal_id"),
                "timestamp": ts,
                "source_csv": csv_path,
                "memo": entry.get("description", ""),
                "parent": entry,
            }
        # Batch (Uruguay)
        for sub_entry in entry.get("je_numbers", []) or []:
            sje = sub_entry.get("je_number")
            if sje and _TRANID_RE.match(str(sje)):
                yield {
                    "tranid": str(sje),
                    "internal_id": sub_entry.get("netsuite_internal_id"),
                    "timestamp": ts,
                    "source_csv": csv_path,
                    "memo": sub_entry.get("memo", ""),
                    "parent": entry,
                }


def find_prior_candidates(
    audit_log_path: str,
    skill: str,
    n: int = 8,
    exclude_csv_path: Optional[str] = None,
) -> list[dict]:
    """Return up to `n` prior JE candidates for this skill from the audit log,
    sorted by timestamp DESC. Caller must validate them against NetSuite.

    `exclude_csv_path` lets the caller skip an entry whose source_csv matches
    the current pay-run CSV (for the dry-run case where the JE is already in
    the audit log but represents the *current* pay run, not a prior).
    """
    with open(audit_log_path, encoding="utf-8") as f:
        log = json.load(f)
    candidates = []
    seen_tranids = set()
    for c in _iter_audit_je_entries(log, skill):
        if c["tranid"] in seen_tranids:
            continue
        if exclude_csv_path and c.get("source_csv"):
            # Normalize path separators for cross-platform comparison
            cur_norm = exclude_csv_path.replace("\\", "/").lower()
            entry_norm = c["source_csv"].replace("\\", "/").lower()
            if cur_norm.endswith(entry_norm) or entry_norm.endswith(cur_norm):
                continue
        seen_tranids.add(c["tranid"])
        candidates.append(c)
    # Sort by timestamp DESC
    candidates.sort(key=lambda c: c["timestamp"], reverse=True)
    return candidates[:n]


# ============================================================================
# Phase 1 — SuiteQL builders
# ============================================================================

def build_validation_sql(candidate_tranids: list[str]) -> str:
    """Build a SuiteQL query that returns the trandate + posting state for
    each candidate tranid. Caller runs this via ns_runCustomSuiteQL.

    Filters to non-voided, posting=T, no reversal_date — voided / reversed
    candidates are dropped on the caller side.
    """
    if not candidate_tranids:
        return ""
    quoted = ", ".join(f"'{t}'" for t in candidate_tranids)
    return (
        "SELECT t.tranid, t.id AS internal_id, t.voided, t.posting, "
        "t.reversaldate, TO_CHAR(t.trandate, 'YYYY-MM-DD') AS trandate, "
        "t.memo "
        "FROM transaction t "
        f"WHERE t.tranid IN ({quoted}) "
        "ORDER BY t.trandate DESC"
    )


def build_priors_search_sql(
    memo_prefix: str,
    subsidiary: str,
    before_date: str,
    n: int = 2,
) -> str:
    """Find prior posted JEs by memo + subsidiary + before-current-date.

    SuiteQL is the authoritative source for priors — the audit_log can be
    stale when JEs are rejected/reposted manually. Memo prefix is
    case-insensitive substring; we filter by subsidiary to prevent
    cross-sub bleed.

    Returns the most recent `n` JEs strictly before `before_date`.
    """
    sub_escaped = subsidiary.replace("'", "''")
    memo_escaped = memo_prefix.replace("'", "''")
    return (
        "SELECT t.id AS internal_id, t.tranid, "
        "TO_CHAR(t.trandate, 'YYYY-MM-DD') AS trandate, "
        "t.memo, t.voided, t.posting, t.reversaldate "
        "FROM transaction t "
        "JOIN transactionline tl ON t.id = tl.transaction "
        "LEFT JOIN subsidiary sub ON tl.subsidiary = sub.id "
        "WHERE t.recordtype = 'journalentry' "
        f"AND UPPER(t.memo) LIKE UPPER('%{memo_escaped}%') "
        f"AND sub.fullname = '{sub_escaped}' "
        f"AND t.trandate < TO_DATE('{before_date}', 'YYYY-MM-DD') "
        "AND t.voided = 'F' "
        "AND t.posting = 'T' "
        "AND t.reversaldate IS NULL "
        "GROUP BY t.id, t.tranid, t.trandate, t.memo, t.voided, t.posting, t.reversaldate "
        f"ORDER BY t.trandate DESC FETCH FIRST {n} ROWS ONLY"
    )


def build_line_fetch_sql(je_number: str, subsidiary: str) -> str:
    """Build a SuiteQL query for the lines of one posted JE, filtered by sub.

    Note: in this NetSuite tenant, JE body lines have mainline='T' (the typical
    'F' filter returns zero rows for journal entries). We don't filter on
    mainline; instead we filter on tl.account being non-null, which excludes
    any pure-header rows.
    """
    sub_escaped = subsidiary.replace("'", "''")
    return (
        "SELECT t.id AS internal_id, t.tranid, "
        "TO_CHAR(t.trandate, 'YYYY-MM-DD') AS trandate, "
        "t.memo AS header_memo, t.voided, t.posting, t.reversaldate, "
        "a.acctnumber, a.fullname AS account_name, "
        "tl.memo AS line_memo, "
        "tl.debitforeignamount AS debit, "
        "tl.creditforeignamount AS credit, "
        "sub.fullname AS subsidiary, "
        "d.fullname AS department, "
        "tl.linesequencenumber "
        "FROM transaction t "
        "JOIN transactionline tl ON t.id = tl.transaction "
        "JOIN account a ON tl.account = a.id "
        "LEFT JOIN subsidiary sub ON tl.subsidiary = sub.id "
        "LEFT JOIN department d ON tl.department = d.id "
        f"WHERE t.tranid = '{je_number}' "
        f"AND sub.fullname = '{sub_escaped}' "
        "AND t.voided = 'F' "
        "AND tl.account IS NOT NULL "
        "ORDER BY tl.linesequencenumber"
    )


def normalize_ns_lines(suiteql_rows: list[dict]) -> list[dict]:
    """Convert raw SuiteQL row dicts into the same shape as parse_csv output."""
    out = []
    for r in suiteql_rows:
        debit = float(r.get("debit") or 0)
        credit = float(r.get("credit") or 0)
        signed = round(debit - credit, 2)
        # NetSuite returns acctnumber as the leading digits; account_name has the rest
        acct_num = (r.get("acctnumber") or "").strip()
        acct_name = (r.get("account_name") or "").strip()
        # Reconstruct the full account string used in CSVs ("611100 Salary and Compensation : ...")
        # NS fullname already includes parent path; CSV uses "{number} {fullname}".
        # We classify by acct_num so the format here is just for display.
        if acct_num and acct_name and not acct_name.startswith(acct_num):
            full = f"{acct_num} {acct_name}"
        else:
            full = acct_name or acct_num
        dept = (r.get("department") or "").strip()
        out.append({
            "date": r.get("trandate", ""),
            "memo_header": r.get("header_memo", ""),
            "account": full,
            "account_num": acct_num,
            "debit": debit,
            "credit": credit,
            "signed_amount": signed,
            "line_memo": (r.get("line_memo") or "").strip(),
            "subsidiary": (r.get("subsidiary") or "").strip(),
            "department": dept,
            "dept_normalized": normalize_dept(dept),
            "family": classify_account(full),
        })
    return out


# ============================================================================
# Phase 2a — Validate priors and pick top 2
# ============================================================================

def filter_validated_priors(
    candidates: list[dict],
    validation_rows: list[dict],
    current_trandate: Optional[str],
    cadence_days: int,
    n: int = 2,
) -> dict:
    """Walk candidates DESC by timestamp; keep first N that NetSuite confirms
    are not voided / not reversed / posting='T'.

    Returns:
      {
        "priors":   list[dict] of length 0..N (with NetSuite trandate added),
        "warnings": list[str] (stale baseline, only-1-prior, etc.),
        "rejected": list[dict] (with reason) for transparency,
      }
    """
    # Build lookup from validation rows (only "valid" rows passed those filters)
    valid_by_tranid = {}
    for v in validation_rows:
        valid_by_tranid[v.get("tranid")] = v

    chosen = []
    rejected = []
    for c in candidates:
        v = valid_by_tranid.get(c["tranid"])
        if not v:
            rejected.append({**c, "reason": "not in validation result (likely voided/reversed/non-posting)"})
            continue
        # Defensive: even though SQL filters voided/posting/reversaldate, double-check
        if str(v.get("voided", "F")).upper() == "T":
            rejected.append({**c, "reason": "voided=T"})
            continue
        if str(v.get("posting", "T")).upper() != "T":
            rejected.append({**c, "reason": f"posting={v.get('posting')}"})
            continue
        if v.get("reversaldate") not in (None, "", "null"):
            rejected.append({**c, "reason": f"reversaldate={v.get('reversaldate')}"})
            continue
        c["ns_trandate"] = v.get("trandate", "")
        c["ns_memo"] = v.get("memo", "")
        chosen.append(c)
        if len(chosen) >= n:
            break

    warnings = []
    if len(chosen) < n:
        warnings.append(
            f"Only {len(chosen)} valid prior(s) available "
            f"(needed {n}); comparisons will be best-effort."
        )

    # Stale-baseline check: gap between current and most recent prior
    if chosen and current_trandate:
        try:
            cur_dt = _dt.datetime.strptime(current_trandate, "%Y-%m-%d").date()
            top_dt = _dt.datetime.strptime(chosen[0]["ns_trandate"], "%Y-%m-%d").date()
            gap = (cur_dt - top_dt).days
            if gap > 1.5 * cadence_days:
                warnings.append(
                    f"Stale baseline: most recent prior is {gap} days old "
                    f"(expected ~{cadence_days})."
                )
        except (ValueError, TypeError):
            pass  # bad date format; skip the gate

    return {"priors": chosen, "warnings": warnings, "rejected": rejected}


# ============================================================================
# Phase 2b — Comparison engine
# ============================================================================

def _line_key(line: dict) -> tuple:
    """(dept_normalized, account_num) — the per-dept-account aggregation key."""
    return (line["dept_normalized"], line["account_num"])


def _aggregate_by_family(lines: list[dict]) -> dict:
    """Sum signed_amount per account family (across all depts)."""
    out = defaultdict(float)
    for ln in lines:
        fam = ln.get("family")
        if fam:
            out[fam] += ln["signed_amount"]
    return {k: round(v, 2) for k, v in out.items()}


def _aggregate_by_dept_account(lines: list[dict]) -> dict:
    """Sum signed_amount per (dept_normalized, account_num)."""
    out = defaultdict(float)
    for ln in lines:
        out[_line_key(ln)] += ln["signed_amount"]
    return {k: round(v, 2) for k, v in out.items()}


def _mean_of(values: list[float]) -> float:
    if not values:
        return 0.0
    return sum(values) / len(values)


def _fmt_money(v: Optional[float]) -> str:
    if v is None:
        return ""
    return f"${v:,.2f}" if v >= 0 else f"-${abs(v):,.2f}"


def _fmt_pct(v: Optional[float]) -> str:
    if v is None:
        return ""
    return f"{v*100:+.1f}%"


def _check_variance(
    current: float,
    priors: list[float],
    floor: float,
    pct: float,
) -> Optional[dict]:
    """Return dict {delta, mean, delta_pct, breach: bool, threshold: str}
    if the line should be flagged; None if within tolerance."""
    if not priors:
        return None
    mean = _mean_of(priors)
    delta = round(current - mean, 2)
    abs_delta = abs(delta)
    threshold_dollar = floor
    threshold_pct = pct * abs(mean)
    threshold = max(threshold_dollar, threshold_pct)
    delta_pct = (delta / mean) if abs(mean) > 0.01 else None
    return {
        "current": round(current, 2),
        "priors": priors,
        "mean": round(mean, 2),
        "delta": delta,
        "delta_pct": delta_pct,
        "threshold": threshold,
        "threshold_str": f"max(${floor:,.0f}, {pct*100:.0f}% mean)",
        "breach": abs_delta > threshold,
    }


def analyze(
    current_lines: list[dict],
    prior_lines_by_je: dict,  # {je_number: list[line_dict]}
    config: dict,
) -> GutCheckResult:
    """Run all checks and return a GutCheckResult."""
    result = GutCheckResult()

    # Convert priors to per-family + per-(dept,acct) maps
    prior_fam_maps = []
    prior_dept_maps = []
    for je_num, lines in prior_lines_by_je.items():
        prior_fam_maps.append(_aggregate_by_family(lines))
        prior_dept_maps.append(_aggregate_by_dept_account(lines))

    cur_fam = _aggregate_by_family(current_lines)
    cur_dept = _aggregate_by_dept_account(current_lines)

    # ---- Tier 1: aggregate variance per family ----
    all_families = set(cur_fam.keys()) | {f for m in prior_fam_maps for f in m}
    for family in sorted(all_families):
        spec = THRESHOLDS.get(family)
        if not spec or spec.get("skip_variance"):
            continue
        cur_val = cur_fam.get(family, 0.0)
        prior_vals = [m.get(family, 0.0) for m in prior_fam_maps]
        if not prior_vals:
            continue
        v = _check_variance(cur_val, prior_vals, spec["agg_floor"], spec["pct"])
        if v is None:
            continue
        if v["breach"]:
            direction = "up" if v["delta"] > 0 else "down"
            msg = (
                f"{family.replace('_', ' ').title()} {direction} "
                f"{_fmt_pct(v['delta_pct']) if v['delta_pct'] is not None else 'N/A'} "
                f"vs prior 2 mean (delta {_fmt_money(v['delta'])})"
            )
            result.add(Finding(
                tier="T1", severity="WARN", check="variance",
                department=None,
                account=family.replace("_", " ").title(),
                current=v["current"],
                prior_1=prior_vals[0] if len(prior_vals) > 0 else None,
                prior_2=prior_vals[1] if len(prior_vals) > 1 else None,
                mean_prior=v["mean"],
                delta=v["delta"],
                delta_pct=v["delta_pct"],
                threshold=v["threshold_str"],
                message=msg,
                suggested_action="Verify expected business change (new hire, comp adjustment, FX move)",
            ))
        else:
            result.add(Finding(
                tier="T1", severity="PASS", check="variance",
                department=None,
                account=family.replace("_", " ").title(),
                current=v["current"],
                prior_1=prior_vals[0] if len(prior_vals) > 0 else None,
                prior_2=prior_vals[1] if len(prior_vals) > 1 else None,
                mean_prior=v["mean"],
                delta=v["delta"],
                delta_pct=v["delta_pct"],
                threshold=v["threshold_str"],
                message="Within threshold",
            ))

    # ---- Tier 2: per-(dept, account) variance + new/missing ----
    all_keys = set(cur_dept.keys()) | {k for m in prior_dept_maps for k in m}
    for key in sorted(all_keys):
        dept, acct = key
        cur_val = cur_dept.get(key, 0.0)
        prior_vals_per = [m.get(key, 0.0) for m in prior_dept_maps]
        family = classify_account(acct)
        if not family:
            continue
        spec = THRESHOLDS.get(family)
        if not spec:
            continue
        # New combo (current has it, neither prior had non-zero)
        if cur_val and not any(prior_vals_per):
            result.add(Finding(
                tier="T2", severity="WARN", check="new_combo",
                department=dept or "(no dept)",
                account=acct,
                current=round(cur_val, 2),
                message=f"NEW combo - dept/account did not exist in prior 2",
                suggested_action="Verify new hire / re-classed employee / new dept assignment",
            ))
            continue
        # Disappeared combo (prior had it, current zero)
        if not cur_val and all(v != 0 for v in prior_vals_per) and prior_vals_per:
            result.add(Finding(
                tier="T2", severity="WARN", check="missing_combo",
                department=dept or "(no dept)",
                account=acct,
                current=0.0,
                prior_1=prior_vals_per[0] if len(prior_vals_per) > 0 else None,
                prior_2=prior_vals_per[1] if len(prior_vals_per) > 1 else None,
                mean_prior=_mean_of(prior_vals_per),
                message="DISAPPEARED - line existed in both prior 2 but absent in current",
                suggested_action="Verify dept transfer / termination",
            ))
            continue
        if spec.get("skip_variance"):
            continue
        if spec["dept_floor"] is None:
            # Aggregate-only family (liability) — already handled in Tier 1
            continue
        v = _check_variance(cur_val, prior_vals_per, spec["dept_floor"], spec["pct"])
        if v is None or not v["breach"]:
            continue  # PASS lines for Tier 2 are not emitted (too noisy); only flagged WARNs surface
        direction = "up" if v["delta"] > 0 else "down"
        msg = (
            f"{family.replace('_', ' ').title()} {direction} "
            f"{_fmt_pct(v['delta_pct']) if v['delta_pct'] is not None else 'N/A'} "
            f"in {dept or '(no dept)'}"
        )
        result.add(Finding(
            tier="T2", severity="WARN", check="variance",
            department=dept or "(no dept)",
            account=acct,
            current=v["current"],
            prior_1=prior_vals_per[0] if len(prior_vals_per) > 0 else None,
            prior_2=prior_vals_per[1] if len(prior_vals_per) > 1 else None,
            mean_prior=v["mean"],
            delta=v["delta"],
            delta_pct=v["delta_pct"],
            threshold=v["threshold_str"],
            message=msg,
            suggested_action="Verify the dept-level change (hire, comp adj, severance, etc.)",
        ))

    # ---- Special: canonical sign check ----
    # Two paths:
    #   (a) accounts with a fixed canonical sign (Salaries, Bonus, Commission,
    #       Payroll Tax, Liabilities, etc.): violation is an outright FAIL.
    #   (b) net accounts (Health Benefits, Other Benefits): no canonical sign,
    #       so we run a prior-2 flip check on the per-(dept, account) line
    #       — if both prior 2 had one sign and current has the opposite,
    #       it's a likely controller-typo sign flip.
    for ln in current_lines:
        canonical = canonical_sign(ln["account"])
        actual = "debit" if ln["signed_amount"] > 0 else ("credit" if ln["signed_amount"] < 0 else None)
        if actual is None:
            continue  # zero-amount line; skip
        if canonical != "unknown":
            # Path (a): fixed canonical
            if actual != canonical:
                result.add(Finding(
                    tier="SPECIAL", severity="FAIL", check="canonical_sign",
                    department=ln["department"] or "(no dept)",
                    account=ln["account"],
                    current=ln["signed_amount"],
                    message=(
                        f"Sign violation: account {ln['account_num']} canonical sign is "
                        f"{canonical.upper()} but line is posted as {actual.upper()} "
                        f"(memo: '{ln['line_memo']}')"
                    ),
                    suggested_action=(
                        "Edit the CSV (or upstream Reclasses tab) to correct the sign, "
                        "then re-run the skill or /gut-check."
                    ),
                ))
        else:
            # Path (b): net account — prior-2 flip check on the same (dept, account)
            key = _line_key(ln)
            prior_signs = []
            for prior_map in prior_dept_maps:
                v = prior_map.get(key)
                if v is None or abs(v) < 0.01:
                    continue
                prior_signs.append("debit" if v > 0 else "credit")
            # Only flag if BOTH priors had the same sign and current is opposite
            if len(prior_signs) >= 2 and len(set(prior_signs)) == 1 and prior_signs[0] != actual:
                result.add(Finding(
                    tier="SPECIAL", severity="FAIL", check="prior2_sign_flip",
                    department=ln["department"] or "(no dept)",
                    account=ln["account"],
                    current=ln["signed_amount"],
                    message=(
                        f"Prior-2 sign flip: {ln['account_num']} on this dept was "
                        f"{prior_signs[0].upper()} in both prior runs but is "
                        f"{actual.upper()} in current (memo: '{ln['line_memo']}')"
                    ),
                    suggested_action=(
                        "Likely a Reclasses-tab sign typo or pivot-side miscalculation. "
                        "Verify and edit the CSV before uploading."
                    ),
                ))

    # ---- Special: severance routing (US-only) ----
    if "severance_routing" in config.get("special_checks", []):
        for ln in current_lines:
            if ln["account_num"] not in ("611250", "511175"):
                continue
            if ln["dept_normalized"] != "ebitda adjustments":
                result.add(Finding(
                    tier="SPECIAL", severity="FAIL", check="severance_routing",
                    department=ln["department"] or "(no dept)",
                    account=ln["account"],
                    current=ln["signed_amount"],
                    message=(
                        f"Severance line ({ln['account_num']}) must be in 'EBITDA Adjustments' "
                        f"but is in '{ln['department'] or '(no dept)'}'"
                    ),
                    suggested_action="Edit the CSV: set Department to 'EBITDA Adjustments' for this line.",
                ))
            elif "severance" not in ln["line_memo"].lower():
                result.add(Finding(
                    tier="SPECIAL", severity="WARN", check="severance_routing",
                    department=ln["department"],
                    account=ln["account"],
                    current=ln["signed_amount"],
                    message=(
                        f"Severance line memo missing 'Severance' keyword: '{ln['line_memo']}'"
                    ),
                    suggested_action="Verify line memo follows 'US PAYROLL - Severance ({Home Dept})' pattern.",
                ))

    # ---- Special: JB EBITDA reclass prior-2 comparison (US-only) ----
    if "jb_ebitda" in config.get("special_checks", []):
        # Find JB rows in current and priors
        def jb_rows(lines):
            return [
                ln for ln in lines
                if ln["dept_normalized"] == "ebitda adjustments"
                and "(jb)" in ln["line_memo"].lower()
            ]
        cur_jb = jb_rows(current_lines)
        prior_jb_per_je = {je: jb_rows(lns) for je, lns in prior_lines_by_je.items()}

        if not cur_jb:
            result.add(Finding(
                tier="SPECIAL", severity="INFO", check="jb_ebitda",
                department="EBITDA Adjustments",
                account=None,
                current=None,
                message="No (JB) lines found in current JE; skipping JB component check.",
                suggested_action="Verify the assistant controller intentionally omitted Borland this run.",
            ))
        else:
            # Compare per account number
            cur_by_acct = {ln["account_num"]: ln for ln in cur_jb}
            prior_by_acct = defaultdict(list)
            for je, jb_list in prior_jb_per_je.items():
                for ln in jb_list:
                    prior_by_acct[ln["account_num"]].append(ln["signed_amount"])
            all_jb_accts = set(cur_by_acct.keys()) | set(prior_by_acct.keys())
            for acct_num in sorted(all_jb_accts):
                cur_ln = cur_by_acct.get(acct_num)
                prior_vals = prior_by_acct.get(acct_num, [])
                cur_val = cur_ln["signed_amount"] if cur_ln else 0.0
                if cur_ln and not prior_vals:
                    result.add(Finding(
                        tier="SPECIAL", severity="WARN", check="jb_ebitda",
                        department="EBITDA Adjustments",
                        account=cur_ln["account"],
                        current=cur_val,
                        message=f"JB component on {acct_num} is new (no prior history)",
                        suggested_action="Verify with controller — first time this component is in JB reclass.",
                    ))
                    continue
                if not cur_ln and prior_vals:
                    result.add(Finding(
                        tier="SPECIAL", severity="WARN", check="jb_ebitda",
                        department="EBITDA Adjustments",
                        account=acct_num,
                        current=0.0,
                        prior_1=prior_vals[0] if len(prior_vals) > 0 else None,
                        prior_2=prior_vals[1] if len(prior_vals) > 1 else None,
                        mean_prior=_mean_of(prior_vals),
                        message=f"JB component {acct_num} missing in current (was in prior runs)",
                        suggested_action="Edit Reclasses tab to include this JB component, then re-run.",
                    ))
                    continue
                # Both present - compare
                v = _check_variance(cur_val, prior_vals, JB_THRESHOLDS["floor"], JB_THRESHOLDS["pct"])
                # Sign flip check (relative to canonical, already handled, but flag here too)
                # If canonical-sign FAIL didn't fire (it would for credit on 6xxxx), handle drift
                if v and v["breach"]:
                    severity = "WARN"
                    msg = f"JB {acct_num} drift: {_fmt_money(cur_val)} vs {_fmt_money(v['mean'])} mean ({_fmt_pct(v['delta_pct']) if v['delta_pct'] is not None else 'N/A'})"
                    result.add(Finding(
                        tier="SPECIAL", severity=severity, check="jb_ebitda",
                        department="EBITDA Adjustments",
                        account=cur_ln["account"],
                        current=v["current"],
                        prior_1=prior_vals[0] if len(prior_vals) > 0 else None,
                        prior_2=prior_vals[1] if len(prior_vals) > 1 else None,
                        mean_prior=v["mean"],
                        delta=v["delta"],
                        delta_pct=v["delta_pct"],
                        threshold=v["threshold_str"],
                        message=msg,
                        suggested_action="Verify JB Reclasses tab amount with controller.",
                    ))
                else:
                    result.add(Finding(
                        tier="SPECIAL", severity="PASS", check="jb_ebitda",
                        department="EBITDA Adjustments",
                        account=cur_ln["account"],
                        current=v["current"] if v else cur_val,
                        prior_1=prior_vals[0] if len(prior_vals) > 0 else None,
                        prior_2=prior_vals[1] if len(prior_vals) > 1 else None,
                        mean_prior=v["mean"] if v else None,
                        delta=v["delta"] if v else None,
                        delta_pct=v["delta_pct"] if v else None,
                        message=f"JB {acct_num} within tolerance",
                    ))

    return result


# ============================================================================
# Phase 2b — Output formatting
# ============================================================================

def format_chat_report(
    result: GutCheckResult,
    header_meta: dict,
) -> str:
    """Render the chat-facing PASS/WARN/FAIL report."""
    lines = []
    lines.append("=" * 73)
    lines.append(f"GUT-CHECK REVIEW - {header_meta['pay_run_label']} ({header_meta['skill']})")
    lines.append(f"CSV: {header_meta['csv_basename']}")
    if header_meta.get("totals"):
        d = header_meta["totals"]
        bal = "balanced" if abs(d['debit'] - d['credit']) < 0.01 else "OUT OF BALANCE"
        lines.append(
            f"     {d['line_count']} lines, "
            f"{d['currency']} {d['debit']:,.2f} debits / {d['credit']:,.2f} credits ({bal})"
        )
    if header_meta.get("priors"):
        plist = ", ".join(f"{p['tranid']} ({p.get('ns_trandate', '?')})" for p in header_meta["priors"])
        lines.append(f"Compared against: {plist}")
    if header_meta.get("baseline_warnings"):
        for w in header_meta["baseline_warnings"]:
            lines.append(f"  ! {w}")
    lines.append("=" * 73)

    counts = result.counts
    by_tier = defaultdict(lambda: defaultdict(int))
    for f in result.findings:
        by_tier[f.tier][f.severity] += 1
    lines.append(
        f"TIER 1 (aggregate):  {by_tier['T1']['PASS']} PASS, "
        f"{by_tier['T1']['WARN']} WARN, {by_tier['T1']['FAIL']} FAIL"
    )
    lines.append(
        f"TIER 2 (per-dept):   {by_tier['T2']['PASS']} PASS (suppressed), "
        f"{by_tier['T2']['WARN']} WARN, {by_tier['T2']['FAIL']} FAIL"
    )
    lines.append(
        f"SPECIAL checks:      {by_tier['SPECIAL']['PASS']} PASS, "
        f"{by_tier['SPECIAL']['WARN']} WARN, "
        f"{by_tier['SPECIAL']['FAIL']} FAIL, "
        f"{by_tier['SPECIAL']['INFO']} INFO"
    )
    lines.append("=" * 73)
    lines.append("")

    # FAIL section (always at top, loud)
    fails = [f for f in result.findings if f.severity == "FAIL"]
    lines.append(f"FAIL ({len(fails)})")
    if fails:
        for f in fails:
            head = f"  {f.tier}  {f.account or ''}"
            if f.department:
                head += f" @ {f.department}"
            lines.append(head)
            lines.append(f"      {f.message}")
            if f.suggested_action:
                lines.append(f"      Fix: {f.suggested_action}")
            if f.current is not None and f.mean_prior is not None:
                lines.append(
                    f"      Current: {_fmt_money(f.current)}  Mean prior: {_fmt_money(f.mean_prior)}  "
                    f"Delta: {_fmt_money(f.delta)} ({_fmt_pct(f.delta_pct) if f.delta_pct is not None else 'N/A'})"
                )
    lines.append("")

    # WARN section
    warns = [f for f in result.findings if f.severity == "WARN"]
    lines.append(f"WARN ({len(warns)})")
    for f in warns:
        head = f"  {f.tier}  "
        if f.account:
            head += f"{f.account}"
            if f.department:
                head += f" @ {f.department}"
        lines.append(head)
        lines.append(f"      {f.message}")
        if f.current is not None and f.mean_prior is not None:
            lines.append(
                f"      Current: {_fmt_money(f.current)}  Mean prior: {_fmt_money(f.mean_prior)}  "
                f"Delta: {_fmt_money(f.delta)}"
            )
        if f.suggested_action:
            lines.append(f"      -> {f.suggested_action}")
    lines.append("")

    # INFO section
    infos = [f for f in result.findings if f.severity == "INFO"]
    if infos:
        lines.append(f"INFO ({len(infos)})")
        for f in infos:
            lines.append(f"  {f.tier}  {f.message}")
        lines.append("")

    pass_count = counts.get("PASS", 0)
    lines.append(f"PASS ({pass_count}) [collapsed - full detail in workbook tab]")
    lines.append("")

    if header_meta.get("workbook_tab_path"):
        lines.append(f"Workbook tab: {header_meta['workbook_tab_path']}")
        lines.append("")

    if result.has_fail:
        lines.append("==> DO NOT UPLOAD. Fix the CSV (or Reclasses tab) first; re-run /gut-check after edits.")
    else:
        lines.append("==> SAFE TO UPLOAD: WARNs are advisory. Review and proceed when ready.")
        lines.append("    NetSuite UI -> Lists -> Import Assistant -> Transactions -> Journal Entry")
        lines.append("    -> upload CSV -> confirm field mapping -> run import.")
    lines.append("=" * 73)
    return "\n".join(lines)


def write_workbook_tab(
    workbook_path: str,
    result: GutCheckResult,
    header_meta: dict,
    timestamp: Optional[_dt.datetime] = None,
) -> str:
    """Append a 'Gut-Check {YYYY-MM-DD HH:MM}' tab to the backup workbook with
    full findings detail. Returns the tab name written.

    Workbook must already exist (it's the payroll backup). If not found,
    raises FileNotFoundError so the skill can surface it.
    """
    if not os.path.isfile(workbook_path):
        raise FileNotFoundError(f"Backup workbook not found: {workbook_path}")
    ts = timestamp or _dt.datetime.now()
    # Sheet names must be <= 31 chars in Excel; "Gut-Check 2026-05-05 14:23" = 25 chars
    tab_name = f"Gut-Check {ts.strftime('%Y-%m-%d %H_%M')}"
    if len(tab_name) > 31:
        tab_name = tab_name[:31]

    wb = openpyxl.load_workbook(workbook_path)
    if tab_name in wb.sheetnames:
        # Disambiguate with seconds
        tab_name = f"Gut-Check {ts.strftime('%m-%d %H_%M_%S')}"
        tab_name = tab_name[:31]
    ws = wb.create_sheet(title=tab_name)

    # Header rows
    ws.append(["Gut-Check Review", header_meta["pay_run_label"]])
    ws.append(["Skill", header_meta["skill"]])
    ws.append(["CSV", header_meta.get("csv_basename", "")])
    ws.append(["Run at", ts.strftime("%Y-%m-%d %H:%M:%S")])
    if header_meta.get("priors"):
        ws.append(["Compared against", ", ".join(p["tranid"] for p in header_meta["priors"])])
    if header_meta.get("baseline_warnings"):
        for w in header_meta["baseline_warnings"]:
            ws.append(["Baseline note", w])
    ws.append([])

    # Findings table
    cols = ["Tier", "Severity", "Department", "Account", "Check",
            "Current", "Prior 1", "Prior 2", "Mean Prior",
            "Delta", "Delta %", "Threshold", "Message", "Suggested Action"]
    ws.append(cols)
    # Order: FAIL, WARN, INFO, PASS
    severity_order = {"FAIL": 0, "WARN": 1, "INFO": 2, "PASS": 3}
    sorted_findings = sorted(
        result.findings,
        key=lambda f: (severity_order.get(f.severity, 9), f.tier, f.account or "", f.department or "")
    )
    for f in sorted_findings:
        ws.append([
            f.tier, f.severity,
            f.department or "", f.account or "",
            f.check,
            f.current if f.current is not None else "",
            f.prior_1 if f.prior_1 is not None else "",
            f.prior_2 if f.prior_2 is not None else "",
            f.mean_prior if f.mean_prior is not None else "",
            f.delta if f.delta is not None else "",
            f.delta_pct * 100 if f.delta_pct is not None else "",
            f.threshold or "",
            f.message,
            f.suggested_action,
        ])
    wb.save(workbook_path)
    return tab_name


# ============================================================================
# High-level entry point (assumes caller pre-fetched NS data)
# ============================================================================

def run_gut_check(
    folder_path: str,
    audit_log_path: str,
    suiteql_runner=None,
) -> dict:
    """Top-level orchestration if a SuiteQL runner is supplied.

    `suiteql_runner` is a callable (sql: str) -> list[dict] of result rows.
    If None, this function returns the SQL strings and the caller must execute
    them and re-call the lower-level functions directly.
    """
    folder_path = os.path.abspath(folder_path)
    skill = detect_skill_from_folder(folder_path)
    config = SKILL_CONFIGS[skill]
    folder_name = os.path.basename(folder_path)

    # Find the CSV
    import glob as _glob
    csv_pattern = config["csv_glob"].replace("{folder_name}", folder_name)
    matches = _glob.glob(os.path.join(folder_path, csv_pattern))
    matches = [m for m in matches if not os.path.basename(m).startswith("~")]
    if not matches:
        raise FileNotFoundError(f"No JE Import CSV found at {os.path.join(folder_path, csv_pattern)}")
    if len(matches) > 1:
        raise RuntimeError(f"Multiple CSVs match: {matches}; expected exactly one.")
    csv_path = matches[0]

    current_lines = parse_csv(csv_path)
    debit_total = round(sum(ln["debit"] for ln in current_lines), 2)
    credit_total = round(sum(ln["credit"] for ln in current_lines), 2)
    current_trandate = None
    if current_lines:
        # Convert M/D/YYYY to YYYY-MM-DD
        try:
            d = _dt.datetime.strptime(current_lines[0]["date"], "%m/%d/%Y").date()
            current_trandate = d.isoformat()
        except (ValueError, KeyError):
            pass

    if not current_trandate:
        return {
            "status": "error",
            "message": "Could not parse current trandate from CSV; aborting.",
            "csv_path": csv_path,
        }

    # Build priors-search SQL (SuiteQL primary; audit_log was unreliable when
    # JEs are rejected/reposted manually).
    priors_search_sql = build_priors_search_sql(
        memo_prefix=config["memo_prefix"],
        subsidiary=config["subsidiary"],
        before_date=current_trandate,
        n=2,
    )

    if suiteql_runner is None:
        # Caller must orchestrate
        return {
            "status": "needs_suiteql",
            "csv_path": csv_path,
            "current_lines": current_lines,
            "priors_search_sql": priors_search_sql,
            "config": config,
            "current_trandate": current_trandate,
            "skill": skill,
            "folder_name": folder_name,
        }

    # Inline SuiteQL flow
    prior_rows = suiteql_runner(priors_search_sql)
    if not prior_rows:
        return {
            "status": "no_priors",
            "message": (
                f"No prior {skill} JEs found in NetSuite "
                f"(memo LIKE '%{config['memo_prefix']}%', "
                f"sub='{config['subsidiary']}', date < {current_trandate})."
            ),
            "csv_path": csv_path,
        }
    priors = []
    baseline_warnings = []
    for r in prior_rows:
        priors.append({
            "tranid": r.get("tranid"),
            "internal_id": r.get("internal_id"),
            "ns_trandate": r.get("trandate"),
            "ns_memo": r.get("memo"),
        })
    if len(priors) < 2:
        baseline_warnings.append(
            f"Only {len(priors)} valid prior(s) available "
            f"(needed 2); comparisons will be best-effort."
        )
    # Stale-baseline check
    if priors:
        try:
            cur_dt = _dt.datetime.strptime(current_trandate, "%Y-%m-%d").date()
            top_dt = _dt.datetime.strptime(priors[0]["ns_trandate"], "%Y-%m-%d").date()
            gap = (cur_dt - top_dt).days
            if gap > 1.5 * config["cadence_days"]:
                baseline_warnings.append(
                    f"Stale baseline: most recent prior is {gap} days old "
                    f"(expected ~{config['cadence_days']})."
                )
        except (ValueError, TypeError):
            pass
    prior_lines_by_je = {}
    for p in priors:
        rows = suiteql_runner(build_line_fetch_sql(p["tranid"], config["subsidiary"]))
        prior_lines_by_je[p["tranid"]] = normalize_ns_lines(rows)
    # Repackage warnings for the report
    sieved = {"warnings": baseline_warnings}

    result = analyze(current_lines, prior_lines_by_je, config)

    header_meta = {
        "skill": skill,
        "pay_run_label": folder_name,
        "csv_basename": os.path.basename(csv_path),
        "priors": priors,
        "baseline_warnings": sieved["warnings"],
        "totals": {
            "line_count": len(current_lines),
            "debit": debit_total,
            "credit": credit_total,
            "currency": config["currency"],
        },
    }

    workbook_pattern = config["workbook_glob"].replace("{folder_name}", folder_name)
    wb_matches = _glob.glob(os.path.join(folder_path, workbook_pattern))
    wb_matches = [m for m in wb_matches if not os.path.basename(m).startswith("~")]
    tab_path = ""
    if wb_matches:
        wb_path = wb_matches[0]
        tab_name = write_workbook_tab(wb_path, result, header_meta)
        tab_path = f"{os.path.basename(wb_path)} -> '{tab_name}'"
        header_meta["workbook_tab_path"] = tab_path

    chat_report = format_chat_report(result, header_meta)

    return {
        "status": "ok",
        "csv_path": csv_path,
        "skill": skill,
        "priors_used": priors,
        "baseline_warnings": sieved["warnings"],
        "result": result,
        "findings_count": result.counts,
        "has_fail": result.has_fail,
        "chat_report": chat_report,
        "workbook_tab": tab_path,
    }


if __name__ == "__main__":
    import sys
    if len(sys.argv) < 2:
        print("Usage: python payroll_gut_check.py <pay-run-folder>")
        sys.exit(2)
    folder = sys.argv[1]
    audit_log_path = os.path.abspath(os.path.join(_THIS_DIR, "..", "..", "audit_log.json"))
    out = run_gut_check(folder, audit_log_path, suiteql_runner=None)
    if out["status"] == "needs_suiteql":
        print("CLI mode does not run NetSuite calls. Use the gut-check skill via Claude.")
        print(f"Detected skill: {out['skill']}")
        print(f"CSV path: {out['csv_path']}")
        print(f"Found {len(out['candidates'])} prior candidates from audit_log:")
        for c in out["candidates"]:
            print(f"  {c['tranid']}  {c['timestamp'][:10]}  {c.get('memo', '')[:60]}")
        print()
        print("Validation SQL to run:")
        print(out["validate_sql"])
    elif out["status"] == "no_priors":
        print(out["message"])
    else:
        print(out["chat_report"])
