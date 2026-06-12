"""
US Payroll JE builder.

Reads the raw ADP payroll export from a pay-run folder, applies the US GL /
department / split routing, and writes `{MM.DD.YYYY} US Payroll Backup.xlsx`
into that same folder. The backup contains every raw ADP sheet copied verbatim
plus a new `{MM.DD.YYYY}_US_Payroll_JE (pivot)` tab that mirrors the format
the accountant already reviews.

Aggregation is at (Department, Account, Memo, Subsidiary) - the exact level of
the legacy Review_Report_Payroll.csv. DO NOT aggregate further.

Usage:
  python payroll_mapper.py "<pay-run folder>"
      e.g. python payroll_mapper.py "../../Monthly Payroll/Pay Runs/US/04.15.2026/"

  --verify-suffix    Write `... US Payroll Backup (verify).xlsx` instead of
                     overwriting any existing backup (for dry-run verification).

Also exposes `generate_je(folder_path, verify=False)` for import from the skill.
"""

import glob
import os
import re
import sys
from datetime import datetime

import openpyxl
import pandas as pd

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
DEPT_MAP_PATH = os.path.join(SCRIPT_DIR, "US Payroll Department Mapping File.csv")
GL_MAP_PATH = os.path.join(SCRIPT_DIR, "Payroll_Mapping_with_GL_Accounts (final).csv")

SUBSIDIARY = "Acme Holdings : Acme, Inc."
CODES_TO_IGNORE = {"DPIID", "DPIIM", "DPIIV", "GTL", "GIFT"}


def clean_amount(val):
    if pd.isna(val) or str(val).strip() == "":
        return 0.0
    try:
        if isinstance(val, str):
            val = val.replace(",", "").strip()
            if val.startswith("(") and val.endswith(")"):
                val = "-" + val[1:-1]
        return float(val)
    except ValueError:
        return 0.0


def parse_folder_date(folder_path: str) -> datetime:
    """Parse MM.DD.YYYY from folder name -> datetime."""
    name = os.path.basename(os.path.normpath(folder_path))
    m = re.match(r"^(\d{2})\.(\d{2})\.(\d{4})$", name)
    if not m:
        raise ValueError(
            f"Folder name must be 'MM.DD.YYYY' format (got {name!r})"
        )
    return datetime(int(m.group(3)), int(m.group(1)), int(m.group(2)))


def find_raw_file(folder_path: str) -> str:
    """Locate the input ADP file in the pay-run folder.

    Priority:
      1. {MM.DD.YYYY} US Payroll Backup.xls (controller-provided pre-built
         backup with raw sheets + Reclasses tab already in it)
      2. {MM.DD.YYYY} US Payroll Backup.xlsx (re-run scenario where our own
         output is the only file present)
      3. Any other raw ADP-named file (legacy fallback for the G* drop)

    Files with `(verify)` in the name are always excluded.
    """
    if not os.path.isdir(folder_path):
        raise FileNotFoundError(f"Not a folder: {folder_path!r}")

    backup_candidates = []
    other_candidates = []
    for ext in ("*.xlsx", "*.xls"):
        for f in glob.glob(os.path.join(folder_path, ext)):
            base = os.path.basename(f)
            if base.startswith("~"):
                continue
            if "Review_Report" in base:
                continue
            if "Desired Output" in base:
                continue
            if "(verify)" in base:
                continue
            if "Backup" in base:
                backup_candidates.append(f)
            else:
                other_candidates.append(f)

    if backup_candidates:
        # Prefer .xls (controller's drop) over .xlsx (our prior output) when both exist
        xls_first = sorted(backup_candidates, key=lambda p: 0 if p.lower().endswith(".xls") else 1)
        if len(xls_first) > 1 and not xls_first[0].lower().endswith(".xls"):
            raise RuntimeError(
                f"Multiple Backup files in {folder_path}: {xls_first}. "
                "Leave only one Backup file in the folder."
            )
        return xls_first[0]

    if not other_candidates:
        raise FileNotFoundError(f"No input file found in {folder_path}")
    if len(other_candidates) > 1:
        raise RuntimeError(
            f"Multiple candidate input files in {folder_path}: {other_candidates}. "
            "Leave only one ADP file in the folder."
        )
    return other_candidates[0]


def _load_mappings():
    dept_map = pd.read_csv(DEPT_MAP_PATH)
    gl_map_raw = pd.read_csv(GL_MAP_PATH)

    dept_map["Cost Center"] = (
        dept_map["Payroll Report Cost Center"]
        .astype(str)
        .str.replace("Cost Center: ", "", regex=False)
        .str.strip()
    )
    gl_map_raw["Payroll Code"] = gl_map_raw["Payroll Code"].astype(str).str.strip()

    cc_to_dept = dict(zip(dept_map["Cost Center"], dept_map["Department"]))
    code_to_gl = (
        gl_map_raw.drop_duplicates(subset=["Payroll Code"], keep="first")
        .set_index("Payroll Code")
        .to_dict(orient="index")
    )

    hsaer_expense = gl_map_raw[
        (gl_map_raw["Payroll Code"] == "HSAER") & (gl_map_raw["Type"] == "Expense")
    ]
    hsaer_expense_dict = hsaer_expense.iloc[0].to_dict() if not hsaer_expense.empty else {}

    hsaer_liability = gl_map_raw[
        (gl_map_raw["Payroll Code"] == "HSAER") & (gl_map_raw["Type"] == "Liability")
    ]
    hsaer_liability_dict = (
        hsaer_liability.iloc[0].to_dict() if not hsaer_liability.empty else {}
    )

    return cc_to_dept, code_to_gl, hsaer_expense_dict, hsaer_liability_dict


def _extract_records(raw_file: str):
    """Extract per-code records plus ADP control totals needed for the JE.

    Returns a dict:
        records:        list of {Cost Center, Payroll Code, Amount, Category}
                        for Earnings, Employee Deds, Employer Ded Exp (401KM only),
                        and Employer Tax Exp. Employee Taxes are NOT included here —
                        231350 is computed from the ADP footer total instead.
        total_net:      ADP "Total Net:" cash amount.
        adp_ee_tax_total: ADP footer Employee Taxes category subtotal.
        adp_er_tax_total: ADP footer Employer Tax Exp category subtotal.
    """
    xl = pd.ExcelFile(raw_file)
    sheets = [s for s in xl.sheet_names if s.startswith("Sheet")]
    all_records = []
    # NYPFLEE / NYSDIEE amounts per cost-center. They're Employee-Tax codes in
    # the Ee Tax section, but ADP rolls them into the per-dept Employer Tax
    # Exp totals, so we need to back them out per dept before routing to
    # 611450 / 511350. Also backed out from the aggregate 231350 credit.
    ny_adj_per_cc = {}  # cc -> {"NYPFLEE": amt, "NYSDIEE": amt}
    last_sheet_header_cols = {}

    for sheet in sheets:
        df = pd.read_excel(xl, sheet_name=sheet, header=None)

        header_row_idx = -1
        for idx, row in df.head(50).iterrows():
            row_str = " ".join(str(x) for x in row.values if pd.notna(x))
            if "Employee Deds" in row_str and "Code" not in row_str:
                header_row_idx = idx
                break
        if header_row_idx == -1:
            continue

        categories = (
            df.iloc[header_row_idx].ffill().fillna("").astype(str).str.strip().tolist()
        )
        col_names = (
            df.iloc[header_row_idx + 1].fillna("").astype(str).str.strip().tolist()
        )

        mapping_cols = {}
        for i, cat in enumerate(categories):
            if cat == "Earnings" and col_names[i] == "Code":
                mapping_cols["Earnings_Code"] = i
            elif cat == "Employee Deds" and col_names[i] == "Code":
                mapping_cols["Employee_Deds_Code"] = i
            elif cat == "Employee Taxes" and col_names[i] == "Code":
                mapping_cols["Employee_Taxes_Code"] = i
            elif cat == "Employer Ded Exp" and col_names[i] == "Code":
                mapping_cols["Employer_Ded_Exp_Code"] = i
            elif cat == "Employer Tax Exp" and col_names[i] == "Code":
                mapping_cols["Employer_Tax_Exp_Code"] = i

        if sheet == sheets[-1]:
            last_sheet_header_cols = dict(mapping_cols)

        current_cc = None
        for idx in range(header_row_idx + 2, len(df)):
            row = df.iloc[idx]

            val0 = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else ""
            if val0.startswith("Cost Center:"):
                current_cc = val0.replace("Cost Center: ", "").strip()
                continue

            if not current_cc:
                continue

            # Capture NYPFLEE / NYSDIEE from the Employee Taxes section for the
            # per-dept Er Tax backout. Employee Taxes aren't otherwise extracted
            # because 231350 comes from the ADP footer total.
            ee_tax_code_idx = mapping_cols.get("Employee_Taxes_Code")
            if ee_tax_code_idx is not None and ee_tax_code_idx < len(row):
                ee_tax_code = row.iloc[ee_tax_code_idx]
                if pd.notna(ee_tax_code):
                    code_str = str(ee_tax_code).strip().upper()
                    if code_str in ("NYPFLEE", "NYSDIEE"):
                        for look in range(5, 0, -1):
                            if ee_tax_code_idx + look >= len(row):
                                continue
                            v = row.iloc[ee_tax_code_idx + look]
                            if pd.notna(v):
                                amt = clean_amount(v)
                                if amt != 0:
                                    bucket = ny_adj_per_cc.setdefault(current_cc, {"NYPFLEE": 0.0, "NYSDIEE": 0.0})
                                    bucket[code_str] += amt
                                    break

            categories_to_extract = [
                ("Earnings", mapping_cols.get("Earnings_Code"), 8),
                ("Employee Deds", mapping_cols.get("Employee_Deds_Code"), 5),
                ("Employer Ded Exp", mapping_cols.get("Employer_Ded_Exp_Code"), 5),
                ("Employer Tax Exp", mapping_cols.get("Employer_Tax_Exp_Code"), 8),
            ]
            # Employee Taxes codes are intentionally NOT extracted — 231350 comes
            # from the ADP footer total directly, so individual employee-tax
            # codes are irrelevant to routing.

            for cat_name, code_idx, search_limit in categories_to_extract:
                if code_idx is None:
                    continue
                code_val = (
                    str(row.iloc[code_idx]).strip() if pd.notna(row.iloc[code_idx]) else ""
                )
                if not code_val or code_val.upper() == "NAN":
                    continue
                if str(code_val).replace(".", "").isdigit():
                    continue
                if code_val.upper() in CODES_TO_IGNORE:
                    continue

                amt_val = 0.0
                for look_ahead in range(search_limit, 0, -1):
                    if (code_idx + look_ahead) >= len(row):
                        continue
                    potential_amt = row.iloc[code_idx + look_ahead]
                    if pd.isna(potential_amt):
                        continue
                    s = str(potential_amt).strip()
                    if s == "" or s == "Z":
                        continue
                    cleaned_amt = clean_amount(potential_amt)
                    if cleaned_amt == 0:
                        continue
                    if cat_name == "Earnings" and search_limit >= 8 and look_ahead < 6:
                        continue
                    amt_val = cleaned_amt
                    break

                if amt_val == 0:
                    continue
                if cat_name == "Employer Ded Exp" and code_val != "401KM":
                    continue

                all_records.append(
                    {
                        "Cost Center": current_cc,
                        "Payroll Code": code_val,
                        "Amount": amt_val,
                        "Category": cat_name,
                    }
                )

    # Collapse duplicate (CC, Code, Category) rows (multi-shift splits)
    unique_records = {}
    for r in all_records:
        key = (r["Cost Center"], r["Payroll Code"], r["Category"])
        if key not in unique_records:
            unique_records[key] = dict(r)
        else:
            unique_records[key]["Amount"] += r["Amount"]
    all_records = list(unique_records.values())

    # ADP footer totals from the last sheet's "Report Totals:" row.
    total_net = 0.0
    adp_ee_tax_total = 0.0
    adp_er_tax_total = 0.0
    if sheets:
        df_last = pd.read_excel(xl, sheet_name=sheets[-1], header=None)
        ee_tax_col = last_sheet_header_cols.get("Employee_Taxes_Code")
        er_tax_col = last_sheet_header_cols.get("Employer_Tax_Exp_Code")

        for idx in range(max(0, len(df_last) - 30), len(df_last)):
            row = df_last.iloc[idx]
            row_str = " ".join(str(x) for x in row.values if pd.notna(x)).strip()

            if "Report Totals:" in row_str and idx + 1 < len(df_last):
                totals_row = df_last.iloc[idx + 1]
                # On the last sheet the category subtotals sit at the same col
                # as each category's Code header (observed: Ee Tax total at the
                # Employee Taxes code col, Er Tax total at the Employer Tax Exp
                # code col).
                if ee_tax_col is not None and ee_tax_col < len(totals_row):
                    adp_ee_tax_total = clean_amount(totals_row.iloc[ee_tax_col])
                if er_tax_col is not None and er_tax_col < len(totals_row):
                    adp_er_tax_total = clean_amount(totals_row.iloc[er_tax_col])

            if "Total Net:" in row_str:
                for val in row.values:
                    if pd.notna(val) and clean_amount(val) > 1000:
                        total_net = clean_amount(val)

    return {
        "records": all_records,
        "total_net": total_net,
        "adp_ee_tax_total": adp_ee_tax_total,
        "adp_er_tax_total": adp_er_tax_total,
        "ny_adj_per_cc": ny_adj_per_cc,
    }


def _build_je_rows(extraction, cc_to_dept, code_to_gl,
                   hsaer_expense_dict, hsaer_liability_dict):
    """Apply routing rules; return (final_rows, unmapped_ccs, unmapped_codes)."""
    all_records = extraction["records"]
    total_net = extraction["total_net"]
    adp_ee_tax_total = extraction["adp_ee_tax_total"]
    ny_adj_per_cc = extraction.get("ny_adj_per_cc", {})

    # NY adjustment per CC (NYPFLEE + NYSDIEE) — subtracted from per-CC Er Tax
    # Exp expense. Also summed for aggregate 231350 backout.
    ny_total_per_cc = {cc: round(sum(d.values()), 2) for cc, d in ny_adj_per_cc.items()}
    ny_aggregate_total = round(sum(ny_total_per_cc.values()), 2)

    # For Er Tax rows, we aggregate per-CC so the NY backout applies cleanly.
    # Build a lookup: cc -> remaining NY budget to subtract across that CC's
    # Er Tax rows. We spend this budget on the largest-amount rows first so
    # it's guaranteed to close out without going negative.
    er_tax_by_cc = {}
    for rec in all_records:
        if rec["Category"] == "Employer Tax Exp":
            er_tax_by_cc.setdefault(rec["Cost Center"], []).append(rec)
    # Reduce the single largest Er Tax row per CC by that CC's NY total. Chosen
    # because the NY amount is tiny relative to USSOCER/USMEDER, so the big
    # row absorbs it without going negative, and per-CC totals still equal
    # ADP-minus-NY.
    for cc, recs in er_tax_by_cc.items():
        remaining = ny_total_per_cc.get(cc, 0.0)
        if remaining <= 0:
            continue
        recs.sort(key=lambda r: -r["Amount"])
        for rec in recs:
            if remaining <= 0:
                break
            take = min(remaining, rec["Amount"])
            rec["Amount"] = round(rec["Amount"] - take, 2)
            remaining = round(remaining - take, 2)

    unmapped_ccs = set()
    unmapped_codes = set()
    final_rows = []

    for rec in all_records:
        cc = rec["Cost Center"]
        code = rec["Payroll Code"]
        amt = rec["Amount"]
        category = rec["Category"]

        dept = cc_to_dept.get(cc, None)
        if not dept:
            unmapped_ccs.add(cc)
            dept = "UNMAPPED_DEPT"

        # Employer Tax Exp: route ALL codes uniformly to the Payroll Taxes
        # account by department. We do not use per-code GL mapping here — every
        # employer tax code (USFUTA, SUIs, state/local ER codes, CODENER, etc.)
        # is a payroll-tax expense. This avoids UNMAPPED_ACCOUNT rows for
        # rarely-seen codes and keeps the sum equal to the ADP per-dept subtotal.
        if category == "Employer Tax Exp":
            opex_acct = "611450 Salary and Compensation : Payroll Taxes"
            cogs_acct = "511350 COGS - Salary and Compensation : COGS - Payroll Taxes"
            memo = "US PAYROLL - PR Tax"
            code_type = "Expense"
            gl_name = "Payroll Taxes"
            gl_info = {}
        elif code == "HSAER":
            gl_info = hsaer_expense_dict if category == "Earnings" else hsaer_liability_dict
        else:
            gl_info = code_to_gl.get(code, {})

        if category != "Employer Tax Exp":
            if not gl_info:
                unmapped_codes.add(code)

            opex_acct = str(gl_info.get("GL Account - Opex", ""))
            cogs_acct = str(gl_info.get("GL Account - COGS", ""))
            memo = str(gl_info.get("Line Memo", ""))
            code_type = str(gl_info.get("Type", ""))
            gl_name = str(gl_info.get("GL Account Name", ""))

            if opex_acct == "nan": opex_acct = ""
            if cogs_acct == "nan": cogs_acct = ""
            if memo == "nan": memo = ""
            if gl_name == "nan": gl_name = ""

        if str(dept).startswith("COGS") and cogs_acct:
            acct = cogs_acct
        elif opex_acct:
            acct = opex_acct
        elif cogs_acct:
            acct = cogs_acct
        else:
            acct = "UNMAPPED_ACCOUNT"

        # Capture whether this row's HOME department is Infrastructure before
        # any dept override (severance / liability moves to EBITDA / Aggregate).
        # Used below so Infrastructure-origin severance still gets 30/70 split.
        home_is_infrastructure = "Infrastructure" in str(dept)

        if code_type.lower() in ["liability", "payable"]:
            dept = "Aggregate"
            if amt > 0:
                amt = -abs(amt)
            if "hsa payable" in gl_name.lower() and code == "HSAER":
                amt = -abs(amt)

        # SEVR-or-Severance routing: posting dept overrides to EBITDA Adjustments;
        # account (611250 OpEx / 511175 COGS) was selected above based on home
        # dept's COGS-vs-OpEx nature. For Infrastructure (hybrid) we ALSO split
        # the principal 30/70 — see the Infrastructure split block below.
        is_severance = code == "SEVR" or gl_name.lower() == "severance"
        if is_severance:
            home_dept_tail = str(dept).split(" : ")[-1] if dept else "Unknown"
            memo = f"US PAYROLL - Severance ({home_dept_tail})"
            dept = "EBITDA Adjustments"

        # Health / Other Benefits sign inversion. ADP reports Employee-Deds
        # withholdings as positive amounts; on the JE they're credits against
        # the expense account, so we negate. If the ADP amount is already
        # negative (e.g., MISCD reversal), negating flips it back to a debit —
        # the correct reversal behavior.
        #
        # Only applies to the "Employee Deds" category. Earnings-category
        # codes (HSAER, MEDCR) routed to Health / Other Benefits are employer
        # contributions and stay as positive debits.
        if ("health benefits" in gl_name.lower() or "other benefits" in gl_name.lower()) \
                and dept != "Aggregate" and category == "Employee Deds":
            amt = -amt

        # Infrastructure 70/30 split: 70% OpEx (611xxx) / 30% COGS (511xxx).
        # Applies to lines posting to the Infrastructure dept directly (regular
        # pay/benefits/tax), AND to severance lines whose home dept was
        # Infrastructure (the dept has since been overridden to EBITDA
        # Adjustments, but we still want the split because the underlying
        # employee's labor was hybrid-coded).
        should_split = (
            "Infrastructure" in str(dept)            # regular Infra lines
            or (is_severance and home_is_infrastructure)  # Infra-origin severance
        )
        if should_split and amt != 0:
            amt_opex = round(amt * 0.7, 2)
            amt_cogs = round(amt - amt_opex, 2)
            split_opex = opex_acct if opex_acct else acct
            split_cogs = cogs_acct if cogs_acct else acct
            # For Infra-origin severance, suffix the memo to mirror the
            # PR-Tax reclass pattern so the split rows aggregate separately.
            if is_severance and home_is_infrastructure:
                memo_opex = memo.replace("(Infrastructure)", "(Infrastructure - Opex)")
                memo_cogs = memo.replace("(Infrastructure)", "(Infrastructure - COGS)")
            else:
                memo_opex = memo
                memo_cogs = memo

            final_rows.append({
                "Department": dept, "Account": split_opex, "Memo": memo_opex,
                "Subsidiary": SUBSIDIARY, "Amount": amt_opex,
            })
            final_rows.append({
                "Department": dept, "Account": split_cogs, "Memo": memo_cogs,
                "Subsidiary": SUBSIDIARY, "Amount": amt_cogs,
            })
        else:
            final_rows.append({
                "Department": dept, "Account": acct, "Memo": memo,
                "Subsidiary": SUBSIDIARY, "Amount": amt,
            })

    if total_net != 0:
        # Total Net from the ADP report is printed as a positive cash amount,
        # but on the JE it's a credit to the Payroll Liability account. WBNK1
        # (employee deduction already routed to 231200 as a liability above)
        # lands on the same 231200 line via the groupby and further grows the
        # credit — net result: 231200 = -(Total Net + WBNK1 total).
        final_rows.append({
            "Department": "Aggregate",
            "Account": "231200 Payroll Liability",
            "Memo": "Payroll Liability",
            "Subsidiary": SUBSIDIARY,
            "Amount": -abs(total_net),
        })

    # Seed 231350 Payroll Tax Liability with the ADP Employee Taxes total
    # (credit). The per-CC Er Tax rows above have already had the per-CC NY
    # amount backed out, so the employer-tax consolidation below produces a
    # debit sum = ADP Er Tax total - NY aggregate. Net: 231350 credit =
    # -(ADP Ee Tax + ADP Er Tax - NY aggregate), matching the accountant's spec.
    if adp_ee_tax_total > 0:
        final_rows.append({
            "Department": "Aggregate",
            "Account": "231350 Payroll Tax Liability",
            "Memo": "Payroll Tax Liability",
            "Subsidiary": SUBSIDIARY,
            "Amount": -adp_ee_tax_total,
        })

    final_df = pd.DataFrame(final_rows)
    final_grouped = final_df.groupby(
        ["Department", "Account", "Memo", "Subsidiary"], as_index=False
    )["Amount"].sum()
    final_grouped["Amount"] = final_grouped["Amount"].round(2)
    final_grouped = final_grouped[final_grouped["Amount"] != 0.0]

    # Consolidate employer-side payroll tax expense into 231350 Payroll Tax
    # Liability so the JE balances. Employer tax is a positive debit on
    # 511350/611450; the offsetting credit on 231350 must grow (more negative).
    employer_tax_total = final_grouped[
        final_grouped["Account"].str.contains("Payroll Taxes", na=False)
        & (final_grouped["Department"] != "Aggregate")
    ]["Amount"].sum()

    mask_tax_liab = (
        (final_grouped["Department"] == "Aggregate")
        & (final_grouped["Account"] == "231350 Payroll Tax Liability")
    )
    if mask_tax_liab.sum() > 0:
        final_grouped.loc[mask_tax_liab, "Amount"] -= employer_tax_total
    else:
        new_row = pd.DataFrame({
            "Department": ["Aggregate"],
            "Account": ["231350 Payroll Tax Liability"],
            "Memo": ["Payroll Tax Liability"],
            "Subsidiary": [SUBSIDIARY],
            "Amount": [-employer_tax_total],
        })
        final_grouped = pd.concat([final_grouped, new_row], ignore_index=True)

    # Same consolidation for employer 401k Match (code 401KM): the expense
    # debit on 511250/611350 must have an offsetting credit on 231250 401K
    # Payable.
    employer_401k_total = final_grouped[
        final_grouped["Account"].str.contains("401k Match", na=False, case=False)
        & (final_grouped["Department"] != "Aggregate")
    ]["Amount"].sum()

    mask_401k_liab = (
        (final_grouped["Department"] == "Aggregate")
        & (final_grouped["Account"] == "231250 401K payable")
    )
    if mask_401k_liab.sum() > 0:
        final_grouped.loc[mask_401k_liab, "Amount"] -= employer_401k_total
    elif employer_401k_total != 0:
        new_row = pd.DataFrame({
            "Department": ["Aggregate"],
            "Account": ["231250 401K payable"],
            "Memo": ["401K Payable"],
            "Subsidiary": [SUBSIDIARY],
            "Amount": [-employer_401k_total],
        })
        final_grouped = pd.concat([final_grouped, new_row], ignore_index=True)

    final_grouped = final_grouped.sort_values(by=["Department", "Account"]).reset_index(drop=True)
    return final_grouped, unmapped_ccs, unmapped_codes


def _to_je_pivot_rows(final_grouped: pd.DataFrame, je_date: datetime,
                       je_memo: str) -> list:
    """Convert aggregated rows to the pivot tab format the accountant reviews.

    Columns: Date | Subsidiary | Department | Journal Entry Memo | Account |
             Line Memo | Debit (single signed column; negatives = credits).
    Department is left blank for Aggregate (liability) rows.
    """
    out = []
    for _, r in final_grouped.iterrows():
        dept = r["Department"]
        display_dept = "" if dept == "Aggregate" else dept
        out.append({
            "Date": je_date,
            "Subsidiary": r["Subsidiary"],
            "Department": display_dept,
            "Journal Entry Memo": je_memo,
            "Account": r["Account"],
            "Line Memo": r["Memo"],
            "Debit": float(r["Amount"]),
        })
    return out


def _write_backup_xlsx(raw_file: str, pivot_rows: list, pivot_sheet_name: str,
                        output_path: str):
    """Copy raw ADP sheets verbatim and append the JE pivot tab."""
    xl = pd.ExcelFile(raw_file)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for sheet_name in xl.sheet_names:
        df = pd.read_excel(xl, sheet_name=sheet_name, header=None)
        ws = wb.create_sheet(title=sheet_name)
        for r_idx in range(len(df)):
            for c_idx in range(df.shape[1]):
                val = df.iat[r_idx, c_idx]
                if pd.isna(val):
                    continue
                if hasattr(val, "item"):
                    val = val.item()
                ws.cell(row=r_idx + 1, column=c_idx + 1, value=val)

    ws = wb.create_sheet(title=pivot_sheet_name)
    headers = ["Date", "Subsidiary", "Department", "Journal Entry Memo",
               "Account", "Line Memo", "Debit"]
    ws.append(headers)
    for row in pivot_rows:
        ws.append([row[h] for h in headers])

    wb.save(output_path)


def generate_je(folder_path: str, verify: bool = False) -> dict:
    """Build the US payroll JE from the pay-run folder.

    Returns a dict: {
        'rows': list[dict]     # JE pivot rows (Date, Subsidiary, Department, ..., Debit)
        'totals': {'debits': float, 'credits': float, 'balanced': bool, 'delta': float}
        'output_path': str
        'raw_input': str
        'je_date': datetime
        'je_memo': str
        'pivot_sheet_name': str
        'unmapped_cost_centers': list[str]
        'unmapped_codes': list[str]
    }
    """
    folder_path = os.path.abspath(folder_path)
    je_date = parse_folder_date(folder_path)
    raw_file = find_raw_file(folder_path)

    folder_name = os.path.basename(folder_path)
    je_memo = f"{folder_name} US PAYROLL"
    pivot_sheet_name = f"{folder_name}_US_Payroll_JE (pivot)"

    cc_to_dept, code_to_gl, hsaer_exp, hsaer_liab = _load_mappings()
    extraction = _extract_records(raw_file)
    final_grouped, unmapped_ccs, unmapped_codes = _build_je_rows(
        extraction, cc_to_dept, code_to_gl, hsaer_exp, hsaer_liab
    )

    pivot_rows = _to_je_pivot_rows(final_grouped, je_date, je_memo)

    suffix = " (verify)" if verify else ""
    output_name = f"{folder_name} US Payroll Backup{suffix}.xlsx"
    output_path = os.path.join(folder_path, output_name)

    _write_backup_xlsx(raw_file, pivot_rows, pivot_sheet_name, output_path)

    debits = sum(r["Debit"] for r in pivot_rows if r["Debit"] > 0)
    credits = -sum(r["Debit"] for r in pivot_rows if r["Debit"] < 0)
    delta = round(debits - credits, 2)

    return {
        "rows": pivot_rows,
        "totals": {
            "debits": round(debits, 2),
            "credits": round(credits, 2),
            "balanced": abs(delta) < 0.01,
            "delta": delta,
        },
        "output_path": output_path,
        "raw_input": raw_file,
        "je_date": je_date,
        "je_memo": je_memo,
        "pivot_sheet_name": pivot_sheet_name,
        "unmapped_cost_centers": sorted(unmapped_ccs),
        "unmapped_codes": sorted(unmapped_codes),
    }


# ============================================================================
# Step 2 — Apply manual reclasses from the Reclasses tab
# ============================================================================

PAYROLL_TAX_OPEX_ACCT = "611450 Salary and Compensation : Payroll Taxes"
PAYROLL_TAX_COGS_ACCT = "511350 COGS - Salary and Compensation : COGS - Payroll Taxes"
SEVERANCE_OPEX_ACCT = "611250 Salary and Compensation : Severance"
SEVERANCE_COGS_ACCT = "511175 COGS - Salary and Compensation : COGS - Severance"
EBITDA_DEPT = "EBITDA Adjustments"
GA_DEPT = "General & Administrative : GA"
JB_INITIALS = "JB"

# Board member component labels → (account_opex, memo)
# JB lives under GA (non-COGS dept), so we only ever need the OpEx account.
JB_COMPONENTS = {
    "salary": ("611100 Salary and Compensation : Salaries and Wages", "US PAYROLL - Salary"),
    "medical": ("611300 Salary and Compensation : Health Benefits", "US PAYROLL - EE Med deductions"),
    "401k": ("611350 Salary and Compensation : 401k Match", "US PAYROLL - 401K"),
    "other": ("611400 Salary and Compensation : Other Benefits", "US PAYROLL - Misc EE withholding for other benefits"),
    "tax": ("611450 Salary and Compensation : Payroll Taxes", "US PAYROLL - PR Tax"),
}


def _normalize_dept_token(s: str) -> str:
    """Normalize a dept short name for matching: lowercase, strip, 'mgmt' <-> 'management'."""
    s = (s or "").strip().lower()
    s = s.replace("management", "mgmt")
    return " ".join(s.split())


def _resolve_full_dept(short_name: str, pivot_depts: list) -> str:
    """Map a Reclasses-tab short dept name to a full pivot dept name.

    Matches on the last ` : ` segment of each full dept (with mgmt/management
    normalization). Returns None if no match or ambiguous.
    """
    target = _normalize_dept_token(short_name)
    matches = []
    for full in pivot_depts:
        last = full.split(" : ")[-1] if full else ""
        if _normalize_dept_token(last) == target:
            matches.append(full)
    if len(matches) == 1:
        return matches[0]
    # Substring fallback — short name contained in any dept segment
    if not matches:
        for full in pivot_depts:
            if target in _normalize_dept_token(full):
                matches.append(full)
        if len(matches) == 1:
            return matches[0]
    return None


def _is_cogs_dept(dept: str) -> bool:
    return str(dept).startswith("COGS")


def _is_infrastructure_dept(dept: str) -> bool:
    return "Infrastructure" in str(dept)


def _short_dept_name(full_dept: str) -> str:
    """Last segment of a full dept path, used in severance memos."""
    return full_dept.split(" : ")[-1] if full_dept else full_dept


def _parse_reclasses_tab(backup_path: str) -> dict:
    """Read the Reclasses tab and return parsed sections.

    Returns:
        {
            "severance_rows": list of {dept_short, name, severance, tax},
            "borland_components": dict {canonical_key: amount} or None,
            "unknown_sections": list of section header strings,
            "raw_dump": list of (row_idx, [cells]) for echoing to the user,
        }
    """
    wb = openpyxl.load_workbook(backup_path, data_only=True)
    if "Reclasses" not in wb.sheetnames:
        wb.close()
        return {"severance_rows": [], "borland_components": None,
                "unknown_sections": [], "raw_dump": [], "present": False}

    ws = wb["Reclasses"]
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    wb.close()

    raw_dump = []
    for i, r in enumerate(rows, start=1):
        cells = [(chr(64 + j), v) for j, v in enumerate(r, start=1)
                 if v is not None and str(v).strip() != ""]
        if cells:
            raw_dump.append((i, cells))

    # Pass 1: locate section headers by scanning col A for section-title strings
    # Section types:
    #   'severance' -> Type A (severance tax reclass)
    #   'borland' + 'ebitda' -> Type B (JB board member)
    #   anything else non-empty -> unknown section
    sections = []  # list of (start_row, type, title_text)
    for i, r in enumerate(rows):
        a = str(r[0]).strip() if r[0] is not None else ""
        if not a:
            continue
        a_low = a.lower()
        b_low = str(r[1]).strip().lower() if len(r) > 1 and r[1] is not None else ""
        c_low = str(r[2]).strip().lower() if len(r) > 2 and r[2] is not None else ""
        # Detect section titles: col B/C/D should be empty-ish so this is a header
        has_bcd = any(r[j] is not None and str(r[j]).strip() != "" for j in (1, 2, 3))
        looks_like_header = not has_bcd
        # Recognize the severance column-header row as a section start, so the
        # tab doesn't need a separate "Severance" title row above it.
        is_severance_header = (
            a_low == "department" and b_low == "name" and c_low.startswith("sever")
        )
        if a_low == "severance" or is_severance_header:
            sections.append((i, "severance", a))
        elif "borland" in a_low and "ebitda" in a_low:
            sections.append((i, "borland", a))
        elif looks_like_header and a_low not in ("department",):
            sections.append((i, "unknown", a))

    # Determine row ranges for each section (from start to next section - 1)
    section_ranges = []
    for idx, (start, stype, title) in enumerate(sections):
        end = sections[idx + 1][0] - 1 if idx + 1 < len(sections) else len(rows) - 1
        section_ranges.append((start, end, stype, title))

    severance_rows = []
    borland_components = None
    unknown_sections = []

    for start, end, stype, title in section_ranges:
        if stype == "severance":
            # Walk rows from start+1 to end. Skip header row and subtotal rows.
            for i in range(start + 1, end + 1):
                r = rows[i]
                dept = str(r[0]).strip() if r[0] is not None else ""
                name = str(r[1]).strip() if r[1] is not None else ""
                sev_c = r[2]
                tax_d = r[3]
                # Need all four A-D populated to be a data row
                if not dept or not name or sev_c is None or tax_d is None:
                    continue
                # Skip the column-header row (dept="Department", name="Name")
                if dept.lower() == "department":
                    continue
                try:
                    sev_amt = float(sev_c)
                    tax_amt = float(tax_d)
                except (TypeError, ValueError):
                    continue
                severance_rows.append({
                    "dept_short": dept,
                    "name": name,
                    "severance": sev_amt,
                    "tax": tax_amt,
                    "row_idx": i + 1,
                })
        elif stype == "borland":
            comps = {}
            for i in range(start + 1, end + 1):
                r = rows[i]
                label = str(r[1]).strip() if r[1] is not None else ""
                amount = r[2]
                if not label or amount is None:
                    continue
                try:
                    amt = float(amount)
                except (TypeError, ValueError):
                    continue
                label_low = label.lower()
                if "salary" in label_low:
                    comps["salary"] = amt
                elif "medical" in label_low or "health" in label_low:
                    comps["medical"] = amt
                elif "401k" in label_low or "401(k)" in label_low:
                    comps["401k"] = amt
                elif "other" in label_low:
                    comps["other"] = amt
                elif "tax" in label_low:
                    comps["tax"] = amt
                # Silently ignore labels we can't match
            if comps:
                borland_components = comps
        elif stype == "unknown":
            unknown_sections.append(title)

    return {
        "severance_rows": severance_rows,
        "borland_components": borland_components,
        "unknown_sections": unknown_sections,
        "raw_dump": raw_dump,
        "present": True,
    }


def _build_severance_tax_reclass_entries(severance_rows: list, pivot_depts: list) -> tuple:
    """Return (entries, resolution_log, unresolved).

    entries is a list of {Department, Account, Memo, Subsidiary, Amount} dicts
    representing the reclass deltas to merge into the pivot: source rows are
    negative, destination rows are positive.

    resolution_log documents each severance employee -> resolved home dept -> routing.
    """
    # Aggregate tax per home dept before emitting (multi-employee in same dept)
    per_dept_tax = {}
    resolution_log = []
    unresolved = []

    for row in severance_rows:
        full = _resolve_full_dept(row["dept_short"], pivot_depts)
        if full is None:
            unresolved.append(row)
            continue
        per_dept_tax.setdefault(full, 0.0)
        per_dept_tax[full] += row["tax"]
        resolution_log.append({
            "name": row["name"],
            "dept_short": row["dept_short"],
            "resolved_dept": full,
            "severance": row["severance"],
            "tax": row["tax"],
        })

    entries = []
    for dept, total_tax in per_dept_tax.items():
        total_tax = round(total_tax, 2)
        if total_tax == 0:
            continue
        short = _short_dept_name(dept)

        if _is_infrastructure_dept(dept):
            # 30/70 split of the tax — mirrors Step 1 Infrastructure split.
            tax_cogs = round(total_tax * 0.3, 2)
            tax_opex = round(total_tax - tax_cogs, 2)
            # Source-side reduction: split source (611450 OpEx + 511350 COGS) using same ratio
            if tax_opex != 0:
                entries.append({
                    "Department": dept,
                    "Account": PAYROLL_TAX_OPEX_ACCT,
                    "Memo": "US PAYROLL - PR Tax",
                    "Subsidiary": SUBSIDIARY,
                    "Amount": -tax_opex,
                })
                entries.append({
                    "Department": EBITDA_DEPT,
                    "Account": SEVERANCE_OPEX_ACCT,
                    "Memo": f"US PAYROLL - PR Tax Severance ({short} - Opex)",
                    "Subsidiary": SUBSIDIARY,
                    "Amount": tax_opex,
                })
            if tax_cogs != 0:
                entries.append({
                    "Department": dept,
                    "Account": PAYROLL_TAX_COGS_ACCT,
                    "Memo": "US PAYROLL - PR Tax",
                    "Subsidiary": SUBSIDIARY,
                    "Amount": -tax_cogs,
                })
                entries.append({
                    "Department": EBITDA_DEPT,
                    "Account": SEVERANCE_COGS_ACCT,
                    "Memo": f"US PAYROLL - PR Tax Severance ({short} - COGS)",
                    "Subsidiary": SUBSIDIARY,
                    "Amount": tax_cogs,
                })
        else:
            cogs = _is_cogs_dept(dept)
            src_acct = PAYROLL_TAX_COGS_ACCT if cogs else PAYROLL_TAX_OPEX_ACCT
            dst_acct = SEVERANCE_COGS_ACCT if cogs else SEVERANCE_OPEX_ACCT
            entries.append({
                "Department": dept,
                "Account": src_acct,
                "Memo": "US PAYROLL - PR Tax",
                "Subsidiary": SUBSIDIARY,
                "Amount": -total_tax,
            })
            entries.append({
                "Department": EBITDA_DEPT,
                "Account": dst_acct,
                "Memo": f"US PAYROLL - PR Tax Severance ({short})",
                "Subsidiary": SUBSIDIARY,
                "Amount": total_tax,
            })
    return entries, resolution_log, unresolved


def _build_boardmember_reclass_entries(borland_components: dict, pivot_rows_df: pd.DataFrame) -> tuple:
    """Return (entries, warnings). JB always moves out of GA to EBITDA Adjustments."""
    entries = []
    warnings = []
    for key, amt in borland_components.items():
        if key not in JB_COMPONENTS:
            warnings.append(f"Unknown JB component '{key}' (amount ${amt:,.2f}) — skipped")
            continue
        acct, base_memo = JB_COMPONENTS[key]
        amt = round(amt, 2)
        if amt == 0:
            continue
        if amt < 0:
            warnings.append(f"JB {key} amount is negative (${amt:,.2f}) — unusual; flag for review")

        # Sanity check: only meaningful for positive-debit accounts (salary,
        # 401k, tax). For medical and other benefits the GA dept total is a
        # net credit from employee withholdings that can legitimately be
        # smaller in magnitude than the JB net amount — skip the check there.
        ga_match = pivot_rows_df[
            (pivot_rows_df["Department"] == GA_DEPT)
            & (pivot_rows_df["Account"] == acct)
        ]
        if len(ga_match) == 0:
            warnings.append(f"JB {key}: no matching row in GA dept for account {acct}")
        elif key in ("salary", "401k", "tax"):
            ga_balance = round(ga_match["Amount"].sum(), 2)
            if amt > ga_balance + 0.01:
                warnings.append(
                    f"JB {key} reclass ${amt:,.2f} exceeds GA balance ${ga_balance:,.2f}"
                )

        # Move `amt` from GA to EBITDA — consistent delta regardless of sign.
        # Source reduction is -amt, destination addition is +amt.
        entries.append({
            "Department": GA_DEPT,
            "Account": acct,
            "Memo": base_memo,
            "Subsidiary": SUBSIDIARY,
            "Amount": -amt,
        })
        entries.append({
            "Department": EBITDA_DEPT,
            "Account": acct,
            "Memo": f"{base_memo} ({JB_INITIALS})",
            "Subsidiary": SUBSIDIARY,
            "Amount": amt,
        })
    return entries, warnings


def _read_pivot_tab(backup_path: str, pivot_sheet_name: str) -> pd.DataFrame:
    """Read the pivot tab back as a DataFrame with our standard columns."""
    wb = openpyxl.load_workbook(backup_path, data_only=True)
    if pivot_sheet_name not in wb.sheetnames:
        # Auto-detect by substring
        candidates = [s for s in wb.sheetnames if "pivot" in s.lower()]
        if len(candidates) == 1:
            pivot_sheet_name = candidates[0]
        else:
            wb.close()
            raise FileNotFoundError(
                f"Pivot sheet not found in {backup_path}. Candidates: {candidates}"
            )
    ws = wb[pivot_sheet_name]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    wb.close()
    # Columns: Date, Subsidiary, Department, Journal Entry Memo, Account, Line Memo, Debit
    data = []
    for r in rows:
        if all(v is None for v in r):
            continue
        data.append({
            "Date": r[0],
            "Subsidiary": r[1],
            "Department": r[2] if r[2] else "Aggregate",
            "Journal Entry Memo": r[3],
            "Account": r[4],
            "Memo": r[5],
            "Amount": float(r[6]) if r[6] is not None else 0.0,
        })
    return pd.DataFrame(data)


def _write_je_with_reclasses_tab(backup_path: str, final_rows: list, sheet_name: str):
    """Append (or replace) the 'JE with Reclasses' tab to the backup workbook."""
    wb = openpyxl.load_workbook(backup_path)
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(title=sheet_name)
    headers = ["Date", "Subsidiary", "Department", "Journal Entry Memo",
               "Account", "Line Memo", "Debit"]
    ws.append(headers)
    for r in final_rows:
        ws.append([r[h] for h in headers])
    wb.save(backup_path)


def apply_reclasses(folder_path: str) -> dict:
    """Step 2: read the Reclasses tab and produce the JE with Reclasses.

    Reads the backup workbook in `folder_path`, parses the Reclasses tab,
    applies the reclasses to the pivot rows, writes a 'JE with Reclasses'
    tab to the same workbook, and returns a summary.
    """
    folder_path = os.path.abspath(folder_path)
    je_date = parse_folder_date(folder_path)
    folder_name = os.path.basename(folder_path)
    je_memo = f"{folder_name} US PAYROLL"
    pivot_sheet_name = f"{folder_name}_US_Payroll_JE (pivot)"
    output_sheet_name = "JE with Reclasses"
    backup_path = os.path.join(folder_path, f"{folder_name} US Payroll Backup.xlsx")

    if not os.path.isfile(backup_path):
        raise FileNotFoundError(
            f"Backup workbook not found at {backup_path}. "
            "Run Step 1 (payroll_mapper.py) first."
        )

    parsed = _parse_reclasses_tab(backup_path)
    if not parsed["present"]:
        return {
            "status": "no_reclasses_tab",
            "message": "No 'Reclasses' tab found in the backup workbook.",
            "backup_path": backup_path,
        }

    pivot_df = _read_pivot_tab(backup_path, pivot_sheet_name)
    pivot_depts = sorted({d for d in pivot_df["Department"].tolist() if d and d != "Aggregate"})

    # Severance tax reclass
    sev_entries, sev_log, sev_unresolved = _build_severance_tax_reclass_entries(
        parsed["severance_rows"], pivot_depts
    )

    # Board member reclass
    jb_entries = []
    jb_warnings = []
    if parsed["borland_components"]:
        jb_entries, jb_warnings = _build_boardmember_reclass_entries(
            parsed["borland_components"], pivot_df
        )

    # Merge: pivot rows + reclass entries → groupby → drop zeros
    pivot_rows_for_merge = pivot_df.copy()
    # Use "Aggregate" for the empty-dept rows so they groupby cleanly
    reclass_df = pd.DataFrame(sev_entries + jb_entries) if (sev_entries or jb_entries) \
        else pd.DataFrame(columns=["Department", "Account", "Memo", "Subsidiary", "Amount"])

    combined = pd.concat([
        pivot_rows_for_merge[["Department", "Account", "Memo", "Subsidiary", "Amount"]],
        reclass_df,
    ], ignore_index=True)
    grouped = combined.groupby(
        ["Department", "Account", "Memo", "Subsidiary"], as_index=False
    )["Amount"].sum()
    grouped["Amount"] = grouped["Amount"].round(2)
    grouped = grouped[grouped["Amount"] != 0.0]
    grouped = grouped.sort_values(by=["Department", "Account", "Memo"]).reset_index(drop=True)

    # Build final rows in JE format
    final_rows = []
    for _, r in grouped.iterrows():
        display_dept = "" if r["Department"] == "Aggregate" else r["Department"]
        final_rows.append({
            "Date": je_date,
            "Subsidiary": r["Subsidiary"],
            "Department": display_dept,
            "Journal Entry Memo": je_memo,
            "Account": r["Account"],
            "Line Memo": r["Memo"],
            "Debit": float(r["Amount"]),
        })

    _write_je_with_reclasses_tab(backup_path, final_rows, output_sheet_name)

    debits = sum(r["Debit"] for r in final_rows if r["Debit"] > 0)
    credits = -sum(r["Debit"] for r in final_rows if r["Debit"] < 0)
    delta = round(debits - credits, 2)

    return {
        "status": "ok",
        "backup_path": backup_path,
        "sheet_name": output_sheet_name,
        "final_rows": final_rows,
        "pivot_rows_count": len(pivot_df),
        "final_rows_count": len(final_rows),
        "totals": {
            "debits": round(debits, 2),
            "credits": round(credits, 2),
            "balanced": abs(delta) < 0.01,
            "delta": delta,
        },
        "reclasses": {
            "severance_resolution_log": sev_log,
            "severance_unresolved": sev_unresolved,
            "severance_entries_count": len(sev_entries),
            "borland_components": parsed["borland_components"],
            "borland_entries_count": len(jb_entries),
            "unknown_sections": parsed["unknown_sections"],
            "jb_warnings": jb_warnings,
        },
        "raw_reclasses_dump": parsed["raw_dump"],
    }


def main():
    args = [a for a in sys.argv[1:] if not a.startswith("--")]
    flags = [a for a in sys.argv[1:] if a.startswith("--")]

    if not args:
        print(__doc__)
        sys.exit(2)

    folder = args[0]
    verify = "--verify-suffix" in flags

    if "--reclasses" in flags:
        print(f"Applying reclasses to: {folder}")
        result = apply_reclasses(folder)
        if result["status"] == "no_reclasses_tab":
            print(result["message"])
            sys.exit(0)
        print(f"\nBackup:       {result['backup_path']}")
        print(f"Pivot rows:   {result['pivot_rows_count']}")
        print(f"Final rows:   {result['final_rows_count']}")
        t = result["totals"]
        print(f"Debits:       ${t['debits']:,.2f}")
        print(f"Credits:      ${t['credits']:,.2f}")
        if t["balanced"]:
            print("Balanced:     YES")
        else:
            print(f"Balanced:     NO (delta ${t['delta']:,.2f})")
        rc = result["reclasses"]
        print(f"\nSeverance reclass entries: {rc['severance_entries_count']}")
        print(f"Borland reclass entries:   {rc['borland_entries_count']}")
        if rc["unknown_sections"]:
            print("\n!! UNKNOWN SECTIONS (halt and ask the accountant):")
            for s in rc["unknown_sections"]:
                print(f"  - {s}")
        if rc["severance_unresolved"]:
            print("\n!! UNRESOLVED SEVERANCE ROWS:")
            for r in rc["severance_unresolved"]:
                print(f"  - {r}")
        if rc["jb_warnings"]:
            print("\n!! JB RECLASS WARNINGS:")
            for w in rc["jb_warnings"]:
                print(f"  - {w}")
        sys.exit(0)

    print(f"Processing pay-run folder: {folder}")
    result = generate_je(folder, verify=verify)

    print(f"\nRaw input:     {os.path.basename(result['raw_input'])}")
    print(f"JE date:       {result['je_date'].strftime('%Y-%m-%d')}")
    print(f"JE memo:       {result['je_memo']}")
    print(f"Pivot rows:    {len(result['rows'])}")
    totals = result["totals"]
    print(f"Total debits:  ${totals['debits']:,.2f}")
    print(f"Total credits: ${totals['credits']:,.2f}")
    if totals["balanced"]:
        print("Balanced:      YES")
    else:
        print(f"Balanced:      NO  (delta ${totals['delta']:,.2f})")
    print(f"\nOutput:        {result['output_path']}")

    if result["unmapped_cost_centers"]:
        print("\n--- UNMAPPED COST CENTERS ---")
        for cc in result["unmapped_cost_centers"]:
            print(f"  {cc}")
    if result["unmapped_codes"]:
        print("\n--- UNMAPPED PAYROLL CODES ---")
        for code in result["unmapped_codes"]:
            print(f"  {code}")


if __name__ == "__main__":
    main()
