"""Canonical paste-back generator for bank-statement-posting skill.

Reads accounts.yaml + a month folder of bank-statement CSVs, parses each via the
account's csv_format, applies an injected classification map (from the SKILL.md
classifier table + Phase 4 PYMT/Check tranids), and writes a multi-tab XLSX
with one tab per account.

Tab columns match the May-26 Banking Sheet:
  Date | Description | Amount | Tag | Owner | Week | Posted? | Ref #

Field rules (DO NOT deviate):
- Date = M/D/YYYY (no leading zeros)
- Description = Customer Conference verbatim Transaction Detail (Chase) / Reference Detail (Wells Fargo) /
  equivalent for foreign banks. Never edited.
- Amount = signed (negative for outflows). No symbols.
- Tag = bare values only (N/A, BEN, DOM, BREX, EXP, BILL, LOCINT, LOANINT, INT, LOC,
  LOAN, TAX, MISC, PAYUS, PAYCA, PAYUK, PAYUY, PAYBV, PAYDE, PAYPN). Never with suffix.
- Owner = the accountant / the AP Specialist.
- Week = WK1..WK5 (formula: WK{((day-1)//7)+1}).
- Posted? = "*" or blank.
- Ref # = PYMT####, JE####, ACH_Debit_MM.DD.YY_<bank>, or blank.

Usage (from skill driver):
    from generate_paste_back import build_workbook
    build_workbook(
        month_folder = "Weekly Cash Activities/2026-05",
        classifier_map = {
            ("chase_x0001", "5/1/2026", 50000.00): ("N/A", "the accountant", "*", "PYMT####"),
            ...
        },
        accounts_yaml = ".claude/skills/bank-statement-posting/accounts.yaml",
        output_filename = "banking-transactions-paste-20260505.xlsx",  # optional
    )
"""
from __future__ import annotations

import csv
import glob
import os
from datetime import datetime
from typing import Iterator

try:
    import yaml
except ImportError:
    yaml = None

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

HEADERS = ["Date", "Description", "Amount", "Tag", "Owner", "Week", "Posted?", "Ref #"]


# ---------------------------------------------------------------------------
# Per-format CSV parsers. Each yields (post_date_str, description, signed_amount, ref_id).
# Description is the bank's transaction-detail field, whitespace-normalized (runs of
# whitespace collapsed to a single space). Chase occasionally pads descriptions with
# extra spaces between Pending (I) and Posted (F) pulls — without normalization, repeat
# pastes into Google Sheets create whitespace-only duplicates that look identical to
# the eye but are distinct cell values. Per the accountant 2026-05-15.
# ref_id is the bank's stable transaction identifier - used for cross-pull dedup so
# the same transaction in Pending and Posted states (which sometimes have differing
# description text) is recognized as one row, not two. Critical for Wells Fargo.
# ---------------------------------------------------------------------------

def _normalize_desc(s: str) -> str:
    """Collapse runs of whitespace to a single space (and trim ends)."""
    import re
    return re.sub(r"\s+", " ", (s or "").strip())


def parse_chase_standard(path: str) -> Iterator[tuple[str, str, float, str]]:
    """Chase: Account Number,...,Post Date,Status,Transaction Description,Amount,
    Bank Reference,Customer Reference,Transaction Detail."""
    with open(path, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            try:
                amount = float(row["Amount"])
            except (KeyError, ValueError):
                continue
            ref_id = (row.get("Bank Reference") or "").strip()
            yield row["Post Date"], _normalize_desc(row["Transaction Detail"]), amount, ref_id


def parse_wellsfargo_activity_detail(path: str) -> Iterator[tuple[str, str, float, str]]:
    """Wells Fargo: Account,...,Date,Status,Type,Reference Number,Withdrawals,Deposits,...,
    Reference Detail. Withdrawals -> negative, Deposits -> positive.
    Wells Fargo's Reference Detail text changes between Pending and Posted statuses (UETR field
    added, RPTID changes, etc.), so we dedup on Reference Number (the column 7 ID like
    "265844232KZY00Z7" or "26120008066416") which is stable across pulls."""
    with open(path, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            wd = (row.get("Withdrawals") or "").strip()
            dp = (row.get("Deposits") or "").strip()
            if wd:
                amount = -float(wd.replace(",", ""))
            elif dp:
                amount = float(dp.replace(",", ""))
            else:
                continue
            ref_id = (row.get("Reference Number") or "").strip()
            yield row["Date"], _normalize_desc(row["Reference Detail"]), amount, ref_id


def parse_td_canada(path: str) -> Iterator[tuple[str, str, float, str]]:
    """TD Canada: Posted Date,Value Date,Company Name,Account Name,Account Nickname,
    Account Number,Transit Number,Description 1..5,Currency,Withdrawals,Deposits,Balance.

    Date is YYYYMMDD (zero-padded, no separators) -> normalize to M/D/YYYY downstream.
    Description = " | "-joined non-empty Description 1-5 fields, whitespace-normalized.
    Withdrawals -> negative, Deposits -> positive.

    TD has no stable per-row bank reference, so we fall back to the desc as the dedup key.
    Multiple TD pulls in one folder are rare since each export covers a wide date range.
    """
    with open(path, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            wd = (row.get("Withdrawals") or "").strip()
            dp = (row.get("Deposits") or "").strip()
            if wd:
                amount = -float(wd.replace(",", ""))
            elif dp:
                amount = float(dp.replace(",", ""))
            else:
                continue
            desc_parts = [
                (row.get(f"Description {i}") or "").strip()
                for i in range(1, 6)
            ]
            desc = _normalize_desc(" | ".join(p for p in desc_parts if p))
            # Convert YYYYMMDD -> MM/DD/YYYY for downstream normalize_date()
            raw_date = row["Posted Date"].strip()
            if len(raw_date) == 8 and raw_date.isdigit():
                raw_date = f"{raw_date[4:6]}/{raw_date[6:8]}/{raw_date[0:4]}"
            yield raw_date, desc, amount, desc


def parse_ing_bv_csv(path: str) -> Iterator[tuple[str, str, float, str]]:
    """Chase Bank (Europe) N.V. — tab-separated CSV, handles all 3 currencies
    (EUR / USD / GBP). EUR uses an 18-col schema with separate Debit/Credit columns;
    USD/GBP use a 15-col schema with a single 'Debit Credit' text column and a
    positive-magnitude 'Amount' column. Both detected by header sniff.

    Skip rules:
    - Record Type != 'Posted' (drops Opening/Closing Balance rows).
    - Transaction Type Name in {Opening Balance, Closing Balance}.

    Date column: Booking Date (format YYYY/MM/DD) -> normalize to MM/DD/YYYY downstream.
    Description column: Detail Information, whitespace-normalized.
    Dedup ref: Bank Reference (e.g., DD20002841 for fees, SEI-5761458 / 255771858CEH4O8B
    for SEPA/Inward). Critical: every customer inbound is TWO rows sharing one Bank
    Reference (FX fee debit + transfer credit). Both rows are intentional - the
    `(date, ref_id, signed_amount)` dedup key keeps them distinct because the amounts
    differ. Don't merge them.

    Per the accountant 2026-05-19 historical analysis (see Weekly Cash Activities/Historical
    Exampels/_analysis_ing_bv_bofa_pln.md).
    """
    with open(path, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f, delimiter="\t")
        header = reader.fieldnames or []
        is_eur_schema = "IBAN" in header and "Debit" in header and "Credit" in header
        for row in reader:
            if (row.get("Record Type") or "").strip().strip('"') != "Posted":
                continue
            ttype = (row.get("Transaction Type Name") or "").strip().strip('"')
            if ttype in ("Opening Balance", "Closing Balance"):
                continue
            raw_date = (row.get("Booking Date") or "").strip().strip('"')
            if not raw_date:
                continue
            # YYYY/MM/DD -> MM/DD/YYYY (downstream normalize_date strips zeros)
            parts = raw_date.split("/")
            if len(parts) != 3:
                continue
            yyyy, mm, dd = parts
            date_mdy = f"{int(mm):02d}/{int(dd):02d}/{yyyy}"
            if is_eur_schema:
                dr_s = (row.get("Debit") or "").strip().strip('"')
                cr_s = (row.get("Credit") or "").strip().strip('"')
                try:
                    dr_f = float(dr_s.replace(",", "")) if dr_s else 0.0
                    cr_f = float(cr_s.replace(",", "")) if cr_s else 0.0
                except ValueError:
                    continue
                amount = cr_f - dr_f
            else:
                dc = (row.get("Debit Credit") or "").strip().strip('"')
                amt_s = (row.get("Amount") or "").strip().strip('"')
                try:
                    amt_f = float(amt_s.replace(",", "")) if amt_s else 0.0
                except ValueError:
                    continue
                amount = amt_f if dc == "Credit" else -amt_f
            if amount == 0:
                continue
            desc = _normalize_desc(row.get("Detail Information") or "")
            ref_id = (row.get("Bank Reference") or "").strip().strip('"')
            yield date_mdy, desc, amount, ref_id


def parse_bofa_pln_xlsx(path: str) -> Iterator[tuple[str, str, float, str]]:
    """Bank of America Handlowy Poland — single-sheet XLSX, sheet name = account IBAN suffix.
    11 columns: Value Date | Statement Date | Currency | Amount | Beneficiary/ Remitter
    | Customer Reference | Type | Bank Reference | Description | Narrative | Payment Details.

    Sign: Amount is already-signed (negative = outflow). All YTD rows are outflows;
    inflows would be customer payments or internal funding wires (none observed Jan-May
    2026 in YTD export).

    Date column: Value Date in M/D/YYYY format (string). Pass through normalize_date
    downstream to strip leading zeros.

    Description: Description field, with Narrative appended via " | " if non-empty.
    Beneficiary name (from "Beneficiary/ Remitter" column) is NOT included in desc
    because the accountant's paste-back format puts the verbatim bank narrative in the desc
    column and the beneficiary name shows up in the classifier output. Originator
    lookup happens against the Beneficiary field separately.

    Dedup ref: Bank Reference (numeric strings like 0002021110); fall back to
    Customer Reference (PL0PRO326002A8KM pattern) when Bank Reference is blank.
    """
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl not installed - required for Bank of America PLN parser") from exc
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows_iter = ws.iter_rows(values_only=True)
    header = next(rows_iter, None)
    if not header:
        return
    # Map header to column index (Bank of America's header rows can drift across exports)
    idx = {(h or "").strip(): i for i, h in enumerate(header)}
    col_date = idx.get("Value Date")
    col_amt = idx.get("Amount")
    col_desc = idx.get("Description")
    col_narr = idx.get("Narrative")
    col_ref_bank = idx.get("Bank Reference")
    col_ref_cust = idx.get("Customer Reference")
    if col_date is None or col_amt is None:
        return
    for row in rows_iter:
        if not row or row[col_date] is None:
            continue
        # Date is sometimes string "M/D/YYYY" (Bank of America export) and sometimes a
        # datetime in older xlsx slices - normalize both to MM/DD/YYYY string.
        raw = row[col_date]
        if hasattr(raw, "strftime"):
            date_mdy = raw.strftime("%m/%d/%Y")
        else:
            s = str(raw).strip()
            parts = s.split("/")
            if len(parts) != 3:
                continue
            mm, dd, yyyy = parts
            date_mdy = f"{int(mm):02d}/{int(dd):02d}/{yyyy}"
        try:
            amt_raw = row[col_amt]
            if amt_raw is None:
                continue
            # Bank of America mixes raw float (-100) and comma-formatted strings ('-3,942.15')
            if isinstance(amt_raw, str):
                amt_raw = amt_raw.replace(",", "").strip()
                if not amt_raw:
                    continue
            amount = float(amt_raw)
        except (TypeError, ValueError):
            continue
        if amount == 0:
            continue
        desc_main = row[col_desc] if col_desc is not None else ""
        desc_narr = row[col_narr] if col_narr is not None else ""
        parts = [str(p).strip() for p in (desc_main, desc_narr) if p]
        desc = _normalize_desc(" | ".join(parts))
        ref_id = ""
        if col_ref_bank is not None and row[col_ref_bank]:
            ref_id = str(row[col_ref_bank]).strip()
        elif col_ref_cust is not None and row[col_ref_cust]:
            ref_id = str(row[col_ref_cust]).strip()
        yield date_mdy, desc, amount, ref_id


def parse_barclays_uk(path: str) -> Iterator[tuple[str, str, float, str]]:
    """Barclays UK Bank PLC (Acme UK Ltd, sub 8, GBP) — wide single-sheet XLSX export.
    25 columns; the ones we use: 'Credit amount', 'Debit amount' (separate signed cols),
    'Post date' (DD/MM/YYYY), 'Narrative', 'Bank reference', 'Customer reference', 'TRN type'.
    Yields (date_mdy, desc, signed_amount, ref_id). Credit -> positive, Debit -> negative.

    Recurring patterns (per Mar/Apr-26 history): AVIVA HEALTH (BEN), ROYAL LON-SCOTLIFE
    pension (BEN), AIRWALLEX = Brex card pmts (EXP), Barclays monthly CHGS (INT), and the
    GBP 50,000 'Transfer in - from Chase US' funding leg (N/A - TRN, booked as JE).
    """
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl required for Barclays UK parser") from exc
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows_iter = ws.iter_rows(values_only=True)
    header = next(rows_iter, None)
    if not header:
        return
    idx = {(h or "").strip(): i for i, h in enumerate(header)}
    c_date = idx.get("Post date")
    c_cr = idx.get("Credit amount")
    c_dr = idx.get("Debit amount")
    c_narr = idx.get("Narrative")
    c_bref = idx.get("Bank reference")
    c_cref = idx.get("Customer reference")
    if c_date is None or (c_cr is None and c_dr is None):
        return
    for row in rows_iter:
        if not row or row[c_date] is None:
            continue
        raw = row[c_date]
        if hasattr(raw, "strftime"):
            date_mdy = raw.strftime("%m/%d/%Y")
        else:
            s = str(raw).strip()
            parts = s.split("/")
            if len(parts) != 3:
                continue
            dd, mm, yyyy = parts  # Barclays is DD/MM/YYYY
            date_mdy = f"{int(mm):02d}/{int(dd):02d}/{yyyy}"
        cr = row[c_cr] if c_cr is not None else None
        dr = row[c_dr] if c_dr is not None else None
        def _num(v):
            if v is None:
                return 0.0
            if isinstance(v, str):
                v = v.replace(",", "").strip()
                if not v:
                    return 0.0
            return float(v)
        try:
            amount = _num(cr) - abs(_num(dr))
        except (TypeError, ValueError):
            continue
        if amount == 0:
            continue
        desc = _normalize_desc(row[c_narr] if c_narr is not None else "")
        ref_id = ""
        if c_bref is not None and row[c_bref]:
            ref_id = str(row[c_bref]).strip()
        elif c_cref is not None and row[c_cref]:
            ref_id = str(row[c_cref]).strip()
        yield date_mdy, desc, amount, ref_id


PARSERS = {
    "chase_standard": parse_chase_standard,
    "wellsfargo_activity_detail": parse_wellsfargo_activity_detail,
    "td_canada": parse_td_canada,
    "ing_bv": parse_ing_bv_csv,
    "bofa_pln": parse_bofa_pln_xlsx,
    "barclays_uk": parse_barclays_uk,
}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def normalize_date(date_str: str) -> str:
    """MM/DD/YYYY -> M/D/YYYY (strip leading zeros)."""
    dt = datetime.strptime(date_str, "%m/%d/%Y")
    return f"{dt.month}/{dt.day}/{dt.year}"


def week_label(date_str_mdy: str) -> str:
    """WK1..WK5 — business-day weeks anchored on Mondays (per the accountant 2026-05-15).
    Algorithm: count Mondays strictly after day 1 of the month through dt.day;
    WK = count + 1. For May 2026 (Fri 5/1 start): 5/1=WK1, 5/4-5/8=WK2,
    5/11-5/15=WK3, 5/18-5/22=WK4, 5/25-5/29=WK5. For June 2026 (Mon 6/1):
    6/1 alone=WK1, 6/2-6/8=WK2."""
    dt = datetime.strptime(date_str_mdy, "%m/%d/%Y")
    mondays = sum(
        1 for d in range(2, dt.day + 1)
        if datetime(dt.year, dt.month, d).weekday() == 0
    )
    return f"WK{mondays + 1}"


def in_target_month(date_str_mdy: str, target_yyyy_mm: str) -> bool:
    """Return True if M/D/YYYY string falls in target_yyyy_mm (YYYY-MM)."""
    dt = datetime.strptime(date_str_mdy, "%m/%d/%Y")
    return f"{dt.year}-{dt.month:02d}" == target_yyyy_mm


def load_accounts_yaml(path: str) -> dict:
    if yaml is None:
        raise RuntimeError("PyYAML not installed. Run: pip install pyyaml")
    with open(path, encoding="utf-8") as f:
        return yaml.safe_load(f)


def detect_account_for_csv(csv_path: str, accounts: dict) -> str | None:
    """Match a CSV/XLSX file to an account_key. Strategy:
    1. For XLSX (Bank of America PLN): match by filename glob only (sheet sniff is expensive).
    2. For CSV: sniff first data row's account column against bank_account_string.
       Handles both comma-separated (Chase x0001/x0002, Wells Fargo) and tab-separated
       (ING BV) formats. ING BV uses one IBAN suffix (0635637510) across EUR/USD/GBP,
       so we additionally match the Currency column to disambiguate.
    3. Fallback to filename glob for everything.
    """
    ext = os.path.splitext(csv_path)[1].lower()
    if ext == ".xlsx":
        # Match by filename glob only
        fname = os.path.basename(csv_path)
        for k, cfg in accounts.items():
            if k.startswith("banking_sheet"):
                continue
            pat = cfg.get("filename_glob")
            if pat and glob.fnmatch.fnmatch(fname, pat):
                return k
        return None

    # CSV: sniff first line to detect delimiter (tab for ING BV, comma otherwise)
    with open(csv_path, newline="", encoding="utf-8-sig") as f:
        sniff = f.readline()
        f.seek(0)
        delim = "\t" if "\t" in sniff and sniff.count("\t") > sniff.count(",") else ","
        reader = csv.DictReader(f, delimiter=delim)
        first = next(iter(reader), None)

    if first is not None:
        # Chase x0001/x0002 header: "Account Number"; Wells Fargo header: "Account"; ING BV: "Account No."
        # TD: also has "Account Number" - same key as Chase so the bank_account_string disambiguates.
        for key in ("Account Number", "Account No.", "Account"):
            if key in first:
                acct_str = (first.get(key) or "").strip().strip('"')
                # ING BV shares one Account No. across EUR/USD/GBP - disambiguate by Currency
                ccy = (first.get("Currency") or "").strip().strip('"')
                for k, cfg in accounts.items():
                    if k.startswith("banking_sheet"):
                        continue
                    if cfg.get("bank_account_string") != acct_str:
                        continue
                    cfg_ccy = cfg.get("currency")
                    if cfg_ccy and ccy and cfg_ccy != ccy:
                        continue
                    return k

    # Fallback: filename glob
    fname = os.path.basename(csv_path)
    for k, cfg in accounts.items():
        if k.startswith("banking_sheet"):
            continue
        pat = cfg.get("filename_glob")
        if pat and glob.fnmatch.fnmatch(fname, pat):
            return k
    return None


# ---------------------------------------------------------------------------
# Workbook builder
# ---------------------------------------------------------------------------

def build_tab(ws, rows: list[tuple]) -> None:
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            if col_idx == 3 and isinstance(val, (int, float)):
                cell.number_format = '#,##0.00;[Red](#,##0.00)'
            if col_idx == 2:
                cell.alignment = Alignment(wrap_text=False, vertical="top")
    widths = [11, 110, 14, 8, 10, 7, 9, 28]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w


def build_workbook(
    month_folder: str,
    classifier_map: dict[tuple[str, str, float], tuple[str, str, str, str]],
    accounts_yaml: str,
    output_filename: str | None = None,
) -> str:
    """Generate the paste-back XLSX.

    classifier_map keys: (account_key, M/D/YYYY date, signed amount rounded to 2dp).
    Values: (tag, owner, posted, ref_num).

    Returns the absolute output path.
    """
    accounts = load_accounts_yaml(accounts_yaml)

    # Determine target month from folder name (last path component "2026-05")
    target_yyyy_mm = os.path.basename(os.path.normpath(month_folder))
    if not output_filename:
        today = datetime.now().strftime("%Y%m%d")
        output_filename = f"banking-transactions-paste-{today}.xlsx"
    output_path = os.path.join(month_folder, output_filename)

    # Group input files by account. Recursive glob picks up per-account subfolders
    # (e.g. "Weekly Cash Activities/2026-05/ING BV EUR/*.csv") plus root drops.
    csvs_by_account: dict[str, list[str]] = {}
    input_paths = (
        glob.glob(os.path.join(month_folder, "*.csv"))
        + glob.glob(os.path.join(month_folder, "*.xlsx"))
        + glob.glob(os.path.join(month_folder, "*", "*.csv"))
        + glob.glob(os.path.join(month_folder, "*", "*.xlsx"))
    )
    for csv_path in input_paths:
        # Skip our own paste-back outputs
        fname = os.path.basename(csv_path)
        if fname.startswith("banking-transactions-paste-") or fname.startswith("_dryrun"):
            continue
        account_key = detect_account_for_csv(csv_path, accounts)
        if account_key is None:
            print(f"WARN unmatched file (skipped): {csv_path}")
            continue
        if accounts[account_key].get("status") != "production":
            print(f"WARN {account_key} is {accounts[account_key].get('status')}; skipping {csv_path}")
            continue
        csvs_by_account.setdefault(account_key, []).append(csv_path)

    # Build the workbook with tabs in accounts.yaml order
    wb = Workbook()
    wb.remove(wb.active)
    for account_key, cfg in accounts.items():
        if account_key.startswith("banking_sheet"):
            continue
        ws = wb.create_sheet(cfg.get("sheet_tab", account_key))
        rows: list[tuple] = []
        parser = PARSERS.get(cfg.get("csv_format"))
        # Cross-pull dedup: Wells Fargo's Reference Detail text changes between Pending and Posted
        # for the same transaction; dedup on the bank's stable ref_id to avoid duplicates.
        seen_refs: set = set()
        for csv_path in csvs_by_account.get(account_key, []):
            if parser is None:
                print(f"WARN no parser for {cfg.get('csv_format')} ({account_key})")
                continue
            for raw_date, desc, amount, ref_id in parser(csv_path):
                date_norm = normalize_date(raw_date)
                if not in_target_month(date_norm, target_yyyy_mm):
                    continue
                dedup_key = (date_norm, ref_id, round(amount, 2))
                if dedup_key in seen_refs:
                    continue
                seen_refs.add(dedup_key)
                cls = classifier_map.get((account_key, date_norm, round(amount, 2)))
                if cls is None:
                    tag, owner, posted, ref = ("MISC", "the accountant", "", "")
                    print(f"WARN no classifier: {account_key} {date_norm} {amount}: {desc[:60]}")
                else:
                    tag, owner, posted, ref = cls
                rows.append((date_norm, desc, amount, tag, owner, week_label(date_norm), posted, ref))
        rows.sort(key=lambda r: datetime.strptime(r[0], "%m/%d/%Y"))
        build_tab(ws, rows)

    wb.save(output_path)
    return output_path


if __name__ == "__main__":
    # CLI smoke test against today's May 2026 folder
    import sys
    here = os.path.dirname(os.path.abspath(__file__))
    repo_root = os.path.normpath(os.path.join(here, "..", "..", "..", ".."))
    accounts_yaml = os.path.normpath(os.path.join(here, "..", "accounts.yaml"))
    month_folder = sys.argv[1] if len(sys.argv) > 1 else os.path.join(
        repo_root, "Weekly Cash Activities", "2026-05"
    )
    print(f"Repo root:    {repo_root}")
    print(f"Accounts:     {accounts_yaml}")
    print(f"Month folder: {month_folder}")
    out = build_workbook(month_folder, classifier_map={}, accounts_yaml=accounts_yaml)
    print(f"Output:       {out}")
